import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import re
import time

# ==============================
# Configuración inicial
# ==============================

desktop_path = os.path.join(
    os.path.expanduser('~'),
    'Desktop/GASTOS RESICO/2026/ENERO26'
)

file_name = input("Ingrese el nombre del archivo Excel (sin extensión .xlsx): ")

file_name = file_name.strip()
file_name = re.sub(r'\.xlsx$', '', file_name, flags=re.IGNORECASE) + '.xlsx'
file_path = os.path.join(desktop_path, file_name)

# Iniciar cronómetro
tiempo_inicio = time.time()

# ==============================
# Estilos
# ==============================

blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
purple_fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

result_style = Font(bold=True)
center_align = Alignment(horizontal='center')

# ==============================
# Funciones auxiliares
# ==============================

def extract_code(value):
    """Extrae el código de un valor que puede contener formato 'CODIGO - Descripción'"""
    if value and '-' in str(value):
        return str(value).split('-')[0].strip()
    return str(value).strip() if value else ''

def find_column(sheet, column_name):
    """Busca una columna por nombre en la primera fila"""
    for cell in sheet[1]:
        if cell.value and str(cell.value).strip().lower() == column_name.strip().lower():
            return cell.column
    return None

def create_column_if_missing(sheet, column_name, fill_color=blue_fill):
    """Crea una columna si no existe"""
    col = find_column(sheet, column_name)
    if col is None:
        last_col = sheet.max_column + 1
        sheet.cell(row=1, column=last_col, value=column_name)
        sheet.cell(row=1, column=last_col).fill = fill_color
        sheet.cell(row=1, column=last_col).alignment = center_align
        print(f"✅ Columna creada: {column_name}")
        return last_col
    return col

def es_gasolina(concepto):
    """
    Detecta si un concepto es relacionado a combustible/gasolina.
    Según Art. 27 LISR: combustibles para vehículos marítimos, aéreos y terrestres
    """
    if not concepto:
        return False
    palabras = [
        'gasolina', 'combustible', 'magna', 'premium',
        'diesel', 'diésel', 'gasohol', 'gasoil',
        'nafta', 'petrol', 'gas', 'energético',
        'turbosina', 'jet fuel', 'bunker'
    ]
    concepto_lower = concepto.lower()
    return any(p in concepto_lower for p in palabras)

# ==============================
# Reglas de deducibilidad
# ==============================

USOS_DEDUCIBLES = ['G01', 'G02', 'G03']
METODOS_VALIDOS = ['PUE', 'PPD']
FORMAS_VALIDAS = ['01', '02', '03', '04', '28']
FORMAS_ELECTRONICAS = ['02', '03', '04', '28']
USO_CFDI_VERDE = "G02 - Devoluciones, descuentos o bonificaciones"
REGIMENES_FACILIDAD_COMBUSTIBLE = ['626']

# ==============================
# Abrir archivo Excel
# ==============================

try:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    print(f"✅ Archivo cargado: {file_path}")
    print(f"📊 Filas totales: {sheet.max_row - 1}")
except Exception as e:
    print(f"❌ Error al abrir el archivo: {e}")
    raise SystemExit(1)

# ==============================
# Columnas requeridas
# ==============================

required_columns = {
    'SubTotal': 'SubTotal',
    'Descuento': 'Descuento',
    'IVA Trasladado 0%': 'IVA Trasladado 0%',
    'IVA Exento': 'IVA Exento',
    'IVA Trasladado 16%': 'IVA Trasladado 16%',
    'Total': 'Total',
    'Uso CFDI': 'Uso CFDI',
    'Metodo pago': 'Metodo pago',
    'Forma pago': 'Forma pago',
    'Regimen receptor': 'Regimen receptor',
    'Razon emisor': 'Razon emisor',
    'Conceptos': 'Conceptos'
}

print("🔍 Buscando / creando columnas base...")
columns = {}

for key, col_name in required_columns.items():
    columns[key] = create_column_if_missing(sheet, col_name)

# ==============================
# OPTIMIZACIÓN #1: Pre-indexar columnas IEPS
# ==============================

print("🚀 Optimizando búsquedas de columnas...")

ieps_encontrado = False
ieps_no_desglosado_encontrado = False
columnas_ieps = []

for col in range(1, sheet.max_column + 1):
    header = sheet.cell(row=1, column=col).value
    if header and "IEPS" in str(header).upper():
        columnas_ieps.append(col)
        header_str = str(header).strip()
        if 'No Desglosado' in header_str:
            columns['IEPS Trasladado No Desglosado'] = col
            ieps_no_desglosado_encontrado = True
            print(f"✅ IEPS No Desglosado: {get_column_letter(col)}")
        else:
            columns['IEPS Trasladado'] = col
            ieps_encontrado = True
            print(f"✅ IEPS Trasladado: {get_column_letter(col)}")

print(f"⚡ {len(columnas_ieps)} columnas IEPS pre-indexadas")

# ==============================
# Columna Efecto (si existe)
# ==============================

efecto_col = find_column(sheet, 'Efecto')
if efecto_col:
    print(f"✅ Columna Efecto: {get_column_letter(efecto_col)}")

# ==============================
# Crear columna de Razón No Deducible
# ==============================

razon_no_ded_col = create_column_if_missing(sheet, 'Razón No Deducible', red_fill)

# ==============================
# Inicializar columnas numéricas
# ==============================

print("📝 Inicializando columnas numéricas...")

for row in range(2, sheet.max_row + 1):
    for col_key in ['IVA Trasladado 16%', 'IVA Trasladado 0%', 'IVA Exento', 'Descuento']:
        cell = sheet.cell(row=row, column=columns[col_key])
        if cell.value is None:
            cell.value = 0
        cell.number_format = "0.00"

# ==============================
# Crear columnas de cálculo
# ==============================

last_column = sheet.max_column

headers = [
    'SUB1-16%',
    'SUB0%',
    'SUB2-16%',
    'IVA ACREDITABLE 16%',
    'C IVA',
    'T2',
    'Comprobación T2',
    'Deducible'
]

print("🧮 Creando columnas de cálculo...")

for i, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=last_column + i, value=header)
    cell.fill = blue_fill
    cell.alignment = center_align

sub1_col = last_column + 1
sub0_col = last_column + 2
sub2_col = last_column + 3
iva_acred_col = last_column + 4
c_iva_col = last_column + 5
t2_col = last_column + 6
comprob_col = last_column + 7
deducible_col = last_column + 8

# ==============================
# OPTIMIZACIÓN #2: Cachear letras de columnas
# ==============================

print("📝 Cacheando referencias de columnas...")

col_letters = {
    'SubTotal': get_column_letter(columns['SubTotal']),
    'Descuento': get_column_letter(columns['Descuento']),
    'IVA16': get_column_letter(columns['IVA Trasladado 16%']),
    'IVA0': get_column_letter(columns['IVA Trasladado 0%']),
    'Total': get_column_letter(columns['Total']),
    'sub1': get_column_letter(sub1_col),
    'sub0': get_column_letter(sub0_col),
    'sub2': get_column_letter(sub2_col),
    'iva_acred': get_column_letter(iva_acred_col),
    't2': get_column_letter(t2_col),
}

# ==============================
# Variables de control
# ==============================

ieps_procesados = 0
ieps_no_desglosado_procesados = 0
gasolina_con_ieps = 0
gasolina_sin_ieps = 0
uso_s01_count = 0
efectivo_mayor_2000 = 0
gasolina_efectivo = 0
gasolina_electronico = 0

print("🔧 Iniciando procesamiento optimizado de filas...")
print("=" * 80)

# ==============================
# LOOP PRINCIPAL OPTIMIZADO
# ==============================

total_filas = sheet.max_row - 1

for row in range(2, sheet.max_row + 1):
    
    # Mostrar progreso cada 100 filas
    if (row - 1) % 100 == 0 or row == sheet.max_row:
        progreso = ((row - 1) / total_filas) * 100
        print(f"📊 Procesando: {row - 1}/{total_filas} facturas ({progreso:.1f}%)")
    
    # ==============================
    # OPTIMIZACIÓN #4: Leer todas las celdas necesarias UNA VEZ
    # ==============================
    
    row_data = {
        'concepto': str(sheet.cell(row=row, column=columns['Conceptos']).value or ''),
        'total': float(sheet.cell(row=row, column=columns['Total']).value or 0),
        'uso_cfdi': sheet.cell(row=row, column=columns['Uso CFDI']).value,
        'metodo_pago': sheet.cell(row=row, column=columns['Metodo pago']).value,
        'forma_pago': sheet.cell(row=row, column=columns['Forma pago']).value,
        'regimen': sheet.cell(row=row, column=columns['Regimen receptor']).value,
    }
    
    # Detección de gasolina (una sola vez)
    es_gasolina_concepto = es_gasolina(row_data['concepto'])
    
    # ==============================
    # Buscar IEPS (OPTIMIZADO)
    # ==============================
    
    tiene_ieps = False
    tiene_ieps_no_desglosado = False
    
    if columnas_ieps:  # Solo si hay columnas IEPS
        for col in columnas_ieps:  # Solo recorre 1-3 columnas
            ieps_val = sheet.cell(row=row, column=col).value
            if ieps_val not in (None, 0, ''):
                header = sheet.cell(row=1, column=col).value
                if 'No Desglosado' in str(header):
                    tiene_ieps_no_desglosado = True
                else:
                    tiene_ieps = True
    
    # Marcar gasolina
    if es_gasolina_concepto:
        cel = sheet.cell(row=row, column=columns['Conceptos'])
        if tiene_ieps or tiene_ieps_no_desglosado:
            cel.fill = blue_fill
            gasolina_con_ieps += 1
        else:
            cel.fill = orange_fill
            gasolina_sin_ieps += 1
    
    # ==============================
    # TODAS LAS FÓRMULAS (NO VALORES)
    # ==============================
    
    # SUB1-16% = SubTotal - Descuento
    sheet.cell(
        row=row,
        column=sub1_col,
        value=f"=({col_letters['SubTotal']}{row}-{col_letters['Descuento']}{row})"
    )
    sheet.cell(row=row, column=sub1_col).number_format = "0.00"
    sheet.cell(row=row, column=sub1_col).font = result_style
    
    # IEPS -> IVA 0% (SOLO GASOLINA)
    if es_gasolina_concepto and columnas_ieps:
        for col in columnas_ieps:
            ieps_val = sheet.cell(row=row, column=col).value
            if ieps_val not in (None, 0, ''):
                sheet.cell(row=row, column=columns['IVA Trasladado 0%'], value=ieps_val)
                sheet.cell(row=row, column=columns['IVA Trasladado 0%']).fill = orange_fill
                ieps_procesados += 1
                break
    
    # SUB0% = IVA 0%
    sheet.cell(
        row=row,
        column=sub0_col,
        value=f"={col_letters['IVA0']}{row}"
    )
    sheet.cell(row=row, column=sub0_col).number_format = "0.00"
    sheet.cell(row=row, column=sub0_col).font = result_style
    
    # ✅ SUB2-16% = SUB1 - SUB0 (SIEMPRE FÓRMULA)
    sheet.cell(
        row=row,
        column=sub2_col,
        value=f"={col_letters['sub1']}{row}-{col_letters['sub0']}{row}"
    )
    sheet.cell(row=row, column=sub2_col).number_format = "0.00"
    sheet.cell(row=row, column=sub2_col).font = result_style
    
    # ✅ IVA ACREDITABLE 16% = SUB2 * 0.16 (SIEMPRE FÓRMULA)
    sheet.cell(
        row=row,
        column=iva_acred_col,
        value=f"={col_letters['sub2']}{row}*0.16"
    )
    sheet.cell(row=row, column=iva_acred_col).number_format = "0.00"
    sheet.cell(row=row, column=iva_acred_col).font = result_style
    
    # Validación visual (calcular para comparar)
    # Solo para el formato de colores, NO para escribir el valor
    try:
        # Leer los valores calculados por Excel
        subtotal_val = float(sheet.cell(row=row, column=columns['SubTotal']).value or 0)
        descuento_val = float(sheet.cell(row=row, column=columns['Descuento']).value or 0)
        iva0_val = float(sheet.cell(row=row, column=columns['IVA Trasladado 0%']).value or 0)
        iva16_val = float(sheet.cell(row=row, column=columns['IVA Trasladado 16%']).value or 0)
        
        sub1_calc = subtotal_val - descuento_val
        sub2_calc = sub1_calc - iva0_val
        iva_acred_calc = round(sub2_calc * 0.16, 2)
        
        if abs(iva_acred_calc - iva16_val) < 0.01:
            sheet.cell(row=row, column=iva_acred_col).fill = green_fill
            sheet.cell(row=row, column=columns['IVA Trasladado 16%']).fill = green_fill
    except:
        pass  # Si hay error en el cálculo, no aplicar color
    
    # C IVA = IVA Acreditable - IVA 16%
    sheet.cell(
        row=row,
        column=c_iva_col,
        value=f"={col_letters['iva_acred']}{row}-{col_letters['IVA16']}{row}"
    )
    sheet.cell(row=row, column=c_iva_col).number_format = "0.00"
    sheet.cell(row=row, column=c_iva_col).font = result_style
    
    # T2 = SUB2 + SUB0 + IVA16
    sheet.cell(
        row=row,
        column=t2_col,
        value=f"={col_letters['sub2']}{row}+{col_letters['sub0']}{row}+{col_letters['IVA16']}{row}"
    )
    sheet.cell(row=row, column=t2_col).number_format = "0.00"
    sheet.cell(row=row, column=t2_col).font = result_style
    
    # Comprobación T2 = Total - T2
    sheet.cell(
        row=row,
        column=comprob_col,
        value=f"={col_letters['Total']}{row}-{col_letters['t2']}{row}"
    )
    sheet.cell(row=row, column=comprob_col).number_format = "0.00"
    sheet.cell(row=row, column=comprob_col).font = result_style
    
    # ==============================
    # Formateo y validaciones
    # ==============================
    
    regimen = extract_code(row_data['regimen'])
    
    if regimen == '626':
        sheet.cell(row=row, column=columns['Regimen receptor']).fill = blue_fill
    elif regimen == '612':
        sheet.cell(row=row, column=columns['Regimen receptor']).fill = purple_fill
    else:
        sheet.cell(row=row, column=columns['Regimen receptor']).fill = orange_fill
        sheet.cell(row=row, column=columns['Razon emisor']).fill = orange_fill
    
    uso_cfdi = extract_code(row_data['uso_cfdi']).upper() if row_data['uso_cfdi'] else ''
    
    if row_data['uso_cfdi'] and str(row_data['uso_cfdi']).strip() == USO_CFDI_VERDE:
        sheet.cell(row=row, column=columns['Uso CFDI']).fill = green_fill
    
    if uso_cfdi == 'S01':
        sheet.cell(row=row, column=columns['Uso CFDI']).fill = red_fill
        uso_s01_count += 1
    
    es_egreso = False
    if efecto_col:
        efecto_val = sheet.cell(row=row, column=efecto_col).value
        if efecto_val and str(efecto_val).strip().upper() in ['EGRESO', 'E']:
            es_egreso = True
    
    # ==============================
    # VALIDACIÓN DE DEDUCIBILIDAD
    # ==============================
    
    metodo_pg = extract_code(row_data['metodo_pago']).upper()
    forma_pg = extract_code(row_data['forma_pago']).upper()
    
    es_deducible = True
    razones_rechazo = []
    
    if uso_cfdi not in USOS_DEDUCIBLES:
        es_deducible = False
        razones_rechazo.append(f"Uso CFDI {uso_cfdi} no deducible")
    
    if metodo_pg not in METODOS_VALIDOS:
        es_deducible = False
        razones_rechazo.append(f"Método {metodo_pg} inválido")
    
    # VALIDACIÓN GASOLINA
    if es_gasolina_concepto:
        if forma_pg == '01':
            if regimen in REGIMENES_FACILIDAD_COMBUSTIBLE and row_data['total'] <= 2000:
                gasolina_efectivo += 1
                sheet.cell(row=row, column=columns['Forma pago']).fill = yellow_fill
                razones_rechazo.append("RESICO: Gasolina efectivo ≤$2,000 (facilidad)")
            else:
                es_deducible = False
                gasolina_efectivo += 1
                razones_rechazo.append("Gasolina NO deducible en efectivo")
                sheet.cell(row=row, column=columns['Forma pago']).fill = red_fill
        else:
            gasolina_electronico += 1
            if forma_pg not in FORMAS_ELECTRONICAS:
                es_deducible = False
                razones_rechazo.append(f"Gasolina: forma de pago {forma_pg} inválida")
    else:
        if forma_pg == '01' and row_data['total'] > 2000:
            es_deducible = False
            efectivo_mayor_2000 += 1
            razones_rechazo.append("Efectivo >$2,000 NO deducible")
            if efecto_col:
                sheet.cell(row=row, column=efecto_col).fill = red_fill
        elif forma_pg not in FORMAS_VALIDAS:
            es_deducible = False
            razones_rechazo.append(f"Forma de pago {forma_pg} inválida")
    
    # Escribir resultado
    ded_cell = sheet.cell(row=row, column=deducible_col, value="SI" if es_deducible else "NO")
    
    if es_deducible:
        ded_cell.fill = blue_fill if es_egreso else green_fill
    else:
        ded_cell.fill = red_fill
    
    ded_cell.font = Font(bold=True, color='FFFFFF')
    ded_cell.alignment = center_align
    
    razon_cell = sheet.cell(row=row, column=razon_no_ded_col)
    if razones_rechazo:
        razon_cell.value = " | ".join(razones_rechazo)
        if not es_deducible:
            razon_cell.fill = red_fill
            razon_cell.font = Font(bold=True, color='FFFFFF')
        else:
            razon_cell.fill = yellow_fill
            razon_cell.font = Font(bold=True)
    else:
        razon_cell.value = "Cumple requisitos"
        razon_cell.fill = green_fill
        razon_cell.font = Font(color='006100')

# ==============================
# Ajustar ancho de columnas
# ==============================

for col in range(last_column + 1, last_column + len(headers) + 1):
    sheet.column_dimensions[get_column_letter(col)].width = 15

sheet.column_dimensions[get_column_letter(razon_no_ded_col)].width = 40

# ==============================
# Guardar archivo
# ==============================

try:
    base_name = re.sub(r'\.xlsx$', '', file_name, flags=re.IGNORECASE)
    output_name = os.path.join(desktop_path, f"{base_name}_validado.xlsx")
    workbook.save(output_name)
    
    tiempo_fin = time.time()
    tiempo_total = tiempo_fin - tiempo_inicio
    velocidad = total_filas / tiempo_total if tiempo_total > 0 else 0

    print("\n" + "=" * 80)
    print("✅ PROCESO COMPLETADO EXITOSAMENTE")
    print("=" * 80)
    print(f"📂 Archivo guardado como: {output_name}")
    print(f"📊 Total de filas procesadas: {total_filas}")
    print(f"⏱️  Tiempo total: {tiempo_total:.2f} segundos")
    print(f"⚡ Velocidad: {velocidad:.0f} facturas/segundo")

    print("\n📌 RESUMEN DE VALIDACIÓN:")
    print(f"  • IEPS Trasladado procesados: {ieps_procesados}")
    print(f"  • IEPS No Desglosado procesados: {ieps_no_desglosado_procesados}")
    print(f"  • Gasolina con IEPS: {gasolina_con_ieps}")
    print(f"  • Gasolina sin IEPS: {gasolina_sin_ieps}")
    print(f"  • Gasolina pagada en efectivo: {gasolina_efectivo}")
    print(f"  • Gasolina pagada electrónicamente: {gasolina_electronico}")
    print(f"  • Usos S01 (sin efectos fiscales): {uso_s01_count}")
    print(f"  • Gastos normales efectivo >$2,000: {efectivo_mayor_2000}")

    print("\n⚠️  REGLAS APLICADAS:")
    print("  1. Art. 27, fracc. III LISR: Efectivo >$2,000 NO deducible")
    print("  2. Art. 27, fracc. III LISR: Gasolina NUNCA en efectivo")
    print("     (Excepción: RESICO hasta $2,000)")
    print("  3. Uso CFDI válido: G01, G02, G03")
    print("  4. Método de pago válido: PUE, PPD")
    print("  5. Forma de pago válida según tipo de gasto")

    print("=" * 80)

except Exception as e:
    print("\n❌ Error al guardar el archivo:")
    print(e)