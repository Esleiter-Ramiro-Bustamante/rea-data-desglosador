"""
motor_pandas.py — Motor pandas TURBO con vectorización completa
ReaDesF1.8

NUEVO EN v1.8:
  ✓ Reemplaza iterrows() por masks vectorizadas (10-50x más rápido)
  ✓ Columnas a tipo 'category' (hasta 70% menos RAM)
  ✓ Fórmulas auditables Excel escritas siempre por openpyxl
    (sub1, sub0, sub2, iva_acred — segunda validación contable)
"""

import zipfile
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from validaciones_fiscales import (
    evaluar_deducibilidad_vectorizado,
    optimizar_tipos_dataframe,
    formulas_auditables,
    PATRON_GASOLINA, PATRON_DULCE, PATRON_INSUMO,
    USO_CFDI_VERDE, REGIMENES_TRABAJADOS,
    LIMITE_EFECTIVO, FORMAS_ELECTRONICAS
)

blue_fill   = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
green_fill  = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill    = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
purple_fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
pink_fill   = PatternFill(start_color='FF69B4', end_color='FF69B4', fill_type='solid')
brown_fill  = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
lime_fill   = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

bold_white   = Font(bold=True, color='FFFFFF')
bold_black   = Font(bold=True)
center_align = Alignment(horizontal='center')
REG_FILLS    = {'626': blue_fill, '612': purple_fill, '616': brown_fill}


def _extraer_serie(s: pd.Series) -> pd.Series:
    s2   = s.fillna('').astype(str).str.strip()
    mask = s2.str.contains('-', regex=False)
    return s2.str.split('-').str[0].str.strip().where(mask, s2).str.upper()


def _parchear_cache_formulas(xlsx_path: str, cache: dict) -> None:
    """Inserta valores cacheados en fórmulas del xlsx vía XML."""
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        contents = {n: zin.read(n) for n in zin.namelist()}
    sheet_key = 'xl/worksheets/sheet1.xml'
    xml = contents[sheet_key].decode('utf-8')
    def replacer(m):
        ref = m.group(1)
        if ref not in cache: return m.group(0)
        val = cache[ref]
        sv = str(int(val)) if isinstance(val, float) and val == int(val) \
             else f'{val:.10f}'.rstrip('0').rstrip('.')
        return m.group(0).replace('<v></v>', f'<v>{sv}</v>')
    xml_patched = re.sub(
        r'<c r="([A-Z]+\d+)"[^>]*><f>[^<]*</f><v></v></c>',
        replacer, xml)
    contents[sheet_key] = xml_patched.encode('utf-8')
    with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in contents.items():
            zout.writestr(name, data)


def procesar_con_pandas(file_path: str, output_path: str,
                        modo: str = 'TURBO') -> dict:

    print(f"  🚀 Motor: pandas ({modo})")

    # ── PASO 1: Leer TODO de un golpe ────────────────────────────
    print("  📂 Cargando con pandas...")
    df = pd.read_excel(file_path, dtype=str)
    df.columns = df.columns.str.strip()
    total_filas = len(df)
    print(f"  ✅ {total_filas:,} filas cargadas")

    # ── PASO 2: Columnas a category (hasta 70% menos RAM) ─────────
    print("  🗜️  Optimizando tipos de columna (category)...")
    df = optimizar_tipos_dataframe(df)

    def gc(n):
        for c in df.columns:
            if c.strip().lower() == n.strip().lower(): return c
        return None
    def ac(n, d=None):
        c = gc(n)
        if c is None:
            df[n] = d if d is not None else ''
            return n
        return c

    C = {k: ac(v, '0' if k != 'concepto' else '') for k,v in {
        'concepto':'Conceptos','total':'Total','uso':'Uso CFDI',
        'metodo':'Metodo pago','forma':'Forma pago',
        'regimen':'Regimen receptor','subtotal':'SubTotal',
        'descuento':'Descuento','iva0':'IVA Trasladado 0%',
        'iva_ex':'IVA Exento','iva16':'IVA Trasladado 16%',
        'razon_em':'Razon emisor',
    }.items()}

    c_ieps8  = gc('IEPS Trasladado 8%')
    c_ieps_g = gc('IEPS Trasladado')
    c_ieps_nd= gc('IEPS Trasladado No Desglosado')

    # Convertir numéricos
    for nc in ['total','subtotal','descuento','iva0','iva_ex','iva16']:
        df[C[nc]] = pd.to_numeric(df[C[nc]], errors='coerce').fillna(0)
    for ic in [c_ieps8, c_ieps_g, c_ieps_nd]:
        if ic: df[ic] = pd.to_numeric(df[ic], errors='coerce').fillna(0)

    # ── PASO 3: Extraer códigos (vectorizado) ────────────────────
    print("  🔑 Extrayendo códigos...")
    df['_regimen'] = _extraer_serie(df[C['regimen']])
    df['_uso']     = _extraer_serie(df[C['uso']])
    df['_metodo']  = _extraer_serie(df[C['metodo']])
    df['_forma']   = _extraer_serie(df[C['forma']])
    df['_cl']      = df[C['concepto']].fillna('').str.lower()

    # ── PASO 4: Detección vectorizada ────────────────────────────
    print("  🔍 Detectando tipos (vectorizado — todas las filas a la vez)...")
    df['_es_gas']    = df['_cl'].str.contains(PATRON_GASOLINA, regex=True, na=False)
    df['_es_dulce']  = df['_cl'].str.contains(PATRON_DULCE,    regex=True, na=False)
    df['_es_insumo'] = df['_cl'].str.contains(PATRON_INSUMO,   regex=True, na=False)
    df['_agrupada']  = df['_es_gas'] & df['_cl'].str.contains(r'\|', na=False)

    # IEPS gasolina → mover a IVA 0%
    i8  = df[c_ieps8]   if c_ieps8  else pd.Series(0, index=df.index)
    ig  = df[c_ieps_g]  if c_ieps_g else pd.Series(0, index=df.index)
    ind = df[c_ieps_nd] if c_ieps_nd else pd.Series(0, index=df.index)
    ieps_gas = ig.where(ig > 0, ind)
    mask_ieps_gas = df['_es_gas'] & (ieps_gas > 0)
    df.loc[mask_ieps_gas, C['iva0']] = ieps_gas[mask_ieps_gas]

    # ── PASO 5: Evaluación vectorizada con MASKS ─────────────────
    # NUEVA en v1.8 — reemplaza iterrows() completamente
    print("  ⚖️  Evaluando deducibilidad (masks vectorizadas)...")
    df['Total'] = df[C['total']]   # alias para evaluar_deducibilidad_vectorizado
    df = evaluar_deducibilidad_vectorizado(df)
    print(f"  ✅ Deducibles : {(df['_deducible']=='SI').sum():,}")
    print(f"  ❌ No ded.    : {(df['_deducible']=='NO').sum():,}")

    # ── PASO 6: Agregar columnas de cálculo al DataFrame ─────────
    # NOTA: Los valores numéricos se guardan aquí para referencia.
    # Las FÓRMULAS AUDITABLES se escriben en el paso 7 vía openpyxl.
    st = df[C['subtotal']]
    dc = df[C['descuento']]
    i3 = df[fc('IEPS Trasladado 3%')].fillna(0).astype(float) if fc('IEPS Trasladado 3%') else pd.Series(0, index=df.index)
    df['_sub1']     = (st - dc + i8.where(i8 > 0, 0) + i3.where(i3 > 0, 0)).round(2)
    df.loc[mask_ieps_gas, '_sub1'] = (st - dc)[mask_ieps_gas].round(2)
    # sub0: gasolina con IEPS → solo iva0 (iva_ex es el mismo valor, no duplicar)
    df['_sub0'] = (df[C['iva0']] + df[C['iva_ex']]).round(2)
    df.loc[mask_ieps_gas, '_sub0'] = df.loc[mask_ieps_gas, C['iva0']].round(2)
    df['_sub2']     = (df['_sub1'] - df['_sub0']).round(2)
    df['_iva_acred']= (df['_sub2'] * 0.16).round(2)
    df['_iva_ok']   = (df['_iva_acred'] - df[C['iva16']]).abs() < 0.01

    # Columnas de salida
    df['Deducible']          = df['_deducible']
    df['Razón No Deducible'] = df['_razon']

    # ── CRÍTICO: agregar columnas de cálculo al df ANTES de to_excel() ──
    # Si no existen en el Excel, openpyxl no las encontrará con fc()
    # y wf() no escribirá ninguna fórmula auditable → celdas en 0.
    # Se escriben con el valor numérico calculado; openpyxl las
    # sobreescribirá con la fórmula Excel viva en el paso 7.
    df['SUB1-16%']            = df['_sub1']
    df['SUB0%']               = df['_sub0']
    df['SUB2-16%']            = df['_sub2']
    df['IVA ACREDITABLE 16%'] = df['_iva_acred']
    iva16_s = df[C['iva16']].fillna(0).astype(float)
    df['C IVA']               = (df['_iva_acred'] - iva16_s).round(2)
    df['T2']                  = (df['_sub2'] + df['_sub0'] + iva16_s).round(2)
    df['Comprobación T2']     = (df[C['total']].fillna(0).astype(float) - df['T2']).round(2)

    # Guardar con pandas (rápido)
    cols_internas = [c for c in df.columns if c.startswith('_')]
    df_salida = df.drop(columns=cols_internas)
    df_salida.to_excel(output_path, index=False)
    print(f"  ✅ Datos guardados")

    # ── PASO 7: openpyxl — FÓRMULAS AUDITABLES + COLORES ─────────
    # CRÍTICO: Las fórmulas sub1/sub0/sub2/iva_acred se escriben
    # siempre como fórmulas Excel vivas para auditoría contable.
    print("  🎨 Escribiendo fórmulas auditables y colores...")
    wb    = openpyxl.load_workbook(output_path)
    sheet = wb.active
    wb.calculation.fullCalcOnLoad = False

    hmap = {str(c.value).strip().lower(): c.column
            for c in sheet[1] if c.value}
    def fc(n): return hmap.get(n.strip().lower())

    # Columnas de cálculo — encabezados
    calc_names = ['SUB1-16%','SUB0%','SUB2-16%',
                  'IVA ACREDITABLE 16%','C IVA','T2',
                  'Comprobación T2']
    for n in calc_names:
        c = fc(n)
        if c:
            sh = sheet.cell(1, c)
            sh.fill = blue_fill
            sh.alignment = center_align
    c_ded   = fc('Deducible')
    c_razon = fc('Razón No Deducible')
    c_ded_h = sheet.cell(1, c_ded) if c_ded else None
    if c_ded_h: c_ded_h.fill = blue_fill; c_ded_h.alignment = center_align
    if c_razon: sheet.cell(1, c_razon).fill = red_fill

    # Columnas para fórmulas
    c_sub1 = fc('SUB1-16%')
    c_sub0 = fc('SUB0%')
    c_sub2 = fc('SUB2-16%')
    c_iva_a= fc('IVA ACREDITABLE 16%')
    c_civa = fc('C IVA')
    c_t2   = fc('T2')
    c_comp = fc('Comprobación T2')
    c_i16  = fc('IVA Trasladado 16%')
    c_uso  = fc('Uso CFDI')
    c_reg  = fc('Regimen receptor')
    c_rem  = fc('Razon emisor')
    c_conc = fc('Conceptos')
    c_form = fc('Forma pago')
    c_ef   = fc('Efecto')

    # Cache letras para fórmulas
    CL = {}
    for key, col_name in {
        'ST':'SubTotal','DC':'Descuento','I16':'IVA Trasladado 16%',
        'I0':'IVA Trasladado 0%','IE':'IVA Exento','TOT':'Total',
        'S1':'SUB1-16%','S0':'SUB0%','S2':'SUB2-16%',
        'IA':'IVA ACREDITABLE 16%','T2':'T2',
    }.items():
        c = fc(col_name)
        if c: CL[key] = get_column_letter(c)

    c_i8_col = fc('IEPS Trasladado 8%')
    if c_i8_col: CL['I8'] = get_column_letter(c_i8_col)
    c_i3_col = fc('IEPS Trasladado 3%')
    if c_i3_col: CL['I3'] = get_column_letter(c_i3_col)

    df['_ieps_gas_ok'] = mask_ieps_gas

    fmt = "0.00"
    _cache_formulas = {}

    def wf(col, formula, num_val):
        if col:
            c = sheet.cell(rn, col, formula)
            c.number_format = fmt
            c.font = Font(bold=True)
            _cache_formulas[f'{get_column_letter(col)}{rn}'] = num_val

    for idx in range(len(df)):
        rn      = idx + 2
        row_df  = df.iloc[idx]

        reg_val  = str(row_df.get('_regimen', ''))
        uso_val  = str(row_df.get('_uso', ''))
        forma_val= str(row_df.get('_forma', ''))
        es_gas_v = bool(row_df.get('_es_gas', False))
        es_ins_v = bool(row_df.get('_es_insumo', False))
        es_dul_v = bool(row_df.get('_es_dulce', False))
        agrup_v  = bool(row_df.get('_agrupada', False))
        iva_ok_v = bool(row_df.get('_iva_ok', False))
        ded_val  = str(row_df.get('_deducible', 'NO'))
        total_v  = float(row_df.get(C['total'], 0) or 0)
        i8_v     = float(row_df.get(c_ieps8, 0) or 0) if c_ieps8 else 0
        i3_v     = float(row_df.get(c_i3_col, 0) or 0) if c_i3_col else 0
        es_ded   = (ded_val == 'SI')

        # ══════════════════════════════════════════════════════════
        # FÓRMULAS AUDITABLES — escritas como fórmulas Excel vivas
        # sub1 = subtotal - descuento
        # sub0 = iva0 + iva_exento
        # sub2 = sub1 - sub0
        # iva_acred = sub2 * 0.16
        # Permiten ver la operación al pararse en la celda
        # ══════════════════════════════════════════════════════════
        fa = formulas_auditables(rn, CL)

        # sub0: gasolina con IEPS → sub0_gas (solo iva0, no duplicar iva_ex)
        ieps_g_v  = float(row_df.get(c_ieps_g,  0) or 0) if c_ieps_g  else 0
        ieps_nd_v = float(row_df.get(c_ieps_nd, 0) or 0) if c_ieps_nd else 0
        f_sub0 = fa['sub0_gas'] if (es_gas_v and (ieps_g_v > 0 or ieps_nd_v > 0)) else fa['sub0']

        # Valores numéricos ya calculados en el DataFrame
        v_sub1      = float(row_df.get('_sub1',      0) or 0)
        v_sub0      = float(row_df.get('_sub0',      0) or 0)
        v_sub2      = float(row_df.get('_sub2',      0) or 0)
        v_iva_acred = float(row_df.get('_iva_acred', 0) or 0)
        iva16_v     = float(row_df.get(C['iva16'],   0) or 0)
        v_c_iva     = round(v_iva_acred - iva16_v, 2)
        v_t2        = round(v_sub2 + v_sub0 + iva16_v, 2)
        v_comprob   = round(total_v - v_t2, 2)

        if i8_v > 0:
            f_sub1 = fa['sub1_ieps8']
        elif i3_v > 0:
            f_sub1 = fa['sub1_ieps3']
        else:
            f_sub1 = fa['sub1']
        wf(c_sub1, f_sub1,          v_sub1)
        wf(c_sub0, f_sub0,          v_sub0)
        wf(c_sub2, fa['sub2'],      v_sub2)
        wf(c_iva_a, fa['iva_acred'], v_iva_acred)
        wf(c_civa,  fa['c_iva'],     v_c_iva)
        wf(c_t2,    fa['t2'],        v_t2)
        wf(c_comp,  fa['comprob'],   v_comprob)

        # Validación visual IVA
        if c_i16 and c_iva_a and iva_ok_v:
            sheet.cell(rn, c_i16).fill  = green_fill
            sheet.cell(rn, c_iva_a).fill = green_fill

        # Color régimen
        if c_reg: sheet.cell(rn, c_reg).fill = REG_FILLS.get(reg_val, orange_fill)
        if c_rem and reg_val not in REG_FILLS: sheet.cell(rn, c_rem).fill = orange_fill

        # Uso CFDI
        if c_uso:
            uso_raw = str(df_salida.iloc[idx].get('Uso CFDI',''))
            if uso_raw.strip() == USO_CFDI_VERDE:
                sheet.cell(rn, c_uso).fill = green_fill
            if uso_val == 'S01':
                sheet.cell(rn, c_uso).fill = red_fill

        # Color concepto
        if c_conc:
            if es_dul_v and i8_v > 0:
                sheet.cell(rn, c_conc).fill = pink_fill
            elif es_gas_v:
                sheet.cell(rn, c_conc).fill = blue_fill if bool(row_df.get('_ieps_gas_ok', False)) else orange_fill
            elif es_ins_v:
                if forma_val == '01' and total_v > LIMITE_EFECTIVO:
                    sheet.cell(rn, c_conc).fill = red_fill
                elif forma_val == '01':
                    sheet.cell(rn, c_conc).fill = yellow_fill
                else:
                    sheet.cell(rn, c_conc).fill = lime_fill

        # Color forma pago
        if c_form:
            if es_gas_v and forma_val == '01':
                sheet.cell(rn, c_form).fill = (
                    yellow_fill if (reg_val == '626' and (total_v <= LIMITE_EFECTIVO or agrup_v))
                    else red_fill)
            elif es_ins_v and forma_val == '01' and total_v > LIMITE_EFECTIVO:
                sheet.cell(rn, c_form).fill = red_fill
            elif es_ins_v and forma_val == '01':
                sheet.cell(rn, c_form).fill = yellow_fill

        # Efecto
        if c_ef and not es_ded and forma_val == '01' and total_v > LIMITE_EFECTIVO:
            sheet.cell(rn, c_ef).fill = red_fill

        # Deducible
        if c_ded:
            es_egreso = False
            if c_ef:
                ev = sheet.cell(rn, c_ef).value
                if ev and str(ev).strip().upper() in {'EGRESO','E'}:
                    es_egreso = True
            dc = sheet.cell(rn, c_ded)
            dc.value     = ded_val
            dc.fill      = (blue_fill if es_egreso else green_fill) if es_ded else red_fill
            dc.font      = bold_white
            dc.alignment = center_align

        # Razón
        if c_razon:
            rc = sheet.cell(rn, c_razon)
            if ded_val == 'NO':
                rc.fill = red_fill;    rc.font = bold_white
            elif rc.value and rc.value != 'Cumple requisitos':
                rc.fill = yellow_fill; rc.font = bold_black
            else:
                rc.fill = green_fill;  rc.font = Font(color='006100')

    # Anchos
    if c_razon: sheet.column_dimensions[get_column_letter(c_razon)].width = 65
    for n in calc_names + ['Deducible']:
        c = fc(n)
        if c: sheet.column_dimensions[get_column_letter(c)].width = 15

    wb.save(output_path)
    print(f"  🔧 Insertando valores cacheados ({len(_cache_formulas):,} celdas)...")
    _parchear_cache_formulas(output_path, _cache_formulas)
    print(f"  ✅ Fórmulas auditables y colores aplicados")

    # Estadísticas
    f  = df['_forma'].fillna('')
    r  = df['_regimen'].fillna('')
    eg = df['_es_gas'].fillna(False)
    ei = df['_es_insumo'].fillna(False)
    t  = df[C['total']].fillna(0).astype(float)

    return {
        'regimenes':    r.value_counts().to_dict(),
        'dulces_ieps8': int((df['_es_dulce'] & (i8 > 0)).sum()),
        'dulces_sin':   int((df['_es_dulce'] & (i8 == 0)).sum()),
        'gas_ieps':     int(mask_ieps_gas.sum()),
        'gas_sin':      int((eg & ~mask_ieps_gas).sum()),
        'gas_626':      int((eg & (f=='01') & (r=='626')).sum()),
        'gas_626_agrup':int(df['_agrupada'].sum()),
        'gas_612':      int((eg & (f=='01') & (r=='612')).sum()),
        'gas_elec':     int((eg & (f!='01')).sum()),
        'ins_nd':       int((ei & ~eg & (f=='01') & (t>LIMITE_EFECTIVO)).sum()),
        'ins_menor':    int((ei & ~eg & (f=='01') & (t<=LIMITE_EFECTIVO)).sum()),
        'ins_elec':     int((ei & ~eg & f.isin(FORMAS_ELECTRONICAS)).sum()),
        's01':          int((df['_uso']=='S01').sum()),
        'ef_mayor':     int((~eg & ~ei & (f=='01') & (t>LIMITE_EFECTIVO)).sum()),
    }
