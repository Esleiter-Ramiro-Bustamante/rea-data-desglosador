"""
generador_reporte.py — Modulo de Reporte Fiscal v2.3
ReaDesF1.8

CORRECCIONES v2.2:
  - uuid_rel extraido de columna COMENTARIOS cuando viene vacio
    (los motores escriben "Complemento parcialidad 1 - factura 5D1EDA96AB6E saldo $0")
  - Cruce PPD<->complemento robusto: por sufijo 12 chars hex + UUID completo
  - PPD ya cubierto NO aparece como PENDIENTE
  - UUID completo visible en HTML con scroll horizontal (no truncado)
  - Ceros mostrados como guion en HTML (no "0.0")
  - classify() unifica color+css+filtro en 1 sola funcion (1 solo .upper() por fila)
  - idx_razones construido UNA vez -> O(1) por busqueda
  - is_cp calculado UNA vez por fila
  - StringIO para HTML (no concatenar string gigante en RAM)
  - Regimen detectado dinamicamente por cliente (612/626/otros)
"""

import os
import re
import time
from io import StringIO
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

USOS_VALIDOS    = {'G01', 'G02', 'G03'}
METODOS_VALIDOS = {'PUE', 'PPD'}
FORMAS_VALIDAS  = {'01', '02', '03', '04', '28'}
LIMITE_EFECTIVO = 2000.0

# Claves de Deducción Personal (Art. 147 LISR - CFDI 4.0)
CLAVES_DEDUCCION_PERSONAL = {'D01','D02','D03','D04','D05','D06','D07','D08','D09','D10'}

_RE_UUID_FULL  = re.compile(
    r'[0-9A-Fa-f]{8}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{12}',
    re.IGNORECASE
)
_RE_UUID_SHORT = re.compile(r'\b([0-9A-Fa-f]{12})\b')


class ReportTemplateManager:
    """Patrón Gestor de Activos (Singleton) para manejar las plantillas HTML, CSS y JS."""
    _html_template = None
    _templates_cache = {}

    @classmethod
    def _load_html(cls):
        if cls._html_template is None:
            path = Path(__file__).parent.parent / 'views' / 'assets' / 'reporte_template.html'
            with open(path, 'r', encoding='utf-8') as fp:
                cls._html_template = fp.read()
            # Extraer sub-templates
            import re
            matches = re.finditer(r'<template id="([^"]+)">\s*(.*?)\s*</template>', cls._html_template, re.DOTALL)
            for m in matches:
                cls._templates_cache[m.group(1)] = m.group(2)

    @classmethod
    def get_html_template(cls):
        cls._load_html()
        # Retorna el HTML sin los tags <template>
        import re
        return re.sub(r'<template id="[^"]+">.*?</template>\s*', '', cls._html_template, flags=re.DOTALL)

    @classmethod
    def get_sub_template(cls, id_name):
        cls._load_html()
        return cls._templates_cache.get(id_name, '')

    @classmethod
    def copy_asset(cls, template_name, out_path):
        src_path = Path(__file__).parent.parent / 'views' / 'assets' / template_name
        with open(src_path, 'r', encoding='utf-8') as fp:
            content = fp.read()
        with open(out_path, 'w', encoding='utf-8') as fp:
            fp.write(content)
        print('  Asset: %s' % out_path)



def extraer_codigo(val, n=3):
    s = str(val or '').strip()
    if '-' in s:
        return s.split('-')[0].strip()[:n].upper()
    return s[:n].upper()

def detectar_deduccion_personal(conceptos: str, regimen: str = '612') -> bool:
    """
    Detecta si el concepto contiene claves de deducción personal (D01-D10).
    Solo aplica para Régimen 612.
    
    Args:
        conceptos: String con descripción/conceptos del CFDI
        regimen: Código de régimen (solo detecta si es 612)
        
    Returns:
        True si es régimen 612 y contiene clave D01-D10
    """
    # Solo aplica para régimen 612
    if regimen != '612':
        return False
    
    if not conceptos:
        return False
    
    conceptos_upper = str(conceptos).upper()
    
    # Buscar cualquiera de las claves de deducción personal
    for clave in CLAVES_DEDUCCION_PERSONAL:
        if clave in conceptos_upper:
            # Validar que sea la clave completa con separadores
            if re.search(r'(\s|^|[|,;-])' + clave + r'(\s|$|[|,;.\)])', conceptos_upper):
                return True
    
    return False

def safe_str(v):
    return str(v).strip() if v is not None else ''

def norm_uuid(u):
    return re.sub(r'[-\s]', '', str(u or '')).lower()

def fmt_num(v):
    try:
        f = float(v)
        return '$%s' % '{:,.2f}'.format(f) if f != 0 else '-'
    except Exception:
        return '-'

def extraer_uuids_de_texto(texto):
    """Extrae UUIDs de texto libre. Busca UUID completo primero, luego sufijo 12 hex."""
    if not texto:
        return []
    completos = _RE_UUID_FULL.findall(texto)
    if completos:
        return completos
    return _RE_UUID_SHORT.findall(texto)

def classify(estatus):
    """Retorna (bg_hex, fg_hex, css_cls, flt_cls). Un solo .upper() por fila."""
    eu = estatus.upper()
    is_efe = 'EFE' in eu
    if 'COMPLEMENTO' in eu: return ('D5C6E0','4A0080','complemento','complemento')
    if 'DED PERSONAL NO DEDUCIBLE' in eu: return ('FFC7CE','9C0006','no-ded','no-ded')
    if 'DED PERSONAL' in eu: return ('B4E7B4','0B6623','ded-personal','ded-personal')
    if '16 Y 0'      in eu: return ('BDD7EE','1F497D','mix',  'efe160' if is_efe else 'ded160')
    if '16%'         in eu: return ('C6EFCE','1B5E20','ded16','efe16' if is_efe else 'ded16')
    if '0%'          in eu: return ('FFF3CD','856404','ded0', 'efe0' if is_efe else 'ded0')
    if 'EGRESO'      in eu: return ('E2CFED','6A1B9A','egreso','egreso')
    if 'PENDIENTE'   in eu: return ('E0E0E0','37474F','pendiente','pendiente')
    if 'ERROR'       in eu: return ('FF4444','FFFFFF','no-ded','no-ded')
    if 'NO DED'      in eu: return ('FFC7CE','9C0006','no-ded','no-ded')
    return                         ('FFFFFF','000000','no-ded','no-ded')


def leer_validado(path):
    """Lee _validado.xlsx. Extrae uuid_rel de COMENTARIOS si viene vacio."""
    wb    = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    hraw  = {}
    hnorm = {}
    for c in sheet[1]:
        if c.value is not None:
            raw = str(c.value).strip()
            hraw[c.column - 1] = raw
            hnorm[raw.lower()] = c.column - 1

    def gc(name):
        n = name.strip().lower()
        if n in hnorm:
            return hnorm[n]
        for k, v in hnorm.items():
            if n in k or k in n:
                return v
        return None

    # ── BÚSQUEDA ROBUSTA DE UUID ──────────────────────────────────
    # Intenta múltiples variantes para encontrar UUID
    i_uuid = None
    for variant in ['uuid', 'folio fiscal', 'folio', 'id cfdi']:
        i_uuid = gc(variant)
        if i_uuid is not None:
            print(f'  ✅ UUID encontrado como: "{variant.upper()}"')
            break
    
    if i_uuid is None:
        print(f'  ⚠️  ADVERTENCIA: Columna UUID NO encontrada')
        print(f'  📋 Columnas disponibles: {", ".join(hraw.values())}')
        i_uuid = None  # Permitir continuar pero sin UUID
    
    i_uuid_rel  = gc('uuids relacionados') or gc('uuid relacionado') or gc('folio fiscal relacionado')
    i_fecha     = gc('fecha certificacion') or gc('fecha emision') or gc('fecha')
    i_razon_em  = gc('razon emisor') or gc('nombre emisor') or gc('razon social')
    i_razon_rec = gc('razon receptor') or gc('nombre receptor')
    i_regimen   = gc('regimen receptor') or gc('regimen')
    i_metodo    = gc('metodo pago') or gc('metodo de pago')
    i_forma     = gc('forma pago') or gc('forma de pago')
    i_uso       = gc('uso cfdi')
    i_subtotal  = gc('subtotal')
    i_descuento = gc('descuento')
    i_iva16     = gc('iva trasladado 16%') or gc('iva 16%')
    i_iva0      = gc('iva trasladado 0%')  or gc('iva 0%')
    i_iva_ex    = gc('iva exento')
    i_ieps      = gc('ieps trasladado') or gc('ieps')
    i_ieps_nd   = gc('ieps trasladado no desglosado')
    i_ieps_g    = gc('ieps trasladado')
    i_ieps3     = gc('ieps trasladado 3%')
    i_total     = gc('total')
    i_conceptos = gc('conceptos') or gc('descripcion')
    i_complem   = gc('complementos')
    i_efecto    = gc('efecto')
    i_sub0      = gc('sub0%')    or gc('sub0')
    i_sub2      = gc('sub2-16%') or gc('sub2')
    i_sub1      = gc('sub1-16%') or gc('sub1')
    i_coment    = gc('comentarios') or gc('comentario') or gc('observaciones')
    i_rfc_em    = gc('rfc emisor') or gc('rfc')

    def rv(row, idx):
        return row[idx] if (idx is not None and idx < len(row)) else None

    def rnum(row, idx):
        try:
            v = rv(row, idx)
            return float(v) if v is not None else 0.0
        except Exception:
            return 0.0

    filas = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        st     = rnum(row, i_subtotal)
        dc     = rnum(row, i_descuento)
        iva16  = rnum(row, i_iva16)
        iva0   = rnum(row, i_iva0)
        iva_ex = rnum(row, i_iva_ex)
        total  = rnum(row, i_total)
        ieps_nd = rnum(row, i_ieps_nd)
        ieps_g  = rnum(row, i_ieps_g)
        ieps3   = rnum(row, i_ieps3) if i_ieps3 is not None else 0.0

        # ── sub0: misma lógica que el motor, usando valores crudos ────
        # No leer la columna SUB0% (es fórmula Excel sin caché en Python)
        # Gasolina con IEPS: iva0 ya tiene el IEPS, iva_ex es el mismo valor
        # → sumarlos duplicaría. Solo usar iva0.
        ieps_activo = ieps_nd if ieps_nd > 0 else ieps_g
        # Gasolina con IEPS: motor copia IEPS a IVA_0% e IVA_Exento
        # Si iva0 == iva_ex > 0 → no duplicar, usar solo iva0
        if iva0 > 0 and abs(iva0 - iva_ex) < 0.02:
            sub0 = round(iva0, 2)
        else:
            sub0 = round(iva0 + iva_ex, 2)

        # SUB2: SIEMPRE calcular desde valores crudos del CFDI
        # IEPS 3% telefonía se suma a SUB1 igual que IEPS 8% dulces
        # Formula garantizada: SUB1 = SubTotal - Descuento + IEPS3
        #                      SUB2 = SUB1 - SUB0
        sub1 = round(st - dc + ieps3, 2)
        sub2 = round(max(sub1 - sub0, 0), 2)

        uuid_rel_raw = safe_str(rv(row, i_uuid_rel))
        comentarios  = safe_str(rv(row, i_coment))
        conceptos    = safe_str(rv(row, i_conceptos))
        rfc_em       = safe_str(rv(row, i_rfc_em))

        # Si uuid_rel vacio, extraer de COMENTARIOS o CONCEPTOS
        uuid_rel = uuid_rel_raw
        if not uuid_rel:
            for fuente in [comentarios, conceptos]:
                encontrados = extraer_uuids_de_texto(fuente)
                if encontrados:
                    uuid_rel = encontrados[0]
                    break

        filas.append({
            'uuid':         safe_str(rv(row, i_uuid)),
            'uuid_rel':     uuid_rel,
            'fecha':        rv(row, i_fecha),
            'razon_em':     safe_str(rv(row, i_razon_em)),
            'razon_rec':    safe_str(rv(row, i_razon_rec)),
            'regimen':      safe_str(rv(row, i_regimen)),
            'metodo':       safe_str(rv(row, i_metodo)),
            'forma':        safe_str(rv(row, i_forma)),
            'uso':          safe_str(rv(row, i_uso)),
            'subtotal':     round(st, 2),
            'descuento':    round(dc, 2),
            'iva16':        round(iva16, 2),
            'iva0':         round(iva0, 2),
            'iva_ex':       round(iva_ex, 2),
            'ieps':         rnum(row, i_ieps),
            'total':        round(total, 2),
            'sub0':         sub0,
            'sub2':         sub2,
            'conceptos':    conceptos,
            'complementos': safe_str(rv(row, i_complem)),
            'efecto':       safe_str(rv(row, i_efecto)),
            'comentarios':  comentarios,
            'rfc_em':       rfc_em,
        })

    wb.close()
    print('  OK %d filas leidas - %d columnas' % (len(filas), len(hraw)))
    return filas


NOMBRES_REGIMEN = {
    '612': 'Actividades Empresariales y Profesionales',
    '626': 'Regimen Simplificado de Confianza (RESICO)',
    '601': 'General de Ley Personas Morales',
    '606': 'Arrendamiento',
    '621': 'Incorporacion Fiscal',
    '616': 'Sin obligaciones fiscales',
}

def detectar_regimen(filas):
    conteo = {}
    for f in filas:
        cod = extraer_codigo(f.get('regimen', ''))
        if cod and cod.isdigit() and len(cod) == 3:
            conteo[cod] = conteo.get(cod, 0) + 1
    if not conteo:
        return '612', NOMBRES_REGIMEN['612']
    dom = max(conteo, key=conteo.get)
    return dom, NOMBRES_REGIMEN.get(dom, 'Regimen %s' % dom)


def es_complemento(f):
    uso   = f.get('uso', '').upper()
    efect = f.get('efecto', '').upper()
    conc  = f.get('conceptos', '').upper()
    comp  = f.get('complementos', '').upper()
    return ('CP01' in uso or efect == 'PAGO' or (conc == 'PAGO' and comp == 'PAGO'))

def construir_indice_razones(filas):
    """Dict {uuid_norm: razon} + {rfc: razon}. Incluye sufijo 12 chars."""
    idx = {}
    for f in filas:
        uuid = f.get('uuid', '')
        if not uuid:
            continue
        razon = f.get('razon_em', '')
        un    = norm_uuid(uuid)
        idx[un] = razon
        if len(un) >= 12:
            idx[un[-12:]] = razon
        # Indexar por RFC para cruce cuando uuid_rel esta vacio
        rfc = f.get('rfc_em', '').strip().upper()
        if rfc:
            idx['rfc:' + rfc] = razon
    return idx

def detectar_ppd(filas):
    """
    UUIDs PPD sin CP01. Cruce en 2 niveles:
      Nivel 1: uuid_rel explicito (cuando el motor lo llena)
      Nivel 2: RFC emisor + total identico (cuando uuid_rel esta vacio)
    Retorna (ppd_pendientes, mapa_cp01_a_uuid_ppd)
    """
    # Todos los PPD: {uuid_norm: fila_completa}
    ppd_map = {}
    for f in filas:
        if extraer_codigo(f.get('metodo', '')) == 'PPD' and f.get('uuid'):
            ppd_map[norm_uuid(f['uuid'])] = f

    cubiertos   = set()   # uuid_norm de PPDs con CP01
    cp01_a_ppd  = {}      # uuid_norm(cp01) -> uuid_original(ppd)

    for f in filas:
        if not es_complemento(f):
            continue

        # NIVEL 1: uuid_rel explicito en el _validado
        rel = f.get('uuid_rel', '').strip()
        if rel:
            rel_n    = norm_uuid(rel)
            rel_tail = rel_n[-12:] if len(rel_n) >= 12 else rel_n
            for uid_n, ppd_f in ppd_map.items():
                if uid_n == rel_n or uid_n.endswith(rel_tail):
                    cubiertos.add(uid_n)
                    cp01_a_ppd[norm_uuid(f['uuid'])] = ppd_f['uuid']
            continue

        # NIVEL 2: uuid_rel vacio → cruzar por RFC emisor + total igual
        rfc_cp01   = f.get('rfc_em', '').strip().upper()
        total_cp01 = round(float(f.get('total', 0)), 2)
        if not rfc_cp01:
            continue
        for uid_n, ppd_f in ppd_map.items():
            rfc_ppd   = ppd_f.get('rfc_em', '').strip().upper()
            total_ppd = round(float(ppd_f.get('total', 0)), 2)
            if rfc_cp01 == rfc_ppd and total_cp01 == total_ppd:
                cubiertos.add(uid_n)
                cp01_a_ppd[norm_uuid(f['uuid'])] = ppd_f['uuid']
                break  # un CP01 cubre un PPD

    ppd_pendientes = {ppd_map[n]['uuid'] for n in ppd_map if n not in cubiertos}
    return ppd_pendientes, cp01_a_ppd

def resolver_uuid_rel(f, idx_razones, cp01_a_ppd=None):
    """
    Retorna (display_12chars, razon_proveedor) para fila complemento.
    Busca en: 1) uuid_rel explicito  2) mapa cp01_a_ppd por RFC+total
    display = ultimos 12 chars del UUID del PPD relacionado (en MAYUSCULAS)
    """
    rel = f.get('uuid_rel', '').strip()

    # NIVEL 1: uuid_rel explicito
    if rel:
        rel_n    = norm_uuid(rel)
        rel_tail = rel_n[-12:] if len(rel_n) >= 12 else rel_n
        razon    = idx_razones.get(rel_n) or idx_razones.get(rel_tail, '')
        display  = rel_tail.upper()
        return display, razon

    # NIVEL 2: buscar en mapa cp01_a_ppd
    if cp01_a_ppd:
        cp_n    = norm_uuid(f.get('uuid', ''))
        ppd_uuid = cp01_a_ppd.get(cp_n, '')
        if ppd_uuid:
            ppd_n    = norm_uuid(ppd_uuid)
            rel_tail = ppd_n[-12:].upper() if len(ppd_n) >= 12 else ppd_n.upper()
            razon    = idx_razones.get(ppd_n) or idx_razones.get(ppd_n[-12:], '')
            if not razon:
                razon = f.get('razon_em', '')  # mismo proveedor
            return rel_tail, razon

    return '', ''


def calc_estatus(f, is_cp, ppd_pend, idx_razones):
    if is_cp:
        # ESTATUS = solo "COMPLEMENTO" limpio (sin monto, sin redundancia)
        # El detalle (uuid_rel, razon, saldo) va SOLO en columna OBS
        return 'COMPLEMENTO'

    uuid_n = norm_uuid(f.get('uuid', ''))
    if any(norm_uuid(p) == uuid_n for p in ppd_pend):
        return 'PENDIENTE'

    uso   = f.get('uso', '');   metodo = f.get('metodo', '')
    forma = f.get('forma', ''); total  = f.get('total', 0)
    sub2  = f.get('sub2', 0);  iva16  = f.get('iva16', 0)
    sub0  = f.get('sub0', 0)
    regimen = f.get('regimen', '')
    conceptos = f.get('conceptos', '')

    # ── DETECCIÓN DE DEDUCCIÓN PERSONAL (D01-D10) ──────────────────
    # Si está en régimen 612 y tiene claves D01-D10
    fp = extraer_codigo(forma, 2)
    if detectar_deduccion_personal(conceptos, extraer_codigo(regimen)):
        if fp == '01':
            return 'DED PERSONAL NO DEDUCIBLE POR QUE LO PAGA EN EFECTIVO'
        return 'DED PERSONAL'

    u  = extraer_codigo(uso)
    m  = extraer_codigo(metodo)

    if u in ('S01', 'CN0'):
        return 'NO DEDUCIBLE'

    u_ok = u in USOS_VALIDOS
    m_ok = m in METODOS_VALIDOS
    # Forma 99 = "Por definir" → válida SOLO para PPD (es temporal hasta que llega el CP01)
    f_ok = fp in FORMAS_VALIDAS or (fp == '99' and m == 'PPD')

    if u_ok and m_ok and f_ok:
        if fp == '01' and total >= LIMITE_EFECTIVO:
            return 'NO DEDUCIBLE: Efectivo >= $2,000'
        if u == 'G02':
            return 'EGRESO'
        if   sub2 > 0 and iva16 > 0 and sub0 == 0: suf = '16%'
        elif sub2 > 0 and iva16 > 0 and sub0 > 0:  suf = '16 Y 0%'
        elif sub2 == 0 and iva16 == 0 and sub0 > 0: suf = '0%'
        else:                                        suf = 'NO DEDUCIBLE'
        return ('EFE ' if fp == '01' else 'DED ') + suf

    err = []
    if not u_ok: err.append('USO (%s) INVALIDO' % u)
    if not m_ok: err.append('METODO (%s) INVALIDO' % m)
    if not f_ok: err.append('FORMA (%s) INVALIDA' % fp)
    return 'ERROR: ' + ' | '.join(err)


def formula_estatus(rn):
    L = LIMITE_EFECTIVO
    return (
        '=IF(OR(LEFT(K%(r)s,3)="S01",LEFT(K%(r)s,3)="CN0"),"NO DEDUCIBLE",'
        'IF(NOT(ISERROR(MATCH(LEFT(K%(r)s,3),{"D01","D02","D03","D04","D05","D06","D07","D08","D09","D10"},0))),'
        'IF(AND(NOT(ISERROR(MATCH(LEFT(I%(r)s,3),{"PUE","PPD"},0))),NOT(ISERROR(MATCH(LEFT(J%(r)s,2),{"01","02","03","04","28"},0)))),'
        'IF(LEFT(J%(r)s,2)="01","DED PERSONAL NO DEDUCIBLE POR QUE LO PAGA EN EFECTIVO","DED PERSONAL"),'
        '"ERROR: "&IF(ISERROR(MATCH(LEFT(I%(r)s,3),{"PUE","PPD"},0)),"METODO INVALIDO | ","")&IF(ISERROR(MATCH(LEFT(J%(r)s,2),{"01","02","03","04","28"},0)),"FORMA INVALIDA","")),'
        'IF(AND(NOT(ISERROR(MATCH(LEFT(K%(r)s,3),{"G01","G02","G03"},0))),NOT(ISERROR(MATCH(LEFT(I%(r)s,3),{"PUE","PPD"},0))),NOT(ISERROR(MATCH(LEFT(J%(r)s,2),{"01","02","03","04","28"},0)))),'
        'IF(AND(LEFT(J%(r)s,2)="01",H%(r)s>=%(L)s),"NO DEDUCIBLE: Efectivo >= $2,000",'
        'IF(LEFT(K%(r)s,3)="G02","EGRESO",'
        'IF(LEFT(J%(r)s,2)="01",'
        '"EFE "&IF(AND(E%(r)s>0,F%(r)s>0,G%(r)s=0),"16%%",IF(AND(E%(r)s>0,F%(r)s>0,G%(r)s>0),"16 Y 0%%",IF(AND(E%(r)s=0,F%(r)s=0,G%(r)s>0),"0%%","NO DEDUCIBLE"))),'
        '"DED "&IF(AND(E%(r)s>0,F%(r)s>0,G%(r)s=0),"16%%",IF(AND(E%(r)s>0,F%(r)s>0,G%(r)s>0),"16 Y 0%%",IF(AND(E%(r)s=0,F%(r)s=0,G%(r)s>0),"0%%","NO DEDUCIBLE")))))),'
        '"ERROR: "&IF(ISERROR(MATCH(LEFT(K%(r)s,3),{"G01","G02","G03","D01","D02","D03","D04","D05","D06","D07","D08","D09","D10"},0)),"USO INVALIDO | ","")&'
        'IF(ISERROR(MATCH(LEFT(I%(r)s,3),{"PUE","PPD"},0)),"METODO INVALIDO | ","")&'
        'IF(ISERROR(MATCH(LEFT(J%(r)s,2),{"01","02","03","04","28"},0)),"FORMA INVALIDA",""))))'
    ) % {'r': rn, 'L': L}


def generar_excel(filas, ppd_pend, cp01_a_ppd, idx_razones, out, mes='', reg_cod='612', reg_nombre=''):
    print('  Generando Excel...')
    wb  = openpyxl.Workbook()
    sh  = wb.active
    sh.title = 'Reporte Fiscal'

    mk  = lambda h: PatternFill(start_color=h, end_color=h, fill_type='solid')
    ct  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lt  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    th  = Side(style='thin',   color='CCCCCC')
    tk  = Side(style='medium', color='0A0A0A')
    bd  = Border(left=th, right=th, top=th, bottom=th)
    bdh = Border(left=tk, right=tk, top=tk, bottom=tk)

    NCOLS = 13
    sh.merge_cells('A1:%s1' % get_column_letter(NCOLS))
    t = sh['A1']
    t.value = 'RESUMEN GASTOS MENSUALES - %s  .  %s' % (reg_nombre.upper(), mes.upper())
    t.font  = Font(bold=True, size=12, name='Calibri')
    t.fill  = mk('D6EAF8'); t.alignment = ct; t.border = bdh
    sh.row_dimensions[1].height = 34

    sh.merge_cells('A2:D2')
    sh['A2'].value = 'PUNTOS DEDUCIBLES DEL SISTEMA'
    sh['A2'].font  = Font(bold=True, size=9)
    sh['A2'].fill  = mk('D5F5E3'); sh['A2'].alignment = ct
    for col, val in [('E','PUE'), ('G','01, 02, 03, 28'), ('I','G01 Y G03')]:
        c = sh['%s2' % col]
        c.value = val; c.font = Font(bold=True, size=9)
        c.fill  = mk('D5F5E3'); c.alignment = ct

    sh.merge_cells('A3:%s3' % get_column_letter(NCOLS))
    sh['A3'].value = 'Regimen %s - %s' % (reg_cod, reg_nombre)
    sh['A3'].font  = Font(italic=True, size=9, color='555555')
    sh['A3'].alignment = ct; sh.row_dimensions[3].height = 18

    RH = 4
    hdrs   = ['FOLIO FISCAL','UUID RELACIONADO','FECHA','RAZON SOCIAL',
              'SUBTOTAL 16%','IVA 16%','SUB 0%','TOTAL',
              'METODO DE PAGO','FORMA DE PAGO','USO CFDI',
              'ESTATUS','COMPLEMENTOS / OBSERVACIONES']
    widths = [44,44,19,32,15,13,13,15,30,26,30,28,60]

    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = sh.cell(row=RH, column=i, value=h)
        c.font = Font(bold=True, size=9, name='Calibri', color='1A1A2E')
        c.fill = mk('AED6F1'); c.alignment = ct; c.border = bdh
        sh.column_dimensions[get_column_letter(i)].width = w
    sh.row_dimensions[RH].height = 34
    sh.freeze_panes = 'A%d' % (RH+1)

    stats = {k:0 for k in ['total','ded','ded_personal','no_ded','pend','egreso','efe','comp',
                             'monto_ded','monto_no_ded','monto_pend']}
    ppd_norms = {norm_uuid(p) for p in ppd_pend}

    rn = RH + 1
    for f in filas:
        is_cp   = es_complemento(f)
        estatus = calc_estatus(f, is_cp, ppd_pend, idx_razones)
        bg, fg, _, _ = classify(estatus)
        eu = estatus.upper()

        if is_cp:
            display, razon = resolver_uuid_rel(f, idx_razones, cp01_a_ppd)
            saldo = f.get('total', 0)
            obs   = ('Complemento parcialidad 1 - factura %s' % display) if display else 'Complemento CP01'
            if razon: obs += ' (%s)' % razon[:30]
            obs  += ' saldo insoluto $%s' % '{:,.2f}'.format(saldo)
        elif norm_uuid(f.get('uuid','')) in ppd_norms:
            obs = 'PPD sin complemento CP01 - pasa al siguiente mes'
        else:
            obs = ''

        stats['total'] += 1
        if   'COMPLEMENTO' in eu: stats['comp']      += 1
        elif 'PENDIENTE'   in eu: stats['pend']      += 1; stats['monto_pend']   += f['total']
        elif 'DED PERSONAL' in eu: stats['ded_personal'] += 1; stats['monto_ded'] += f['total']
        elif 'NO DED' in eu or 'ERROR' in eu:
                                   stats['no_ded']   += 1; stats['monto_no_ded'] += f['total']
        elif 'EGRESO'      in eu: stats['egreso']    += 1
        elif 'EFE'         in eu: stats['efe']       += 1; stats['monto_ded']    += f['total']
        else:                      stats['ded']      += 1; stats['monto_ded']    += f['total']

        # CP01 → amarillo en _validado (igual que imagen de referencia)
        if is_cp:
            cf = 'FFFF00'  # Amarillo para CP01
        elif norm_uuid(f.get('uuid','')) in ppd_norms:
            cf = 'FFFF00'  # Amarillo tambien para PPD relacionado
        else:
            cf = 'FFFFFF' if (rn-RH-1)%2==0 else 'F8F9FA'

        # UUID relacionado: display de 12 chars del PPD
        display_rel, _ = resolver_uuid_rel(f, idx_razones, cp01_a_ppd) if is_cp else ('', '')

        vals = [f['uuid'], display_rel or f['uuid_rel'],f['fecha'],f['razon_em'],
                f['sub2'] or None,f['iva16'] or None,
                f['sub0'] or None,f['total'] or None,
                f['metodo'],f['forma'],f['uso'],None,obs]

        for ci, val in enumerate(vals, 1):
            c = sh.cell(row=rn, column=ci, value=val)
            c.border = bd
            is_num = ci in (5,6,7,8)
            c.font = Font(size=9, name='Calibri', bold=is_num, color='1A1A2E')
            c.alignment = lt if ci in (1,2,4,13) else ct
            if is_num: c.number_format = '"$"#,##0.00'
            if ci==12: c.fill=mk(bg); c.font=Font(bold=True,color=fg,size=9,name='Calibri')
            elif is_cp: c.fill=mk('EDE0F5')
            else: c.fill=mk(cf)

        ec = sh.cell(row=rn, column=12)
        ec.value = estatus if (is_cp or 'PENDIENTE' in eu) else formula_estatus(rn)
        ec.fill=mk(bg); ec.font=Font(bold=True,color=fg,size=9,name='Calibri')
        ec.alignment=ct; ec.border=bd
        sh.row_dimensions[rn].height = 28
        rn += 1

    sh.merge_cells('A%d:D%d' % (rn,rn))
    tc = sh.cell(row=rn, column=1, value='TOTALES')
    tc.font=Font(bold=True,size=11); tc.fill=mk('D6EAF8')
    tc.alignment=ct; tc.border=bdh
    for cl,ci in [('E',5),('F',6),('G',7),('H',8)]:
        c = sh.cell(row=rn, column=ci)
        c.value='=SUM(%s%d:%s%d)' % (cl,RH+1,cl,rn-1)
        c.number_format='"$"#,##0.00'; c.fill=mk('D6EAF8')
        c.border=bdh; c.font=Font(bold=True,size=10); c.alignment=ct

    wb.save(out)
    print('  Excel: %s' % out)
    return stats


_OPCIONES = [
    'DED 16%','DED 0%','DED 16 Y 0%','DED PERSONAL','EFE 16%','EFE 0%','EFE 16 Y 0%',
    'EGRESO','NO DEDUCIBLE','NO DEDUCIBLE: Efectivo >= $2,000','PENDIENTE','COMPLEMENTO',
    'OTRO MES', 'MES ANTERIOR'
]

def generar_html(filas, ppd_pend, cp01_a_ppd, idx_razones, out, mes='', stats=None, reg_cod='612', reg_nombre=''):
    print('  Generando HTML...')
    s         = stats or {}
    total_f   = s.get('total', len(filas))
    ppd_norms = {norm_uuid(p) for p in ppd_pend}
    now_str   = datetime.now().strftime('%d/%m/%Y %H:%M')
    n_ded=s.get('ded',0); n_efe=s.get('efe',0); n_no_ded=s.get('no_ded',0)
    n_pend=s.get('pend',0); n_egreso=s.get('egreso',0); n_comp=s.get('comp',0)
    m_ded=s.get('monto_ded',0); m_no_ded=s.get('monto_no_ded',0); m_pend=s.get('monto_pend',0)

    buf = StringIO()
    for i, f in enumerate(filas, 1):
        is_cp   = es_complemento(f)
        estatus = calc_estatus(f, is_cp, ppd_pend, idx_razones)
        bg, fg, css, flt = classify(estatus)

        opts = ''.join(
            '<option value="%s"%s>%s</option>' % (op, ' selected' if op==estatus else '', op)
            for op in _OPCIONES
        )
        if estatus not in _OPCIONES:
            opts += '<option value="%s" selected>%s</option>' % (estatus, estatus)
        sel = '<select class="se %s" data-original="%s" onchange="ce(this)">%s</select>' % (css, estatus, opts)

        uuid_full = f['uuid']
        uuid_disp = uuid_full.upper()

        if is_cp:
            display, razon = resolver_uuid_rel(f, idx_razones, cp01_a_ppd)
            saldo = f.get('total', 0)
            rel_disp = display if display else '–'
            if display:
                obs_txt = 'Complemento parcialidad 1 - factura %s' % display
                if razon: obs_txt += ' (%s)' % razon[:30]
            else:
                obs_txt = 'Complemento CP01'
            obs_txt += ' saldo insoluto $%s' % '{:,.2f}'.format(saldo)
            obs_html = '<span class="cb" contenteditable="true" title="Clic para editar">%s</span>' % obs_txt
        elif norm_uuid(f.get('uuid','')) in ppd_norms:
            rel_raw  = f['uuid_rel']
            rel_disp = (rel_raw[-12:].upper() if len(rel_raw)>=12 else rel_raw.upper()) if rel_raw else '–'
            obs_html = '<span class="pb" contenteditable="true">PPD sin CP01 - siguiente mes</span>'
        else:
            rel_raw  = f['uuid_rel']
            rel_disp = (rel_raw[-12:].upper() if len(rel_raw)>=12 else rel_raw.upper()) if rel_raw else '–'
            obs_html = '<span class="ob-txt" contenteditable="true">-</span>'

        fecha_raw = f['fecha']
        try:
            if hasattr(fecha_raw, 'strftime'):
                fecha_str = fecha_raw.strftime('%d/%m/%Y')
            else:
                s = str(fecha_raw or '').split(' ')[0]
                if '-' in s:
                    p = s.split('-')
                    fecha_str = f"{p[2]}/{p[1]}/{p[0]}" if len(p)==3 else s
                else:
                    fecha_str = s[:10]
        except Exception:
            fecha_str = str(fecha_raw or '')[:10]

        fc_cls = ' fc' if is_cp else ''
        tpl_row = ReportTemplateManager.get_sub_template('tpl-row')
        buf.write(tpl_row.format(
            flt=flt, fc_cls=fc_cls, rfc=f.get('rfc_em','').strip().upper(),
            uuid_full=uuid_full, i=i, uuid_disp=uuid_disp,
            uuid_rel=f.get('uuid_rel', ''), rel_disp=rel_disp,
            fecha_str=fecha_str, razon_em=f.get('razon_em', ''),
            sub2=f.get('sub2', 0), iva16=f.get('iva16', 0), sub0=f.get('sub0', 0),
            total=fmt_num(f.get('total', 0)), metodo=f.get('metodo', ''),
            forma=f.get('forma', ''), uso=f.get('uso', ''),
            sel=sel, obs_html=obs_html
        ) + '\n')

    fs = buf.getvalue(); buf.close()

    css_out = re.sub(r'\.html$', '', out, flags=re.IGNORECASE) + '.css'
    js_out  = re.sub(r'\.html$', '', out, flags=re.IGNORECASE) + '.js'
    ReportTemplateManager.copy_asset('reporte_template.css', css_out)
    ReportTemplateManager.copy_asset('reporte_template.js', js_out)
    css_name = os.path.basename(css_out)
    js_name  = os.path.basename(js_out)

    template = ReportTemplateManager.get_html_template()
    html = template.format(
        mes=mes, reg_cod=reg_cod, reg_nombre=reg_nombre, now=now_str,
        tf=total_f, nn=n_no_ded, np=n_pend, nd2=(n_ded+n_efe), fs=fs,
        diot_html=generar_diot_html(filas, ppd_pend, idx_razones, mes),
        css_name=css_name, js_name=js_name
    )

    with open(out, 'w', encoding='utf-8') as fp:
        fp.write(html)
    print('  HTML: %s' % out)


def generar_diot_html(filas, ppd_pend, idx_razones, mes):
    """
    Genera la sección HTML de la DIOT (Informativa de Operaciones con Terceros).
    Agrupa por RFC emisor, solo facturas deducibles (DED + EFE).

    FIX v2.4: acumula sub2/iva16/sub0 sin round() intermedio para evitar
    deriva de centavos al sumar N facturas. Se redondea solo al mostrar.
    """
    # Acumular por RFC — sin redondeo intermedio (float64 nativo)
    proveedores = {}
    for f in filas:
        is_cp   = es_complemento(f)
        estatus = calc_estatus(f, is_cp, ppd_pend, idx_razones).upper()

        # Solo deducibles — excluir NO DED, PENDIENTE, COMPLEMENTO, EGRESO, ERROR
        if not any(k in estatus for k in ('DED', 'EFE')):
            continue
        if any(k in estatus for k in ('NO DED', 'NO DEDUCIBLE', 'ERROR')):
            continue

        rfc   = f.get('rfc_em', '').strip().upper() or 'SIN RFC'
        razon = f.get('razon_em', '').strip() or rfc

        if rfc not in proveedores:
            proveedores[rfc] = {
                'razon': razon,
                'rfc':   rfc,
                'sub2':  0.0,
                'iva16': 0.0,
                'sub0':  0.0,
                'total': 0.0,
            }
        p = proveedores[rfc]
        # Usar float() sin round() — acumular con máxima precisión
        p['sub2']  += float(f.get('sub2',  0.0) or 0.0)
        p['iva16'] += float(f.get('iva16', 0.0) or 0.0)
        p['sub0']  += float(f.get('sub0',  0.0) or 0.0)
        p['total'] += float(f.get('total', 0.0) or 0.0)

    if not proveedores:
        return ''

    # Redondear acumulados a 2 decimales (solo al final, no por fila)
    for p in proveedores.values():
        p['sub2']  = round(p['sub2'],  2)
        p['iva16'] = round(p['iva16'], 2)
        p['sub0']  = round(p['sub0'],  2)
        p['total'] = round(p['total'], 2)

    # Ordenar por razón social
    filas_diot = sorted(proveedores.values(), key=lambda x: x['razon'])

    # Totales
    tot_sub2  = sum(p['sub2']  for p in filas_diot)
    tot_iva16 = sum(p['iva16'] for p in filas_diot)
    tot_sub0  = sum(p['sub0']  for p in filas_diot)
    tot_total = sum(p['total'] for p in filas_diot)

    def m(v):
        return '${:,.2f}'.format(v) if v else '-'

    tpl_row = ReportTemplateManager.get_sub_template('tpl-diot-row')
    tpl_tot = ReportTemplateManager.get_sub_template('tpl-diot-totrow')
    tpl_wrap = ReportTemplateManager.get_sub_template('tpl-diot-wrap')

    # Construir filas HTML
    rows = ''
    for i, p in enumerate(filas_diot):
        alt = ' diot-alt' if i % 2 == 0 else ''
        rows += tpl_row.format(
            alt=alt, razon=p['razon'], rfc=p['rfc'],
            sub2=m(p['sub2']), iva16=m(p['iva16']),
            sub0=m(p['sub0']), total=m(p['total'])
        ) + '\n'

    # Fila de totales
    rows += tpl_tot.format(
        tot_sub2=m(tot_sub2), tot_iva16=m(tot_iva16),
        tot_sub0=m(tot_sub0), tot_total=m(tot_total)
    ) + '\n'

    # Quitar el 16%% literal y usar 16% del HTML limpio, no hay .format params para la cabecera
    # solo pasamos mes y rows
    return tpl_wrap.format(mes=mes.upper(), rows=rows)


def generar_reporte(validado_path, mes_reporte=''):
    t0 = time.time()
    print('\n' + '='*60)
    print('  GENERADOR DE REPORTE FISCAL - ReaDesF1.8 v2.2')
    print('='*60)

    if not os.path.exists(validado_path):
        print('  ERROR: No encontrado: %s' % validado_path)
        return

    base     = re.sub(r'_validado\.xlsx$', '', validado_path, flags=re.IGNORECASE)
    base     = re.sub(r'\.xlsx$',          '', base,           flags=re.IGNORECASE)
    xlsx_out = base + '_reporte.xlsx'
    html_out = base + '_reporte.html'

    filas               = leer_validado(validado_path)
    reg_cod, reg_nombre = detectar_regimen(filas)
    idx_razones         = construir_indice_razones(filas)
    ppd_pend, cp01_a_ppd = detectar_ppd(filas)

    print('  Regimen: %s - %s' % (reg_cod, reg_nombre))
    print('  PPD pendientes (sin CP01): %d' % len(ppd_pend))
    print('  CP01 cruzados por RFC+total: %d' % len(cp01_a_ppd))

    stats = generar_excel(filas, ppd_pend, cp01_a_ppd, idx_razones,
                          xlsx_out, mes_reporte, reg_cod, reg_nombre)
    generar_html(filas, ppd_pend, cp01_a_ppd, idx_razones,
                 html_out, mes_reporte, stats, reg_cod, reg_nombre)

    t = time.time() - t0
    print('\n' + '='*60)
    print('  Listo en %.2fs' % t)
    print('  %s' % xlsx_out)
    print('  %s' % html_out)
    print('  DED:%d  NO:%d  PEND:%d  CP01:%d' % (
        stats['ded']+stats['efe'], stats['no_ded'], stats['pend'], stats['comp']))
    print('='*60 + '\n')


if __name__ == '__main__':
    import sys
    if len(sys.argv) >= 2:
        path = sys.argv[1]
        mes  = sys.argv[2] if len(sys.argv) >= 3 else ''
    else:
        carpeta = input('\n  Carpeta (Enter = Escritorio/GASTOS RESICO/2026/FEBRERO26): ').strip()
        if not carpeta:
            carpeta = os.path.join(os.path.expanduser('~'),'Desktop','GASTOS RESICO','2026','FEBRERO26')
        nombre = input('  Nombre del archivo _validado (sin .xlsx): ').strip()
        nombre = re.sub(r'\.xlsx$','',nombre,flags=re.IGNORECASE)
        if not nombre.endswith('_validado'):
            nombre += '_validado'
        mes  = input('  Mes del reporte (ej: FEBRERO 2026): ').strip()
        path = os.path.join(carpeta, nombre+'.xlsx')
    if not os.path.exists(path):
        print('\n  ERROR: No encontrado: %s' % path)
        carpeta_check = os.path.dirname(path)
        if os.path.exists(carpeta_check):
            excels = [f for f in os.listdir(carpeta_check) if f.endswith('.xlsx')]
            if excels:
                print('\n  Archivos disponibles:')
                for f in excels[:10]:
                    print('    . %s' % f)
    else:
        generar_reporte(path, mes_reporte=mes)
    input('\n>>> ENTER para salir...')

