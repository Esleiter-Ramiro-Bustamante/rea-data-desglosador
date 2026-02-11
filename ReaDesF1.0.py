import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import re
import time

# ==============================
# Configuración inicial
# ==============================

desktop_path = os.path.join(
    os.path.expanduser('~'),
    'Desktop/GASTOS RESICO/2026/ENERO26'
)

file_name = input("Ingrese el nombre del archivo Excel (sin extensión .xlsx): ")

file_name = file_name.strip()
file_name = re.sub(r'\.xlsx$', '', file_name, flags=re.IGNORECASE) + '.xlsx'
file_path = os.path.join(desktop_path, file_name)

# Iniciar cronómetro
tiempo_inicio = time.time()

# ==============================
# Estilos
# ==============================

blue_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
purple_fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
pink_fill = PatternFill(start_color='FF69B4', end_color='FF69B4', fill_type='solid')

result_style = Font(bold=True)
center_align = Alignment(horizontal='center')

# ==============================
# Funciones auxiliares
# ==============================

def extract_code(value):
    """Extrae el código de un valor que puede contener formato 'CODIGO - Descripción'"""
    if value and '-' in str(value):
        return str(value).split('-')[0].strip()
    return str(value).strip() if value else ''

def find_column(sheet, column_name):
    """Busca una columna por nombre en la primera fila"""
    for cell in sheet[1]:
        if cell.value and str(cell.value).strip().lower() == column_name.strip().lower():
            return cell.column
    return None

def create_column_if_missing(sheet, column_name, fill_color=blue_fill):
    """Crea una columna si no existe"""
    col = find_column(sheet, column_name)
    if col is None:
        last_col = sheet.max_column + 1
        sheet.cell(row=1, column=last_col, value=column_name)
        sheet.cell(row=1, column=last_col).fill = fill_color
        sheet.cell(row=1, column=last_col).alignment = center_align
        print(f"✅ Columna creada: {column_name}")
        return last_col
    return col

def es_gasolina(concepto):
    """
    Detecta si un concepto es relacionado a combustible/gasolina.
    Según Art. 27 LISR: combustibles para vehículos marítimos, aéreos y terrestres
    """
    if not concepto:
        return False
    palabras = [
        'gasolina', 'combustible', 'magna', 'premium',
        'diesel', 'diésel', 'gasohol', 'gasoil',
        'nafta', 'petrol', 'gas', 'energético',
        'turbosina', 'jet fuel', 'bunker'
    ]
    concepto_lower = concepto.lower()
    return any(p in concepto_lower for p in palabras)

def es_producto_dulce(concepto):
    """
    Detecta si un concepto es relacionado a productos dulces/botanas con IEPS 8%.
    Ejemplos: pan dulce, galletas, chocolates, botanas, papas, etc.
    """
    if not concepto:
        return False
    palabras = [
        # Pan dulce
        'pan', 'roles', 'conchas', 'mantecadas', 'donas', 'panque',
        'gansito', 'pinguinos', 'submarinos', 'chocorol', 'principe',
        'pastisetas', 'canelitas', 'polvorones', 'triki', 'duo',
        
        # Chocolates y dulces
        'chocolate', 'bon o bon', 'hershey', 'reese', 'kit kat',
        
        # Botanas
        'papas', 'chips', 'sabritas', 'doritos', 'cheetos', 'ruffles',
        'barcel', 'takis', 'hot nuts', 'cacahuates', 'kiyakis',
        'runners', 'churrumais', 'tostitos', 'fritos',
        
        # Galletas
        'galletas', 'oreo', 'emperador', 'marías', 'animalitos',
        'chokis', 'principe', 'sponch', 'barrita'
    ]
    concepto_lower = concepto.lower()
    return any(p in concepto_lower for p in palabras)

# ==============================
# Reglas de deducibilidad
# ==============================

USOS_DEDUCIBLES = ['G01', 'G02', 'G03']
METODOS_VALIDOS = ['PUE', 'PPD']
FORMAS_VALIDAS = ['01', '02', '03', '04', '28']
FORMAS_ELECTRONICAS = ['02', '03', '04', '28']
USO_CFDI_VERDE = "G02 - Devoluciones, descuentos o bonificaciones"
REGIMENES_FACILIDAD_COMBUSTIBLE = ['626']

# ==============================
# Abrir archivo Excel
# ==============================

try:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    print(f"✅ Archivo cargado: {file_path}")
    print(f"📊 Filas totales: {sheet.max_row - 1}")
except Exception as e:
    print(f"❌ Error al abrir el archivo: {e}")
    raise SystemExit(1)

# ==============================
# Columnas requeridas
# ==============================

required_columns = {
    'SubTotal': 'SubTotal',
    'Descuento': 'Descuento',
    'IVA Trasladado 0%': 'IVA Trasladado 0%',
    'IVA Exento': 'IVA Exento',
    'IVA Trasladado 16%': 'IVA Trasladado 16%',
    'Total': 'Total',
    'Uso CFDI': 'Uso CFDI',
    'Metodo pago': 'Metodo pago',
    'Forma pago': 'Forma pago',
    'Regimen receptor': 'Regimen receptor',
    'Razon emisor': 'Razon emisor',
    'Conceptos': 'Conceptos'
}

print("🔍 Buscando / creando columnas base...")
columns = {}

for key, col_name in required_columns.items():
    columns[key] = create_column_if_missing(sheet, col_name)

# ==============================
# OPTIMIZACIÓN: Pre-indexar columnas IEPS
# ==============================

print("🚀 Optimizando búsquedas de columnas IEPS...")

# Inicializar variables IEPS
ieps_gasolina_encontrado = False
ieps_8_encontrado = False
ieps_no_desglosado_encontrado = False
columnas_ieps = []

# Buscar TODAS las columnas IEPS
for col in range(1, sheet.max_column + 1):
    header = sheet.cell(row=1, column=col).value
    if header and "IEPS" in str(header).upper():
        columnas_ieps.append(col)
        header_str = str(header).strip()
        
        # Clasificar tipo de IEPS
        if '8%' in header_str or '8 %' in header_str:
            # IEPS 8% (dulces/botanas)
            columns['IEPS Trasladado 8%'] = col
            ieps_8_encontrado = True
            print(f"✅ IEPS 8% (dulces): {get_column_letter(col)}")
            
        elif 'No Desglosado' in header_str:
            # IEPS No Desglosado (puede ser gasolina)
            columns['IEPS Trasladado No Desglosado'] = col
            ieps_no_desglosado_encontrado = True
            print(f"✅ IEPS No Desglosado: {get_column_letter(col)}")
            
        else:
            # IEPS genérico (asumimos gasolina si no es 8%)
            columns['IEPS Trasladado'] = col
            ieps_gasolina_encontrado = True
            print(f"✅ IEPS Trasladado (gasolina): {get_column_letter(col)}")

print(f"⚡ {len(columnas_ieps)} columnas IEPS pre-indexadas")

# ==============================
# Columna Efecto (si existe)
# ==============================

efecto_col = find_column(sheet, 'Efecto')
if efecto_col:
    print(f"✅ Columna Efecto: {get_column_letter(efecto_col)}")

# ==============================
# Crear columna de Razón No Deducible
# ==============================

razon_no_ded_col = create_column_if_missing(sheet, 'Razón No Deducible', red_fill)

# ==============================
# Inicializar columnas numéricas
# ==============================

print("📝 Inicializando columnas numéricas...")

for row in range(2, sheet.max_row + 1):
    for col_key in ['IVA Trasladado 16%', 'IVA Trasladado 0%', 'IVA Exento', 'Descuento']:
        cell = sheet.cell(row=row, column=columns[col_key])
        if cell.value is None:
            cell.value = 0
        cell.number_format = "0.00"

# ==============================
# Crear columnas de cálculo
# ==============================

last_column = sheet.max_column

headers = [
    'SUB1-16%',
    'SUB0%',
    'SUB2-16%',
    'IVA ACREDITABLE 16%',
    'C IVA',
    'T2',
    'Comprobación T2',
    'Deducible'
]

print("🧮 Creando columnas de cálculo...")

for i, header in enumerate(headers, start=1):
    cell = sheet.cell(row=1, column=last_column + i, value=header)
    cell.fill = blue_fill
    cell.alignment = center_align

sub1_col = last_column + 1
sub0_col = last_column + 2
sub2_col = last_column + 3
iva_acred_col = last_column + 4
c_iva_col = last_column + 5
t2_col = last_column + 6
comprob_col = last_column + 7
deducible_col = last_column + 8

# ==============================
# OPTIMIZACIÓN: Cachear letras de columnas
# ==============================

print("📝 Cacheando referencias de columnas...")

col_letters = {
    'SubTotal': get_column_letter(columns['SubTotal']),
    'Descuento': get_column_letter(columns['Descuento']),
    'IVA16': get_column_letter(columns['IVA Trasladado 16%']),
    'IVA0': get_column_letter(columns['IVA Trasladado 0%']),
    'IVAExento': get_column_letter(columns['IVA Exento']),
    'Total': get_column_letter(columns['Total']),
    'sub1': get_column_letter(sub1_col),
    'sub0': get_column_letter(sub0_col),
    'sub2': get_column_letter(sub2_col),
    'iva_acred': get_column_letter(iva_acred_col),
    't2': get_column_letter(t2_col),
}

# Agregar letras de columnas IEPS
if ieps_8_encontrado:
    col_letters['IEPS8'] = get_column_letter(columns['IEPS Trasladado 8%'])
if ieps_gasolina_encontrado:
    col_letters['IEPS_GAS'] = get_column_letter(columns['IEPS Trasladado'])
if ieps_no_desglosado_encontrado:
    col_letters['IEPS_ND'] = get_column_letter(columns['IEPS Trasladado No Desglosado'])

# ==============================
# Variables de control
# ==============================

# Contadores IEPS
ieps_gasolina_procesados = 0
ieps_8_procesados = 0
ieps_no_desglosado_procesados = 0

# Contadores gasolina
gasolina_con_ieps = 0
gasolina_sin_ieps = 0
gasolina_efectivo = 0
gasolina_electronico = 0

# Contadores dulces
dulces_con_ieps8 = 0
dulces_sin_ieps8 = 0

# Contadores generales
uso_s01_count = 0
efectivo_mayor_2000 = 0

print("🔧 Iniciando procesamiento optimizado de filas...")
print("=" * 80)

# ==============================
# LOOP PRINCIPAL OPTIMIZADO
# ==============================

total_filas = sheet.max_row - 1

for row in range(2, sheet.max_row + 1):
    
    # Mostrar progreso cada 100 filas
    if (row - 1) % 100 == 0 or row == sheet.max_row:
        progreso = ((row - 1) / total_filas) * 100
        print(f"📊 Procesando: {row - 1}/{total_filas} facturas ({progreso:.1f}%)")
    
    # ==============================
    # Leer celdas necesarias UNA VEZ
    # ==============================
    
    row_data = {
        'concepto': str(sheet.cell(row=row, column=columns['Conceptos']).value or ''),
        'total': float(sheet.cell(row=row, column=columns['Total']).value or 0),
        'uso_cfdi': sheet.cell(row=row, column=columns['Uso CFDI']).value,
        'metodo_pago': sheet.cell(row=row, column=columns['Metodo pago']).value,
        'forma_pago': sheet.cell(row=row, column=columns['Forma pago']).value,
        'regimen': sheet.cell(row=row, column=columns['Regimen receptor']).value,
    }
    
    # Detectar tipo de producto (una sola vez)
    es_gasolina_concepto = es_gasolina(row_data['concepto'])
    es_dulce_concepto = es_producto_dulce(row_data['concepto'])
    
    # ==============================
    # Buscar valores IEPS (OPTIMIZADO)
    # ==============================
    
    ieps_8_val = 0.0
    ieps_gas_val = 0.0
    ieps_nd_val = 0.0
    
    # Leer valores de IEPS existentes
    if ieps_8_encontrado:
        ieps_8_val = float(sheet.cell(row=row, column=columns['IEPS Trasladado 8%']).value or 0)
    
    if ieps_gasolina_encontrado:
        ieps_gas_val = float(sheet.cell(row=row, column=columns['IEPS Trasladado']).value or 0)
    
    if ieps_no_desglosado_encontrado:
        ieps_nd_val = float(sheet.cell(row=row, column=columns['IEPS Trasladado No Desglosado']).value or 0)
    
    # ==============================
    # LÓGICA DE SUB1-16% SEGÚN TIPO DE IEPS
    # ==============================
    
    # CASO 1: Productos con IEPS 8% (dulces, botanas, pan)
    if ieps_8_val > 0:
        # SUB1 = (SubTotal - Descuento) + IEPS 8%
        formula_sub1 = f"=({col_letters['SubTotal']}{row}-{col_letters['Descuento']}{row})+{col_letters['IEPS8']}{row}"
        
        # Marcar concepto como dulce con IEPS
        if es_dulce_concepto:
            sheet.cell(row=row, column=columns['Conceptos']).fill = pink_fill
            dulces_con_ieps8 += 1
        
        ieps_8_procesados += 1
        
    # CASO 2: Gasolina con IEPS (va a IVA 0%)
    elif es_gasolina_concepto and (ieps_gas_val > 0 or ieps_nd_val > 0):
        # IEPS de gasolina → copiar a IVA 0%
        ieps_gasolina = ieps_gas_val if ieps_gas_val > 0 else ieps_nd_val
        sheet.cell(row=row, column=columns['IVA Trasladado 0%'], value=ieps_gasolina)
        sheet.cell(row=row, column=columns['IVA Trasladado 0%']).fill = orange_fill
        
        # SUB1 normal (sin sumar IEPS)
        formula_sub1 = f"=({col_letters['SubTotal']}{row}-{col_letters['Descuento']}{row})"
        
        # Marcar concepto como gasolina con IEPS
        sheet.cell(row=row, column=columns['Conceptos']).fill = blue_fill
        gasolina_con_ieps += 1
        ieps_gasolina_procesados += 1
        
    # CASO 3: Sin IEPS (normal)
    else:
        # SUB1 normal
        formula_sub1 = f"=({col_letters['SubTotal']}{row}-{col_letters['Descuento']}{row})"
        
        # Marcar si es gasolina sin IEPS (alerta)
        if es_gasolina_concepto:
            sheet.cell(row=row, column=columns['Conceptos']).fill = orange_fill
            gasolina_sin_ieps += 1
        elif es_dulce_concepto:
            dulces_sin_ieps8 += 1
    
    # Escribir fórmula SUB1
    sheet.cell(row=row, column=sub1_col, value=formula_sub1)
    sheet.cell(row=row, column=sub1_col).number_format = "0.00"
    sheet.cell(row=row, column=sub1_col).font = result_style
    
    # ==============================
    # SUB0% = IVA 0% + IVA Exento
    # ==============================
    
    sheet.cell(
        row=row,
        column=sub0_col,
        value=f"={col_letters['IVA0']}{row}+{col_letters['IVAExento']}{row}"
    )
    sheet.cell(row=row, column=sub0_col).number_format = "0.00"
    sheet.cell(row=row, column=sub0_col).font = result_style
    
    # ==============================
    # SUB2-16% = SUB1 - SUB0 (SIEMPRE)
    # ==============================
    
    sheet.cell(
        row=row,
        column=sub2_col,
        value=f"={col_letters['sub1']}{row}-{col_letters['sub0']}{row}"
    )
    sheet.cell(row=row, column=sub2_col).number_format = "0.00"
    sheet.cell(row=row, column=sub2_col).font = result_style
    
    # ==============================
    # IVA ACREDITABLE 16% = SUB2 * 0.16
    # ==============================
    
    sheet.cell(
        row=row,
        column=iva_acred_col,
        value=f"={col_letters['sub2']}{row}*0.16"
    )
    sheet.cell(row=row, column=iva_acred_col).number_format = "0.00"
    sheet.cell(row=row, column=iva_acred_col).font = result_style
    
    # Validación visual (para colores)
    try:
        subtotal_val = float(sheet.cell(row=row, column=columns['SubTotal']).value or 0)
        descuento_val = float(sheet.cell(row=row, column=columns['Descuento']).value or 0)
        iva0_val = float(sheet.cell(row=row, column=columns['IVA Trasladado 0%']).value or 0)
        iva_exento_val = float(sheet.cell(row=row, column=columns['IVA Exento']).value or 0)
        iva16_val = float(sheet.cell(row=row, column=columns['IVA Trasladado 16%']).value or 0)
        
        # Calcular según el tipo de IEPS
        if ieps_8_val > 0:
            sub1_calc = subtotal_val - descuento_val + ieps_8_val
        else:
            sub1_calc = subtotal_val - descuento_val
        
        sub0_calc = iva0_val + iva_exento_val
        sub2_calc = sub1_calc - sub0_calc
        iva_acred_calc = round(sub2_calc * 0.16, 2)
        
        if abs(iva_acred_calc - iva16_val) < 0.01:
            sheet.cell(row=row, column=iva_acred_col).fill = green_fill
            sheet.cell(row=row, column=columns['IVA Trasladado 16%']).fill = green_fill
    except:
        pass
    
    # ==============================
    # C IVA = IVA Acreditable - IVA 16%
    # ==============================
    
    sheet.cell(
        row=row,
        column=c_iva_col,
        value=f"={col_letters['iva_acred']}{row}-{col_letters['IVA16']}{row}"
    )
    sheet.cell(row=row, column=c_iva_col).number_format = "0.00"
    sheet.cell(row=row, column=c_iva_col).font = result_style
    
    # ==============================
    # T2 = SUB2 + SUB0 + IVA16
    # ==============================
    
    sheet.cell(
        row=row,
        column=t2_col,
        value=f"={col_letters['sub2']}{row}+{col_letters['sub0']}{row}+{col_letters['IVA16']}{row}"
    )
    sheet.cell(row=row, column=t2_col).number_format = "0.00"
    sheet.cell(row=row, column=t2_col).font = result_style
    
    # ==============================
    # Comprobación T2 = Total - T2
    # ==============================
    
    sheet.cell(
        row=row,
        column=comprob_col,
        value=f"={col_letters['Total']}{row}-{col_letters['t2']}{row}"
    )
    sheet.cell(row=row, column=comprob_col).number_format = "0.00"
    sheet.cell(row=row, column=comprob_col).font = result_style
    
    # ==============================
    # Formateo y validaciones
    # ==============================
    
    regimen = extract_code(row_data['regimen'])
    
    if regimen == '626':
        sheet.cell(row=row, column=columns['Regimen receptor']).fill = blue_fill
    elif regimen == '612':
        sheet.cell(row=row, column=columns['Regimen receptor']).fill = purple_fill
    else:
        sheet.cell(row=row, column=columns['Regimen receptor']).fill = orange_fill
        sheet.cell(row=row, column=columns['Razon emisor']).fill = orange_fill
    
    uso_cfdi = extract_code(row_data['uso_cfdi']).upper() if row_data['uso_cfdi'] else ''
    
    if row_data['uso_cfdi'] and str(row_data['uso_cfdi']).strip() == USO_CFDI_VERDE:
        sheet.cell(row=row, column=columns['Uso CFDI']).fill = green_fill
    
    if uso_cfdi == 'S01':
        sheet.cell(row=row, column=columns['Uso CFDI']).fill = red_fill
        uso_s01_count += 1
    
    es_egreso = False
    if efecto_col:
        efecto_val = sheet.cell(row=row, column=efecto_col).value
        if efecto_val and str(efecto_val).strip().upper() in ['EGRESO', 'E']:
            es_egreso = True
    
    # ==============================
    # VALIDACIÓN DE DEDUCIBILIDAD
    # ==============================
    
    metodo_pg = extract_code(row_data['metodo_pago']).upper()
    forma_pg = extract_code(row_data['forma_pago']).upper()
    
    es_deducible = True
    razones_rechazo = []
    
    if uso_cfdi not in USOS_DEDUCIBLES:
        es_deducible = False
        razones_rechazo.append(f"Uso CFDI {uso_cfdi} no deducible")
    
    if metodo_pg not in METODOS_VALIDOS:
        es_deducible = False
        razones_rechazo.append(f"Método {metodo_pg} inválido")
    
    # VALIDACIÓN GASOLINA
    if es_gasolina_concepto:
        if forma_pg == '01':
            if regimen in REGIMENES_FACILIDAD_COMBUSTIBLE and row_data['total'] <= 2000:
                gasolina_efectivo += 1
                sheet.cell(row=row, column=columns['Forma pago']).fill = yellow_fill
                razones_rechazo.append("RESICO: Gasolina efectivo ≤$2,000 (facilidad)")
            else:
                es_deducible = False
                gasolina_efectivo += 1
                razones_rechazo.append("Gasolina NO deducible en efectivo")
                sheet.cell(row=row, column=columns['Forma pago']).fill = red_fill
        else:
            gasolina_electronico += 1
            if forma_pg not in FORMAS_ELECTRONICAS:
                es_deducible = False
                razones_rechazo.append(f"Gasolina: forma de pago {forma_pg} inválida")
    else:
        if forma_pg == '01' and row_data['total'] > 2000:
            es_deducible = False
            efectivo_mayor_2000 += 1
            razones_rechazo.append("Efectivo >$2,000 NO deducible")
            if efecto_col:
                sheet.cell(row=row, column=efecto_col).fill = red_fill
        elif forma_pg not in FORMAS_VALIDAS:
            es_deducible = False
            razones_rechazo.append(f"Forma de pago {forma_pg} inválida")
    
    # Escribir resultado
    ded_cell = sheet.cell(row=row, column=deducible_col, value="SI" if es_deducible else "NO")
    
    if es_deducible:
        ded_cell.fill = blue_fill if es_egreso else green_fill
    else:
        ded_cell.fill = red_fill
    
    ded_cell.font = Font(bold=True, color='FFFFFF')
    ded_cell.alignment = center_align
    
    razon_cell = sheet.cell(row=row, column=razon_no_ded_col)
    if razones_rechazo:
        razon_cell.value = " | ".join(razones_rechazo)
        if not es_deducible:
            razon_cell.fill = red_fill
            razon_cell.font = Font(bold=True, color='FFFFFF')
        else:
            razon_cell.fill = yellow_fill
            razon_cell.font = Font(bold=True)
    else:
        razon_cell.value = "Cumple requisitos"
        razon_cell.fill = green_fill
        razon_cell.font = Font(color='006100')

# ==============================
# Ajustar ancho de columnas
# ==============================

for col in range(last_column + 1, last_column + len(headers) + 1):
    sheet.column_dimensions[get_column_letter(col)].width = 15

sheet.column_dimensions[get_column_letter(razon_no_ded_col)].width = 40

# ==============================
# Guardar archivo
# ==============================

try:
    base_name = re.sub(r'\.xlsx$', '', file_name, flags=re.IGNORECASE)
    output_name = os.path.join(desktop_path, f"{base_name}_validado.xlsx")
    workbook.save(output_name)
    
    tiempo_fin = time.time()
    tiempo_total = tiempo_fin - tiempo_inicio
    velocidad = total_filas / tiempo_total if tiempo_total > 0 else 0

    print("\n" + "=" * 80)
    print("✅ PROCESO COMPLETADO EXITOSAMENTE - ReaDesF1.0")
    print("=" * 80)
    print(f"📂 Archivo guardado como: {output_name}")
    print(f"📊 Total de filas procesadas: {total_filas}")
    print(f"⏱️  Tiempo total: {tiempo_total:.2f} segundos")
    print(f"⚡ Velocidad: {velocidad:.0f} facturas/segundo")

    print("\n📌 RESUMEN DE VALIDACIÓN:")
    print(f"\n🍬 IEPS 8% (Dulces/Botanas):")
    print(f"  • Facturas con IEPS 8% procesadas: {ieps_8_procesados}")
    print(f"  • Dulces detectados con IEPS 8%: {dulces_con_ieps8}")
    print(f"  • Dulces sin IEPS 8%: {dulces_sin_ieps8}")
    
    print(f"\n⛽ IEPS Gasolina:")
    print(f"  • Facturas con IEPS gasolina: {ieps_gasolina_procesados}")
    print(f"  • Gasolina con IEPS: {gasolina_con_ieps}")
    print(f"  • Gasolina sin IEPS: {gasolina_sin_ieps}")
    print(f"  • Gasolina pagada en efectivo: {gasolina_efectivo}")
    print(f"  • Gasolina pagada electrónicamente: {gasolina_electronico}")
    
    print(f"\n📋 General:")
    print(f"  • IEPS No Desglosado procesados: {ieps_no_desglosado_procesados}")
    print(f"  • Usos S01 (sin efectos fiscales): {uso_s01_count}")
    print(f"  • Gastos normales efectivo >$2,000: {efectivo_mayor_2000}")

    print("\n⚠️  REGLAS APLICADAS:")
    print("  1. Art. 27, fracc. III LISR: Efectivo >$2,000 NO deducible")
    print("  2. Art. 27, fracc. III LISR: Gasolina NUNCA en efectivo")
    print("     (Excepción: RESICO hasta $2,000)")
    print("  3. IEPS 8% se suma a SUB1: SUB1 = (SubTotal - Descuento) + IEPS 8%")
    print("  4. IEPS Gasolina va a IVA 0%")
    print("  5. Uso CFDI válido: G01, G02, G03")
    print("  6. Método de pago válido: PUE, PPD")

    print("\n🎨 CÓDIGO DE COLORES:")
    print("  🟦 AZUL: Gasolina con IEPS / Régimen 626")
    print("  🟩 VERDE: Deducible / IVA correcto")
    print("  🟥 ROJO: NO deducible")
    print("  🟪 MORADO: Régimen 612")
    print("  🟧 NARANJA: Gasolina sin IEPS / Otros regímenes")
    print("  🟨 AMARILLO: Advertencia")
    print("  🌸 ROSA: Dulces/Botanas con IEPS 8%")

    print("=" * 80)

except Exception as e:
    print("\n❌ Error al guardar el archivo:")
    print(e)
