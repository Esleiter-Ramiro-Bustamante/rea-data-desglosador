"""
generador_reporte.py — Modulo de Reporte Fiscal v2.3
ReaDesF1.8

CORRECCIONES v2.2:
  - uuid_rel extraido de columna COMENTARIOS cuando viene vacio
    (los motores escriben "Complemento parcialidad 1 - factura 5D1EDA96AB6E saldo $0")
  - Cruce PPD<->complemento robusto: por sufijo 12 chars hex + UUID completo
  - PPD ya cubierto NO aparece como PENDIENTE
  - UUID completo visible en HTML con scroll horizontal (no truncado)
  - Ceros mostrados como guion en HTML (no "0.0")
  - classify() unifica color+css+filtro en 1 sola funcion (1 solo .upper() por fila)
  - idx_razones construido UNA vez -> O(1) por busqueda
  - is_cp calculado UNA vez por fila
  - StringIO para HTML (no concatenar string gigante en RAM)
  - Regimen detectado dinamicamente por cliente (612/626/otros)
"""

import os
import re
import time
from io import StringIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

USOS_VALIDOS    = {'G01', 'G02', 'G03'}
METODOS_VALIDOS = {'PUE', 'PPD'}
FORMAS_VALIDAS  = {'01', '02', '03', '04', '28'}
LIMITE_EFECTIVO = 2000.0

_RE_UUID_FULL  = re.compile(
    r'[0-9A-Fa-f]{8}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{4}-[0-9A-Fa-f]{12}',
    re.IGNORECASE
)
_RE_UUID_SHORT = re.compile(r'\b([0-9A-Fa-f]{12})\b')


def extraer_codigo(val, n=3):
    s = str(val or '').strip()
    if '-' in s:
        return s.split('-')[0].strip()[:n].upper()
    return s[:n].upper()

def safe_str(v):
    return str(v).strip() if v is not None else ''

def norm_uuid(u):
    return re.sub(r'[-\s]', '', str(u or '')).lower()

def fmt_num(v):
    try:
        f = float(v)
        return '$%s' % '{:,.2f}'.format(f) if f != 0 else '-'
    except Exception:
        return '-'

def extraer_uuids_de_texto(texto):
    """Extrae UUIDs de texto libre. Busca UUID completo primero, luego sufijo 12 hex."""
    if not texto:
        return []
    completos = _RE_UUID_FULL.findall(texto)
    if completos:
        return completos
    return _RE_UUID_SHORT.findall(texto)

def classify(estatus):
    """Retorna (bg_hex, fg_hex, css_cls, flt_cls). Un solo .upper() por fila."""
    eu = estatus.upper()
    if 'COMPLEMENTO' in eu: return ('D5C6E0','4A0080','complemento','complemento')
    if '16 Y 0'      in eu: return ('BDD7EE','1F497D','mix','ded')
    if '16%'         in eu:
        return ('C6EFCE','1B5E20','ded16','efe' if 'EFE' in eu else 'ded')
    if '0%'          in eu:
        return ('FFF3CD','856404','ded0','efe' if 'EFE' in eu else 'ded')
    if 'EGRESO'      in eu: return ('E2CFED','6A1B9A','egreso','egreso')
    if 'PENDIENTE'   in eu: return ('E0E0E0','37474F','pendiente','pendiente')
    if 'ERROR'       in eu: return ('FF4444','FFFFFF','no-ded','no-ded')
    if 'NO DED'      in eu: return ('FFC7CE','9C0006','no-ded','no-ded')
    return                         ('FFFFFF','000000','no-ded','no-ded')


def leer_validado(path):
    """Lee _validado.xlsx. Extrae uuid_rel de COMENTARIOS si viene vacio."""
    wb    = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = wb.active
    hraw  = {}
    hnorm = {}
    for c in sheet[1]:
        if c.value is not None:
            raw = str(c.value).strip()
            hraw[c.column - 1] = raw
            hnorm[raw.lower()] = c.column - 1

    def gc(name):
        n = name.strip().lower()
        if n in hnorm:
            return hnorm[n]
        for k, v in hnorm.items():
            if n in k or k in n:
                return v
        return None

    i_uuid      = gc('uuid') or gc('folio fiscal')
    i_uuid_rel  = gc('uuids relacionados') or gc('uuid relacionado') or gc('folio fiscal relacionado')
    i_fecha     = gc('fecha certificacion') or gc('fecha emision') or gc('fecha')
    i_razon_em  = gc('razon emisor') or gc('nombre emisor') or gc('razon social')
    i_razon_rec = gc('razon receptor') or gc('nombre receptor')
    i_regimen   = gc('regimen receptor') or gc('regimen')
    i_metodo    = gc('metodo pago') or gc('metodo de pago')
    i_forma     = gc('forma pago') or gc('forma de pago')
    i_uso       = gc('uso cfdi')
    i_subtotal  = gc('subtotal')
    i_descuento = gc('descuento')
    i_iva16     = gc('iva trasladado 16%') or gc('iva 16%')
    i_iva0      = gc('iva trasladado 0%')  or gc('iva 0%')
    i_iva_ex    = gc('iva exento')
    i_ieps      = gc('ieps trasladado') or gc('ieps')
    i_ieps_nd   = gc('ieps trasladado no desglosado')
    i_ieps_g    = gc('ieps trasladado')
    i_ieps3     = gc('ieps trasladado 3%')
    i_total     = gc('total')
    i_conceptos = gc('conceptos') or gc('descripcion')
    i_complem   = gc('complementos')
    i_efecto    = gc('efecto')
    i_sub0      = gc('sub0%')    or gc('sub0')
    i_sub2      = gc('sub2-16%') or gc('sub2')
    i_sub1      = gc('sub1-16%') or gc('sub1')
    i_coment    = gc('comentarios') or gc('comentario') or gc('observaciones')
    i_rfc_em    = gc('rfc emisor') or gc('rfc')

    def rv(row, idx):
        return row[idx] if (idx is not None and idx < len(row)) else None

    def rnum(row, idx):
        try:
            v = rv(row, idx)
            return float(v) if v is not None else 0.0
        except Exception:
            return 0.0

    filas = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        st     = rnum(row, i_subtotal)
        dc     = rnum(row, i_descuento)
        iva16  = rnum(row, i_iva16)
        iva0   = rnum(row, i_iva0)
        iva_ex = rnum(row, i_iva_ex)
        total  = rnum(row, i_total)
        ieps_nd = rnum(row, i_ieps_nd)
        ieps_g  = rnum(row, i_ieps_g)
        ieps3   = rnum(row, i_ieps3) if i_ieps3 is not None else 0.0

        # ── sub0: misma lógica que el motor, usando valores crudos ────
        # No leer la columna SUB0% (es fórmula Excel sin caché en Python)
        # Gasolina con IEPS: iva0 ya tiene el IEPS, iva_ex es el mismo valor
        # → sumarlos duplicaría. Solo usar iva0.
        ieps_activo = ieps_nd if ieps_nd > 0 else ieps_g
        # Gasolina con IEPS: motor copia IEPS a IVA_0% e IVA_Exento
        # Si iva0 == iva_ex > 0 → no duplicar, usar solo iva0
        if iva0 > 0 and abs(iva0 - iva_ex) < 0.02:
            sub0 = round(iva0, 2)
        else:
            sub0 = round(iva0 + iva_ex, 2)

        # SUB2: SIEMPRE calcular desde valores crudos del CFDI
        # IEPS 3% telefonía se suma a SUB1 igual que IEPS 8% dulces
        # Formula garantizada: SUB1 = SubTotal - Descuento + IEPS3
        #                      SUB2 = SUB1 - SUB0
        sub1 = round(st - dc + ieps3, 2)
        sub2 = round(max(sub1 - sub0, 0), 2)

        uuid_rel_raw = safe_str(rv(row, i_uuid_rel))
        comentarios  = safe_str(rv(row, i_coment))
        conceptos    = safe_str(rv(row, i_conceptos))
        rfc_em       = safe_str(rv(row, i_rfc_em))

        # Si uuid_rel vacio, extraer de COMENTARIOS o CONCEPTOS
        uuid_rel = uuid_rel_raw
        if not uuid_rel:
            for fuente in [comentarios, conceptos]:
                encontrados = extraer_uuids_de_texto(fuente)
                if encontrados:
                    uuid_rel = encontrados[0]
                    break

        filas.append({
            'uuid':         safe_str(rv(row, i_uuid)),
            'uuid_rel':     uuid_rel,
            'fecha':        rv(row, i_fecha),
            'razon_em':     safe_str(rv(row, i_razon_em)),
            'razon_rec':    safe_str(rv(row, i_razon_rec)),
            'regimen':      safe_str(rv(row, i_regimen)),
            'metodo':       safe_str(rv(row, i_metodo)),
            'forma':        safe_str(rv(row, i_forma)),
            'uso':          safe_str(rv(row, i_uso)),
            'subtotal':     round(st, 2),
            'descuento':    round(dc, 2),
            'iva16':        round(iva16, 2),
            'iva0':         round(iva0, 2),
            'iva_ex':       round(iva_ex, 2),
            'ieps':         rnum(row, i_ieps),
            'total':        round(total, 2),
            'sub0':         sub0,
            'sub2':         sub2,
            'conceptos':    conceptos,
            'complementos': safe_str(rv(row, i_complem)),
            'efecto':       safe_str(rv(row, i_efecto)),
            'comentarios':  comentarios,
            'rfc_em':       rfc_em,
        })

    wb.close()
    print('  OK %d filas leidas - %d columnas' % (len(filas), len(hraw)))
    return filas


NOMBRES_REGIMEN = {
    '612': 'Actividades Empresariales y Profesionales',
    '626': 'Regimen Simplificado de Confianza (RESICO)',
    '601': 'General de Ley Personas Morales',
    '606': 'Arrendamiento',
    '621': 'Incorporacion Fiscal',
    '616': 'Sin obligaciones fiscales',
}

def detectar_regimen(filas):
    conteo = {}
    for f in filas:
        cod = extraer_codigo(f.get('regimen', ''))
        if cod and cod.isdigit() and len(cod) == 3:
            conteo[cod] = conteo.get(cod, 0) + 1
    if not conteo:
        return '612', NOMBRES_REGIMEN['612']
    dom = max(conteo, key=conteo.get)
    return dom, NOMBRES_REGIMEN.get(dom, 'Regimen %s' % dom)


def es_complemento(f):
    uso   = f.get('uso', '').upper()
    efect = f.get('efecto', '').upper()
    conc  = f.get('conceptos', '').upper()
    comp  = f.get('complementos', '').upper()
    return ('CP01' in uso or efect == 'PAGO' or (conc == 'PAGO' and comp == 'PAGO'))

def construir_indice_razones(filas):
    """Dict {uuid_norm: razon} + {rfc: razon}. Incluye sufijo 12 chars."""
    idx = {}
    for f in filas:
        uuid = f.get('uuid', '')
        if not uuid:
            continue
        razon = f.get('razon_em', '')
        un    = norm_uuid(uuid)
        idx[un] = razon
        if len(un) >= 12:
            idx[un[-12:]] = razon
        # Indexar por RFC para cruce cuando uuid_rel esta vacio
        rfc = f.get('rfc_em', '').strip().upper()
        if rfc:
            idx['rfc:' + rfc] = razon
    return idx

def detectar_ppd(filas):
    """
    UUIDs PPD sin CP01. Cruce en 2 niveles:
      Nivel 1: uuid_rel explicito (cuando el motor lo llena)
      Nivel 2: RFC emisor + total identico (cuando uuid_rel esta vacio)
    Retorna (ppd_pendientes, mapa_cp01_a_uuid_ppd)
    """
    # Todos los PPD: {uuid_norm: fila_completa}
    ppd_map = {}
    for f in filas:
        if extraer_codigo(f.get('metodo', '')) == 'PPD' and f.get('uuid'):
            ppd_map[norm_uuid(f['uuid'])] = f

    cubiertos   = set()   # uuid_norm de PPDs con CP01
    cp01_a_ppd  = {}      # uuid_norm(cp01) -> uuid_original(ppd)

    for f in filas:
        if not es_complemento(f):
            continue

        # NIVEL 1: uuid_rel explicito en el _validado
        rel = f.get('uuid_rel', '').strip()
        if rel:
            rel_n    = norm_uuid(rel)
            rel_tail = rel_n[-12:] if len(rel_n) >= 12 else rel_n
            for uid_n, ppd_f in ppd_map.items():
                if uid_n == rel_n or uid_n.endswith(rel_tail):
                    cubiertos.add(uid_n)
                    cp01_a_ppd[norm_uuid(f['uuid'])] = ppd_f['uuid']
            continue

        # NIVEL 2: uuid_rel vacio → cruzar por RFC emisor + total igual
        rfc_cp01   = f.get('rfc_em', '').strip().upper()
        total_cp01 = round(float(f.get('total', 0)), 2)
        if not rfc_cp01:
            continue
        for uid_n, ppd_f in ppd_map.items():
            rfc_ppd   = ppd_f.get('rfc_em', '').strip().upper()
            total_ppd = round(float(ppd_f.get('total', 0)), 2)
            if rfc_cp01 == rfc_ppd and total_cp01 == total_ppd:
                cubiertos.add(uid_n)
                cp01_a_ppd[norm_uuid(f['uuid'])] = ppd_f['uuid']
                break  # un CP01 cubre un PPD

    ppd_pendientes = {ppd_map[n]['uuid'] for n in ppd_map if n not in cubiertos}
    return ppd_pendientes, cp01_a_ppd

def resolver_uuid_rel(f, idx_razones, cp01_a_ppd=None):
    """
    Retorna (display_12chars, razon_proveedor) para fila complemento.
    Busca en: 1) uuid_rel explicito  2) mapa cp01_a_ppd por RFC+total
    display = ultimos 12 chars del UUID del PPD relacionado (en MAYUSCULAS)
    """
    rel = f.get('uuid_rel', '').strip()

    # NIVEL 1: uuid_rel explicito
    if rel:
        rel_n    = norm_uuid(rel)
        rel_tail = rel_n[-12:] if len(rel_n) >= 12 else rel_n
        razon    = idx_razones.get(rel_n) or idx_razones.get(rel_tail, '')
        display  = rel_tail.upper()
        return display, razon

    # NIVEL 2: buscar en mapa cp01_a_ppd
    if cp01_a_ppd:
        cp_n    = norm_uuid(f.get('uuid', ''))
        ppd_uuid = cp01_a_ppd.get(cp_n, '')
        if ppd_uuid:
            ppd_n    = norm_uuid(ppd_uuid)
            rel_tail = ppd_n[-12:].upper() if len(ppd_n) >= 12 else ppd_n.upper()
            razon    = idx_razones.get(ppd_n) or idx_razones.get(ppd_n[-12:], '')
            if not razon:
                razon = f.get('razon_em', '')  # mismo proveedor
            return rel_tail, razon

    return '', ''


def calc_estatus(f, is_cp, ppd_pend, idx_razones):
    if is_cp:
        # ESTATUS = solo "COMPLEMENTO" limpio (sin monto, sin redundancia)
        # El detalle (uuid_rel, razon, saldo) va SOLO en columna OBS
        return 'COMPLEMENTO'

    uuid_n = norm_uuid(f.get('uuid', ''))
    if any(norm_uuid(p) == uuid_n for p in ppd_pend):
        return 'PENDIENTE'

    uso   = f.get('uso', '');   metodo = f.get('metodo', '')
    forma = f.get('forma', ''); total  = f.get('total', 0)
    sub2  = f.get('sub2', 0);  iva16  = f.get('iva16', 0)
    sub0  = f.get('sub0', 0)

    u  = extraer_codigo(uso)
    m  = extraer_codigo(metodo)
    fp = extraer_codigo(forma, 2)

    if u in ('S01', 'CN0'):
        return 'NO DEDUCIBLE'

    u_ok = u in USOS_VALIDOS
    m_ok = m in METODOS_VALIDOS
    # Forma 99 = "Por definir" → válida SOLO para PPD (es temporal hasta que llega el CP01)
    f_ok = fp in FORMAS_VALIDAS or (fp == '99' and m == 'PPD')

    if u_ok and m_ok and f_ok:
        if fp == '01' and total >= LIMITE_EFECTIVO:
            return 'NO DEDUCIBLE: Efectivo >= $2,000'
        if u == 'G02':
            return 'EGRESO'
        if   sub2 > 0 and iva16 > 0 and sub0 == 0: suf = '16%'
        elif sub2 > 0 and iva16 > 0 and sub0 > 0:  suf = '16 Y 0%'
        elif sub2 == 0 and iva16 == 0 and sub0 > 0: suf = '0%'
        else:                                        suf = 'NO DEDUCIBLE'
        return ('EFE ' if fp == '01' else 'DED ') + suf

    err = []
    if not u_ok: err.append('USO (%s) INVALIDO' % u)
    if not m_ok: err.append('METODO (%s) INVALIDO' % m)
    if not f_ok: err.append('FORMA (%s) INVALIDA' % fp)
    return 'ERROR: ' + ' | '.join(err)


def formula_estatus(rn):
    L = LIMITE_EFECTIVO
    return (
        '=IF(OR(LEFT(K%(r)s,3)="S01",LEFT(K%(r)s,3)="CN0"),"NO DEDUCIBLE",'
        'IF(AND(NOT(ISERROR(MATCH(LEFT(K%(r)s,3),{"G01","G02","G03"},0))),'
        'NOT(ISERROR(MATCH(LEFT(I%(r)s,3),{"PUE","PPD"},0))),'
        'NOT(ISERROR(MATCH(LEFT(J%(r)s,2),{"01","02","03","04","28"},0)))),'
        'IF(AND(LEFT(J%(r)s,2)="01",H%(r)s>=%(L)s),"NO DEDUCIBLE: Efectivo >= $2,000",'
        'IF(LEFT(K%(r)s,3)="G02","EGRESO",'
        'IF(LEFT(J%(r)s,2)="01",'
        '"EFE "&IF(AND(E%(r)s>0,F%(r)s>0,G%(r)s=0),"16%%",IF(AND(E%(r)s>0,F%(r)s>0,G%(r)s>0),"16 Y 0%%",IF(AND(E%(r)s=0,F%(r)s=0,G%(r)s>0),"0%%","NO DEDUCIBLE"))),'
        '"DED "&IF(AND(E%(r)s>0,F%(r)s>0,G%(r)s=0),"16%%",IF(AND(E%(r)s>0,F%(r)s>0,G%(r)s>0),"16 Y 0%%",IF(AND(E%(r)s=0,F%(r)s=0,G%(r)s>0),"0%%","NO DEDUCIBLE")))'
        '))),'
        '"ERROR: "&IF(ISERROR(MATCH(LEFT(K%(r)s,3),{"G01","G02","G03"},0)),"USO INVALIDO | ","")&'
        'IF(ISERROR(MATCH(LEFT(I%(r)s,3),{"PUE","PPD"},0)),"METODO INVALIDO | ","")&'
        'IF(ISERROR(MATCH(LEFT(J%(r)s,2),{"01","02","03","04","28"},0)),"FORMA INVALIDA","")))'
    ) % {'r': rn, 'L': L}


def generar_excel(filas, ppd_pend, cp01_a_ppd, idx_razones, out, mes='', reg_cod='612', reg_nombre=''):
    print('  Generando Excel...')
    wb  = openpyxl.Workbook()
    sh  = wb.active
    sh.title = 'Reporte Fiscal'

    mk  = lambda h: PatternFill(start_color=h, end_color=h, fill_type='solid')
    ct  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    lt  = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    th  = Side(style='thin',   color='CCCCCC')
    tk  = Side(style='medium', color='0A0A0A')
    bd  = Border(left=th, right=th, top=th, bottom=th)
    bdh = Border(left=tk, right=tk, top=tk, bottom=tk)

    NCOLS = 13
    sh.merge_cells('A1:%s1' % get_column_letter(NCOLS))
    t = sh['A1']
    t.value = 'RESUMEN GASTOS MENSUALES - %s  .  %s' % (reg_nombre.upper(), mes.upper())
    t.font  = Font(bold=True, size=12, name='Calibri')
    t.fill  = mk('D6EAF8'); t.alignment = ct; t.border = bdh
    sh.row_dimensions[1].height = 34

    sh.merge_cells('A2:D2')
    sh['A2'].value = 'PUNTOS DEDUCIBLES DEL SISTEMA'
    sh['A2'].font  = Font(bold=True, size=9)
    sh['A2'].fill  = mk('D5F5E3'); sh['A2'].alignment = ct
    for col, val in [('E','PUE'), ('G','01, 02, 03, 28'), ('I','G01 Y G03')]:
        c = sh['%s2' % col]
        c.value = val; c.font = Font(bold=True, size=9)
        c.fill  = mk('D5F5E3'); c.alignment = ct

    sh.merge_cells('A3:%s3' % get_column_letter(NCOLS))
    sh['A3'].value = 'Regimen %s - %s' % (reg_cod, reg_nombre)
    sh['A3'].font  = Font(italic=True, size=9, color='555555')
    sh['A3'].alignment = ct; sh.row_dimensions[3].height = 18

    RH = 4
    hdrs   = ['FOLIO FISCAL','UUID RELACIONADO','FECHA','RAZON SOCIAL',
              'SUBTOTAL 16%','IVA 16%','SUB 0%','TOTAL',
              'METODO DE PAGO','FORMA DE PAGO','USO CFDI',
              'ESTATUS','COMPLEMENTOS / OBSERVACIONES']
    widths = [44,44,19,32,15,13,13,15,30,26,30,28,60]

    for i, (h, w) in enumerate(zip(hdrs, widths), 1):
        c = sh.cell(row=RH, column=i, value=h)
        c.font = Font(bold=True, size=9, name='Calibri', color='1A1A2E')
        c.fill = mk('AED6F1'); c.alignment = ct; c.border = bdh
        sh.column_dimensions[get_column_letter(i)].width = w
    sh.row_dimensions[RH].height = 34
    sh.freeze_panes = 'A%d' % (RH+1)

    stats = {k:0 for k in ['total','ded','no_ded','pend','egreso','efe','comp',
                             'monto_ded','monto_no_ded','monto_pend']}
    ppd_norms = {norm_uuid(p) for p in ppd_pend}

    rn = RH + 1
    for f in filas:
        is_cp   = es_complemento(f)
        estatus = calc_estatus(f, is_cp, ppd_pend, idx_razones)
        bg, fg, _, _ = classify(estatus)
        eu = estatus.upper()

        if is_cp:
            display, razon = resolver_uuid_rel(f, idx_razones, cp01_a_ppd)
            saldo = f.get('total', 0)
            obs   = ('Complemento parcialidad 1 - factura %s' % display) if display else 'Complemento CP01'
            if razon: obs += ' (%s)' % razon[:30]
            obs  += ' saldo insoluto $%s' % '{:,.2f}'.format(saldo)
        elif norm_uuid(f.get('uuid','')) in ppd_norms:
            obs = 'PPD sin complemento CP01 - pasa al siguiente mes'
        else:
            obs = ''

        stats['total'] += 1
        if   'COMPLEMENTO' in eu: stats['comp']   += 1
        elif 'PENDIENTE'   in eu: stats['pend']   += 1; stats['monto_pend']   += f['total']
        elif 'NO DED' in eu or 'ERROR' in eu:
                                   stats['no_ded']+= 1; stats['monto_no_ded'] += f['total']
        elif 'EGRESO'      in eu: stats['egreso'] += 1
        elif 'EFE'         in eu: stats['efe']    += 1; stats['monto_ded']    += f['total']
        else:                      stats['ded']   += 1; stats['monto_ded']    += f['total']

        # CP01 → amarillo en _validado (igual que imagen de referencia)
        if is_cp:
            cf = 'FFFF00'  # Amarillo para CP01
        elif norm_uuid(f.get('uuid','')) in ppd_norms:
            cf = 'FFFF00'  # Amarillo tambien para PPD relacionado
        else:
            cf = 'FFFFFF' if (rn-RH-1)%2==0 else 'F8F9FA'

        # UUID relacionado: display de 12 chars del PPD
        display_rel, _ = resolver_uuid_rel(f, idx_razones, cp01_a_ppd) if is_cp else ('', '')

        vals = [f['uuid'], display_rel or f['uuid_rel'],f['fecha'],f['razon_em'],
                f['sub2'] or None,f['iva16'] or None,
                f['sub0'] or None,f['total'] or None,
                f['metodo'],f['forma'],f['uso'],None,obs]

        for ci, val in enumerate(vals, 1):
            c = sh.cell(row=rn, column=ci, value=val)
            c.border = bd
            is_num = ci in (5,6,7,8)
            c.font = Font(size=9, name='Calibri', bold=is_num, color='1A1A2E')
            c.alignment = lt if ci in (1,2,4,13) else ct
            if is_num: c.number_format = '"$"#,##0.00'
            if ci==12: c.fill=mk(bg); c.font=Font(bold=True,color=fg,size=9,name='Calibri')
            elif is_cp: c.fill=mk('EDE0F5')
            else: c.fill=mk(cf)

        ec = sh.cell(row=rn, column=12)
        ec.value = estatus if (is_cp or 'PENDIENTE' in eu) else formula_estatus(rn)
        ec.fill=mk(bg); ec.font=Font(bold=True,color=fg,size=9,name='Calibri')
        ec.alignment=ct; ec.border=bd
        sh.row_dimensions[rn].height = 28
        rn += 1

    sh.merge_cells('A%d:D%d' % (rn,rn))
    tc = sh.cell(row=rn, column=1, value='TOTALES')
    tc.font=Font(bold=True,size=11); tc.fill=mk('D6EAF8')
    tc.alignment=ct; tc.border=bdh
    for cl,ci in [('E',5),('F',6),('G',7),('H',8)]:
        c = sh.cell(row=rn, column=ci)
        c.value='=SUM(%s%d:%s%d)' % (cl,RH+1,cl,rn-1)
        c.number_format='"$"#,##0.00'; c.fill=mk('D6EAF8')
        c.border=bdh; c.font=Font(bold=True,size=10); c.alignment=ct

    wb.save(out)
    print('  Excel: %s' % out)
    return stats


_OPCIONES = [
    'DED 16%','DED 0%','DED 16 Y 0%','EFE 16%','EFE 0%','EFE 16 Y 0%',
    'EGRESO','NO DEDUCIBLE','NO DEDUCIBLE: Efectivo >= $2,000','PENDIENTE','COMPLEMENTO'
]

def generar_html(filas, ppd_pend, cp01_a_ppd, idx_razones, out, mes='', stats=None, reg_cod='612', reg_nombre=''):
    print('  Generando HTML...')
    s         = stats or {}
    total_f   = s.get('total', len(filas))
    ppd_norms = {norm_uuid(p) for p in ppd_pend}
    now_str   = datetime.now().strftime('%d/%m/%Y %H:%M')
    n_ded=s.get('ded',0); n_efe=s.get('efe',0); n_no_ded=s.get('no_ded',0)
    n_pend=s.get('pend',0); n_egreso=s.get('egreso',0); n_comp=s.get('comp',0)
    m_ded=s.get('monto_ded',0); m_no_ded=s.get('monto_no_ded',0); m_pend=s.get('monto_pend',0)

    buf = StringIO()
    for f in filas:
        is_cp   = es_complemento(f)
        estatus = calc_estatus(f, is_cp, ppd_pend, idx_razones)
        bg, fg, css, flt = classify(estatus)

        opts = ''.join(
            '<option value="%s"%s>%s</option>' % (op, ' selected' if op==estatus else '', op)
            for op in _OPCIONES
        )
        if estatus not in _OPCIONES:
            opts += '<option value="%s" selected>%s</option>' % (estatus, estatus)
        sel = '<select class="se %s" data-original="%s" onchange="ce(this)">%s</select>' % (css, estatus, opts)

        uuid_full = f['uuid']
        uuid_disp = uuid_full.upper()

        if is_cp:
            display, razon = resolver_uuid_rel(f, idx_razones, cp01_a_ppd)
            saldo = f.get('total', 0)
            # UUID REL → mostrar los 12 chars del PPD relacionado
            rel_disp = display if display else '–'
            # OBS: "Complemento parcialidad 1 – 5D1EDA96AB6E saldo insoluto $X"
            if display:
                obs_txt = 'Complemento parcialidad 1 - factura %s' % display
                if razon: obs_txt += ' (%s)' % razon[:30]
            else:
                obs_txt = 'Complemento CP01'
            obs_txt += ' saldo insoluto $%s' % '{:,.2f}'.format(saldo)
            obs_html = '<span class="cb" contenteditable="true" title="Clic para editar">%s</span>' % obs_txt
        elif norm_uuid(f.get('uuid','')) in ppd_norms:
            rel_raw  = f['uuid_rel']
            rel_disp = (rel_raw[-12:].upper() if len(rel_raw)>=12 else rel_raw.upper()) if rel_raw else '–'
            obs_html = '<span class="pb" contenteditable="true">PPD sin CP01 - siguiente mes</span>'
        else:
            rel_raw  = f['uuid_rel']
            rel_disp = (rel_raw[-12:].upper() if len(rel_raw)>=12 else rel_raw.upper()) if rel_raw else '–'
            obs_html = '<span class="ob-txt" contenteditable="true">-</span>'

        fecha_raw = f['fecha']
        try:
            fecha_str = (fecha_raw.strftime('%Y-%m-%d %H:%M')
                         if hasattr(fecha_raw,'strftime')
                         else str(fecha_raw or '').split('.')[0][:16])
        except Exception:
            fecha_str = str(fecha_raw or '')

        fc_cls = ' fc' if is_cp else ''
        buf.write(
            '<tr class="fr %s%s" data-est="%s" data-rfc="%s" data-uuid="%s">'
            '<td class="uu" title="%s">%s</td>'
            '<td class="uu" title="%s">%s</td>'
            '<td class="fd">%s</td>'
            '<td class="rz" title="%s">%s</td>'
            '<td class="nm"><input type="text" class="ip" value="%s" oninput="rc(this)"></td>'
            '<td class="nm"><input type="text" class="ip" value="%s" oninput="rc(this)"></td>'
            '<td class="nm"><input type="text" class="ip" value="%s" oninput="rc(this)"></td>'
            '<td class="tt" title="SUB2 + IVA16 + SUB0 = TOTAL"><b class="tot-val">%s</b><span class="tot-formula"></span></td>'
            '<td class="mt">%s</td>'
            '<td class="mt">%s</td>'
            '<td class="mt">%s</td>'
            '<td class="ec">%s</td>'
            '<td class="ob">%s</td>'
            '</tr>\n' % (
                flt, fc_cls, flt,
                f.get('rfc_em','').strip().upper(),
                uuid_full,
                uuid_full, uuid_disp,
                f['uuid_rel'], rel_disp,
                fecha_str,
                f['razon_em'], f['razon_em'],
                f['sub2'], f['iva16'], f['sub0'],
                fmt_num(f['total']),
                f['metodo'], f['forma'], f['uso'],
                sel, obs_html
            )
        )

    fs = buf.getvalue(); buf.close()

    # ── CSS ─────────────────────────────────────────────────────────
    _CSS = """
/* ── Variables de color ───────────────────────────────────────── */
:root {
  --rs: #FF2D78; --am: #FFD600; --az: #0057FF;
  --fo: #F4F4F8; --ca: #FFF;    --fa: #F7F8FC; --bo: #E0E0E8;
  --tx: #1A1A2E; --ts: #5A5A6E;
  --vr: #16A34A; --rj: #DC2626; --am2: #B8860B;
}
* { margin:0; padding:0; box-sizing:border-box; }
body { background:var(--fo); color:var(--tx); font-family:'DM Sans',sans-serif; }

/* ── Header ───────────────────────────────────────────────────── */
.hd { background:linear-gradient(135deg,#1C1C28,#222230); padding:22px 40px;
      border-bottom:3px solid transparent;
      border-image:linear-gradient(90deg,var(--rs),var(--am),var(--az)) 1; }
.ht { display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:16px; }
.lg { font-family:'Bebas Neue',sans-serif; font-size:42px; letter-spacing:3px; line-height:1; }
.lg .r { color:var(--rs); } .lg .d { color:#E0E0DC; } .lg .f { color:var(--am); }
.lg .v { font-size:13px; color:#3D8BFF; margin-left:6px; vertical-align:middle; }
.hm { font-size:12px; color:#8E8E9A; text-align:right; line-height:1.8;
      font-family:'JetBrains Mono',monospace; }
.hm strong { color:var(--am); font-size:13px; }

/* ── Puntos deducibles ────────────────────────────────────────── */
.ps { display:flex; gap:14px; padding:10px 40px;
      background:rgba(22,163,74,0.04); border-bottom:1px solid var(--bo);
      align-items:center; flex-wrap:wrap; }
.pl { font-family:'JetBrains Mono',monospace; font-size:10px; color:var(--ts); letter-spacing:2px; }
.pv { background:rgba(22,163,74,0.1); color:var(--vr); padding:4px 13px;
      border-radius:4px; font-family:'JetBrains Mono',monospace; font-size:11px;
      font-weight:700; border:1px solid rgba(22,163,74,0.2); }

/* ── Tarjetas de resumen ──────────────────────────────────────── */
.cs { display:grid; grid-template-columns:repeat(auto-fit,minmax(160px,1fr));
      gap:14px; padding:22px 40px; }
.cd { background:var(--ca); border-radius:10px; padding:16px 18px;
      border-left:4px solid transparent; box-shadow:0 2px 10px rgba(0,0,0,0.07); }
.cd1 { border-color:var(--vr); }  .cd2 { border-color:var(--rj); }
.cd3 { border-color:#999; }       .cd4 { border-color:var(--am); }
.cd5 { border-color:#E68A00; }
.cd-click { cursor:pointer; transition:all .2s; }
.cd-click:hover { box-shadow:0 4px 18px rgba(22,163,74,0.18);
                  border-left-width:6px; transform:translateY(-2px); }
.cd-click.diot-open { border-left-color:var(--am);
                      box-shadow:0 4px 18px rgba(184,134,11,0.25); }
.diot-ico { font-size:8px; color:var(--am); letter-spacing:1px;
            font-family:'JetBrains Mono',monospace; margin-left:4px;
            opacity:.7; }
.cl { font-size:9px; color:var(--ts); letter-spacing:1.5px; text-transform:uppercase;
      font-family:'JetBrains Mono',monospace; }
.cn { font-family:'Bebas Neue',sans-serif; font-size:36px; color:var(--tx); margin:4px 0; }
.cm { font-family:'JetBrains Mono',monospace; font-size:12px; color:var(--am2); }

/* ── Leyenda de colores ───────────────────────────────────────── */
.leyenda { display:flex; gap:8px; padding:10px 40px 14px; flex-wrap:wrap; align-items:center; }
.ley-tit { font-family:'JetBrains Mono',monospace; font-size:10px; color:var(--ts); letter-spacing:2px; }
.ley-item { display:flex; align-items:center; gap:5px; font-size:10px;
            font-family:'JetBrains Mono',monospace; color:var(--ts); }
.ley-dot { width:14px; height:14px; border-radius:3px; border:1px solid rgba(0,0,0,0.15); }

/* ── Botones de filtro ────────────────────────────────────────── */
.fls { padding:0 40px 14px; display:flex; gap:8px; flex-wrap:wrap; align-items:center; }
.fll { font-family:'JetBrains Mono',monospace; font-size:10px; color:var(--ts); letter-spacing:2px; }
.bf  { padding:7px 15px; border:1px solid var(--bo); border-radius:5px;
       background:transparent; color:var(--ts); font-size:11px;
       font-family:'JetBrains Mono',monospace; cursor:pointer; transition:all .2s; white-space:nowrap; }
.bf:hover { border-color:var(--rs); color:var(--rs); }
.bf.ac    { background:var(--rs); border-color:var(--rs); color:#fff; }
.bdd.ac   { background:var(--vr);  border-color:var(--vr);  color:#fff; }
.bno.ac   { background:var(--rj);  border-color:var(--rj);  color:#fff; }
.bpp.ac   { background:#666;       border-color:#666;        color:#fff; }
.bef.ac   { background:#E68A00;    border-color:#E68A00;     color:#fff; }
.beg.ac   { background:#7B1FA2;    border-color:#7B1FA2;     color:#fff; }
.bcp.ac   { background:#4A0080;    border-color:#4A0080;     color:#fff; }
.breset   { border-color:#999; color:#999; margin-left:8px; }
.breset:hover { border-color:var(--rj); color:var(--rj); }

/* ── Buscador personalizado ───────────────────────────────────── */
.bw { padding:0 40px 16px; }
.bq { width:100%; max-width:480px; padding:10px 16px; background:#FFF;
      border:1px solid #D0D0D8; border-radius:6px; color:var(--tx); font-size:13px; outline:none; }
.bq:focus { border-color:var(--rs); box-shadow:0 0 0 3px rgba(255,45,120,0.08); }

/* ── Tabla ────────────────────────────────────────────────────── */
.tw { padding:0 40px 40px; }
table.dataTable { font-size:12px; }
table.dataTable thead th {
  background:linear-gradient(180deg,#24242E,#1C1C26) !important;
  color:var(--am) !important; padding:11px 8px; text-align:center;
  font-family:'JetBrains Mono',monospace; font-size:10px; letter-spacing:0.8px;
  border-bottom:2px solid var(--rs) !important; white-space:nowrap;
}
table.dataTable tbody td { padding:9px 8px; border-bottom:1px solid var(--bo); vertical-align:middle; }
table.dataTable tbody tr:nth-child(even) { background:var(--fa); }
table.dataTable tbody tr:hover { background:rgba(255,45,120,0.04) !important; }

/* ── Celdas específicas ───────────────────────────────────────── */
.uu  { font-family:'JetBrains Mono',monospace; font-size:9px; color:var(--ts); word-break:break-all; }
.fd  { font-size:11px; color:var(--ts); white-space:nowrap; text-align:center; }
.rz  { font-size:11px; }
.nm  { text-align:right; }
.tt  { font-weight:700; color:var(--am2); text-align:right;
       font-family:'JetBrains Mono',monospace; white-space:nowrap; }
.tot-formula { display:block; font-size:8px; color:var(--ts);
               font-weight:400; opacity:0.7; margin-top:2px;
               font-family:'JetBrains Mono',monospace; }
.mt  { font-size:10px; color:var(--ts); max-width:150px;
       white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.ec  { text-align:center; }
.ob  { font-size:10px; max-width:320px; }

/* ── Input editable de montos ─────────────────────────────────── */
.ip { background:transparent; border:1px solid transparent; color:var(--tx);
      font-family:'JetBrains Mono',monospace; font-size:11px; font-weight:700;
      text-align:right; width:90px; padding:4px 5px; border-radius:4px; }
.ip:hover { border-color:#ccc; }
.ip:focus { border-color:#3D8BFF; background:rgba(0,87,255,0.04); outline:none; }

/* ── Select de estatus ────────────────────────────────────────── */
.se           { padding:6px 10px; border-radius:5px; font-size:10px;
                font-family:'JetBrains Mono',monospace; font-weight:700;
                border:none; cursor:pointer; min-width:155px; }
.se.ded16     { background:#C6EFCE; color:#1B5E20; }
.se.ded0      { background:#FFF3CD; color:#856404; }
.se.mix       { background:#BDD7EE; color:#1F497D; }
.se.egreso    { background:#E2CFED; color:#6A1B9A; }
.se.pendiente { background:#E0E0E0; color:#37474F; }
.se.complemento { background:#D5C6E0; color:#4A0080; }
.se.no-ded    { background:#FFC7CE; color:#9C0006; }
.se.changed   { box-shadow:0 0 0 2px var(--am); }

/* ── Etiquetas OBS ────────────────────────────────────────────── */
.cb { display:inline-block; padding:3px 8px; border-radius:3px;
      font-family:'JetBrains Mono',monospace; font-size:9px; font-weight:700;
      background:rgba(74,0,128,0.08); color:#7B1FA2;
      border:1px solid rgba(74,0,128,0.15); cursor:text; }
.pb { display:inline-block; padding:3px 8px; border-radius:3px;
      font-family:'JetBrains Mono',monospace; font-size:9px; font-weight:700;
      background:rgba(255,214,0,0.08); color:var(--am2);
      border:1px solid rgba(255,214,0,0.2); cursor:text; }
.ob-txt { font-size:10px; color:var(--ts); cursor:text; }
[contenteditable]:focus { outline:1px dashed #3D8BFF; border-radius:2px; padding:2px; }

/* ── Filas complemento / oculto ──────────────────────────────── */
.fc, .fc:nth-child(even) { background:rgba(213,198,224,0.15) !important; }
.oculto { display:none !important; }

/* ── Controles DataTables ─────────────────────────────────────── */
.dataTables_wrapper .dataTables_filter { display:none !important; }
.dataTables_wrapper .dataTables_length select {
  border:1px solid var(--bo); border-radius:4px; padding:4px 8px; }
.dataTables_info, .dataTables_paginate {
  font-family:'JetBrains Mono',monospace; font-size:11px; color:var(--ts); margin-top:12px; }
.paginate_button {
  padding:4px 10px !important; border:1px solid var(--bo) !important;
  border-radius:4px !important; margin:0 2px !important; cursor:pointer; }
.paginate_button.current {
  background:var(--rs) !important; color:#fff !important; border-color:var(--rs) !important; }
.paginate_button:hover:not(.current) {
  background:rgba(255,45,120,0.08) !important; border-color:var(--rs) !important; }
/* Filas de DataTables — respetar colores del reporte */
table.dataTable { border-collapse: collapse; width:100%; font-size:12px; }
table.dataTable thead th {
  background:linear-gradient(180deg,#24242E,#1C1C26) !important;
  color:var(--am) !important; padding:11px 8px; text-align:center;
  font-family:'JetBrains Mono',monospace; font-size:10px; letter-spacing:0.8px;
  border-bottom:2px solid var(--rs) !important; white-space:nowrap;
  border-top:none !important;
}
table.dataTable tbody td {
  padding:9px 8px; border-bottom:1px solid var(--bo); vertical-align:middle;
  background:inherit !important;
}
table.dataTable tbody tr { background:inherit !important; }
table.dataTable tbody tr:hover td { background:rgba(255,45,120,0.04) !important; }


/* ── DIOT ─────────────────────────────────────────────────────── */
.diot-wrap { padding:0 40px 30px; }
.diot-hdr  { display:flex; align-items:center; gap:12px;
             background:linear-gradient(135deg,#2C2C1E,#3A3A28);
             padding:14px 24px; border-radius:8px 8px 0 0;
             border-left:4px solid var(--am); }
.diot-dash  { color:var(--am); font-size:20px; font-weight:700; }
.diot-title { font-family:'Bebas Neue',sans-serif; font-size:20px;
              letter-spacing:2px; color:var(--am); }
.diot-mes   { font-family:'JetBrains Mono',monospace; font-size:11px;
              color:#8E8E6A; margin-left:auto; }
.diot-tw    { overflow-x:auto; border-radius:0 0 8px 8px;
              box-shadow:0 4px 16px rgba(0,0,0,0.10); }
.diot-tbl   { width:100%; border-collapse:collapse; font-size:12px; }
.diot-tbl thead th {
  background:linear-gradient(180deg,#24242E,#1C1C26);
  color:var(--am); padding:10px 14px; text-align:center;
  font-family:'JetBrains Mono',monospace; font-size:10px;
  letter-spacing:0.8px; border-bottom:2px solid var(--am);
  white-space:nowrap;
}
.diot-tbl thead th:first-child,
.diot-tbl thead th:nth-child(2) { text-align:left; }
.diot-row   { background:#FAFAF5; }
.diot-alt   { background:#F3F3EC; }
.diot-row:hover, .diot-alt:hover { background:rgba(184,134,11,0.06) !important; }
.diot-rs    { padding:9px 14px; font-size:11px; font-weight:500; color:var(--tx); }
.diot-rfc   { padding:9px 14px; font-family:'JetBrains Mono',monospace;
              font-size:10px; color:var(--ts); }
.diot-num   { padding:9px 14px; text-align:right;
              font-family:'JetBrains Mono',monospace; font-size:11px;
              color:var(--tx); border-left:1px solid var(--bo); }
.diot-tot   { padding:9px 14px; text-align:right;
              font-family:'JetBrains Mono',monospace; font-size:11px;
              font-weight:700; color:var(--am2);
              border-left:1px solid var(--bo); }
.diot-totrow { background:linear-gradient(135deg,#1C1C28,#222230); }
.diot-totrow td { border-top:2px solid var(--am); }
.diot-tl    { padding:10px 14px; font-family:'JetBrains Mono',monospace;
              font-size:11px; font-weight:700; color:var(--am);
              letter-spacing:2px; }

/* ── Footer ───────────────────────────────────────────────────── */
footer { padding:28px 40px; border-top:1px solid var(--bo);
         display:flex; justify-content:space-between; align-items:center;
         flex-wrap:wrap; gap:16px; background:linear-gradient(180deg,#1C1C28,#12121A); }
.fl2     { font-family:'Bebas Neue',sans-serif; font-size:26px; letter-spacing:4px; color:#E0E0DC; }
.fl2 .r  { color:var(--rs); }
.fl2 .f  { color:var(--am); }
.fi      { font-family:'JetBrains Mono',monospace; font-size:10px; color:#8E8E9A;
           text-align:right; line-height:1.9; opacity:.65; }
"""

    # ── JavaScript ──────────────────────────────────────────────────
    _JS = r"""
/* ── Estado global ────────────────────────────────────────────── */
let fa = 'todos';
let dtable;

const STORE_KEY = 'reaf_cambios_' + document.title.replace(/\s+/g,'_');

/* ── getUUID(): obtiene UUID del <tr> robusto con scrollX ─────── */
function getUUID(tr) {
  if (!tr) return '';
  let uuid = tr.getAttribute('data-uuid') || '';
  if (!uuid) {
    const tdUu = tr.querySelector('td.uu[title]');
    uuid = tdUu ? tdUu.getAttribute('title') : '';
  }
  if (!uuid && tr.cells[0]) uuid = tr.cells[0].getAttribute('title') || '';
  return uuid.trim().toUpperCase();
}

/* ── localStorage: guardar / borrar / leer cambios ───────────── */
function guardarCambio(uuid, estatus) {
  try {
    const d = JSON.parse(localStorage.getItem(STORE_KEY) || '{}');
    d[uuid] = estatus;
    localStorage.setItem(STORE_KEY, JSON.stringify(d));
  } catch(e) { console.warn('localStorage no disponible', e); }
}
function borrarCambio(uuid) {
  try {
    const d = JSON.parse(localStorage.getItem(STORE_KEY) || '{}');
    delete d[uuid];
    localStorage.setItem(STORE_KEY, JSON.stringify(d));
  } catch(e) {}
}
function getCambios() {
  try { return JSON.parse(localStorage.getItem(STORE_KEY) || '{}'); }
  catch(e) { return {}; }
}

/* ── guardarFila(): persiste sub2+iva+sub0+estatus ───────────── */
function guardarFila(tr) {
  try {
    const uuid = getUUID(tr);
    if (!uuid) return;
    const inputs = tr.querySelectorAll('.ip');
    const sel    = tr.querySelector('.se');
    const s2  = inputs[0] ? inputs[0].value : '';
    const iva = inputs[1] ? inputs[1].value : '';
    const s0  = inputs[2] ? inputs[2].value : '';
    const estatus = sel ? sel.value : '';
    const data = JSON.parse(localStorage.getItem(STORE_KEY) || '{}');
    data[uuid] = { sub2: s2, iva: iva, sub0: s0, estatus: estatus };
    localStorage.setItem(STORE_KEY, JSON.stringify(data));
  } catch(e) { console.warn('localStorage error', e); }
}

/* ── restaurarCambios(): restaura sub2+iva+sub0+estatus al cargar */
function restaurarCambios() {
  const cambios = getCambios();
  if (!Object.keys(cambios).length) return;
  $(dtable.rows().nodes()).each(function() {
    const tr   = this;
    const uuid = getUUID(tr);
    if (!uuid || !cambios[uuid]) return;
    const fila   = cambios[uuid];
    const inputs = tr.querySelectorAll('.ip');
    if (inputs.length >= 3) {
      if (fila.sub2 !== undefined) inputs[0].value = fila.sub2;
      if (fila.iva  !== undefined) inputs[1].value = fila.iva;
      if (fila.sub0 !== undefined) inputs[2].value = fila.sub0;
      recalcularTotal(tr,
        parseFloat(inputs[0].value) || 0,
        parseFloat(inputs[1].value) || 0,
        parseFloat(inputs[2].value) || 0);
    }
    const sel = tr.querySelector('.se');
    if (sel && fila.estatus) { sel.value = fila.estatus; ce(sel, false); }
  });
  dtable.draw(false);
  actualizarDiot();
  actualizarContadores();
}

/* ── exportarCambiosCSV(): descarga cambios como CSV de respaldo ─ */
function exportarCambiosCSV() {
  const cambios = getCambios();
  const keys = Object.keys(cambios);
  if (!keys.length) { alert('No hay cambios guardados.'); return; }
  let csv = 'UUID,ESTATUS_MANUAL\n';
  keys.forEach(uuid => { csv += uuid + ',' + cambios[uuid] + '\n'; });
  const blob = new Blob([csv], { type: 'text/csv;charset=utf-8;' });
  const url  = URL.createObjectURL(blob);
  const a    = document.createElement('a');
  a.href = url; a.download = 'cambios_diot_' + STORE_KEY.slice(-8) + '.csv';
  a.click(); URL.revokeObjectURL(url);
}

/* ── actualizarDiot(): recalcula DIOT con estatus actuales ──────
   FIX: usa dtable.rows().nodes() en lugar de $('#tbl tbody tr')
        para incluir filas de TODAS las páginas, no solo la visible.
   ─────────────────────────────────────────────────────────────── */
function actualizarDiot() {
  const sec = document.getElementById('diot-section');
  if (!sec || sec.style.display === 'none') return;

  const provs = {};
  $(dtable.rows().nodes()).each(function() {
    const est = this.getAttribute('data-est') || '';
    if (est !== 'ded' && est !== 'efe') return;
    const cells = this.cells;
    const razon = cells[3] ? cells[3].innerText.trim() : '';
    const rfc   = this.getAttribute('data-rfc') || '';
    if (!rfc) return;
    const s2  = parseFloat((cells[4]?.querySelector('input')?.value) || 0) || 0;
    const i16 = parseFloat((cells[5]?.querySelector('input')?.value) || 0) || 0;
    const s0  = parseFloat((cells[6]?.querySelector('input')?.value) || 0) || 0;
    const tot = parseFloat(((cells[7]?.innerText || '0').replace(/[^0-9.]/g,''))) || 0;
    if (!provs[rfc]) provs[rfc] = { razon, rfc, s2:0, i16:0, s0:0, tot:0 };
    provs[rfc].s2  += s2;  provs[rfc].i16 += i16;
    provs[rfc].s0  += s0;  provs[rfc].tot += tot;
  });

  const lista = Object.values(provs).sort((a,b) => a.razon.localeCompare(b.razon));
  const fmt   = v => v > 0 ? '$' + v.toFixed(2).replace(/\B(?=(\d{3})+(?!\d))/g,',') : '-';
  let tS2=0, tI16=0, tS0=0, tTot=0;
  lista.forEach(p => { tS2+=p.s2; tI16+=p.i16; tS0+=p.s0; tTot+=p.tot; });

  let rows = lista.map((p,i) =>
    `<tr class="diot-row${i%2===0?' diot-alt':''}">
      <td class="diot-rs">${p.razon}</td><td class="diot-rfc">${p.rfc}</td>
      <td class="diot-num">${fmt(p.s2)}</td><td class="diot-num">${fmt(p.i16)}</td>
      <td class="diot-num">${fmt(p.s0)}</td><td class="diot-tot">${fmt(p.tot)}</td>
    </tr>`
  ).join('');
  rows += `<tr class="diot-totrow">
    <td colspan="2" class="diot-tl">TOTALES</td>
    <td class="diot-tot">${fmt(tS2)}</td><td class="diot-tot">${fmt(tI16)}</td>
    <td class="diot-tot">${fmt(tS0)}</td><td class="diot-tot">${fmt(tTot)}</td>
  </tr>`;
  const tbody = sec.querySelector('.diot-tbl tbody');
  if (tbody) tbody.innerHTML = rows;
}

/* ── Filtro por deducibilidad (data-est del <tr>) ─────────────── */
$.fn.dataTable.ext.search.push(function(settings, data, dataIndex) {
  if (fa === 'todos') return true;
  const row = dtable.row(dataIndex).node();
  const est = (row ? row.getAttribute('data-est') : '') || '';
  if (fa === 'ded')         return est === 'ded';
  if (fa === 'efe')         return est === 'efe';
  if (fa === 'no-ded')      return est === 'no-ded';
  if (fa === 'pendiente')   return est === 'pendiente';
  if (fa === 'egreso')      return est === 'egreso';
  if (fa === 'complemento') return est === 'complemento';
  return true;
});

/* ── Inicialización DataTable ─────────────────────────────────── */
$(document).ready(function() {
  dtable = $('#tbl').DataTable({
    pageLength : 100,
    lengthMenu : [25, 50, 100, 200, 500],
    deferRender: true,
    scrollX    : true,
    autoWidth  : true,
    order      : [[2, 'asc']],
    stateSave  : true,
    stateDuration: 0,
    language   : {
      search     : 'Buscar:',
      lengthMenu : 'Mostrar _MENU_',
      info       : '_START_ a _END_ de _TOTAL_',
      paginate   : { previous: 'Ant', next: 'Sig' },
      zeroRecords: 'Sin resultados'
    },
    columnDefs: [
      { targets: [0, 1],     width: '160px' },
      { targets: [2],        width: '100px' },
      { targets: [3],        width: '160px' },
      { targets: [4, 5, 6],  width: '80px'  },
      { targets: [7],        width: '90px'  },
      { targets: [8, 9, 10], width: '140px' },
      { targets: [11],       width: '140px' },
      { targets: [12],       width: '220px' }
    ]
  });

  $('#bq-custom').on('keyup', function() { dtable.search(this.value).draw(); });

  // Inicializar fórmulas visibles en todas las filas al cargar
  dtable.rows().every(function() {
    const tr  = this.node();
    const ins = tr ? tr.querySelectorAll('.ip') : [];
    if (ins.length >= 3) {
      recalcularTotal(tr,
        parseFloat(ins[0].value)||0,
        parseFloat(ins[1].value)||0,
        parseFloat(ins[2].value)||0
      );
    }
  });

  restaurarCambios();
  actualizarContadores();
});

/* ── toggleDiot(): mostrar/ocultar DIOT ──────────────────────── */
function toggleDiot(card) {
  const sec = document.getElementById('diot-section');
  if (!sec) return;
  const visible = sec.style.display !== 'none';
  sec.style.display = visible ? 'none' : 'block';
  card.classList.toggle('diot-open', !visible);
  if (!visible) actualizarDiot();
}

/* ── actualizarContadores(): recuenta data-est en TODAS las filas */
function actualizarContadores() {
  let ded=0, efe=0, nod=0, pen=0, eg=0, cp=0, tot=0;
  let mDed=0, mNod=0, mPen=0;
  $(dtable.rows().nodes()).each(function() {
    const est  = this.getAttribute('data-est') || '';
    const tots = parseFloat(
      ($(this).find('td:eq(7)').text() || '0').replace(/[^0-9.]/g,'')
    ) || 0;
    tot++;
    if      (est==='ded')         { ded++; mDed+=tots; }
    else if (est==='efe')         { efe++; mDed+=tots; }
    else if (est==='no-ded')      { nod++; mNod+=tots; }
    else if (est==='pendiente')   { pen++; mPen+=tots; }
    else if (est==='egreso')      { eg++;  }
    else if (est==='complemento') { cp++;  }
  });
  const bf = document.querySelectorAll('.fls .bf');
  if (bf[0]) bf[0].textContent = 'TODOS (' +tot+ ')';
  if (bf[1]) bf[1].textContent = 'DED ('   +ded+ ')';
  if (bf[2]) bf[2].textContent = 'EFE ('   +efe+ ')';
  if (bf[3]) bf[3].textContent = 'NO DED ('+nod+ ')';
  if (bf[4]) bf[4].textContent = 'PEND ('  +pen+ ')';
  if (bf[5]) bf[5].textContent = 'EGRESO ('+eg+  ')';
  if (bf[6]) bf[6].textContent = 'CP01 ('  +cp+  ')';
  const fmt = v => '$'+v.toFixed(2).replace(/\B(?=(\d{3})+(?!\d))/g,',');
  const el  = id => document.getElementById(id);
  if (el('cnt-ded')) el('cnt-ded').textContent = ded+efe;
  if (el('cnt-efe')) el('cnt-efe').textContent = efe;
  if (el('cnt-nod')) el('cnt-nod').textContent = nod;
  if (el('cnt-pen')) el('cnt-pen').textContent = pen;
  if (el('cnt-tot')) el('cnt-tot').textContent = tot;
  if (el('mnt-ded')) el('mnt-ded').textContent = fmt(mDed);
  if (el('mnt-nod')) el('mnt-nod').textContent = fmt(mNod);
  if (el('mnt-pen')) el('mnt-pen').textContent = fmt(mPen);
  actualizarDiot();
}

/* ── ft(): activar filtro por botón ──────────────────────────── */
function ft(t, btn) {
  fa = t;
  document.querySelectorAll('.fls .bf').forEach(b => b.classList.remove('ac'));
  if (btn) btn.classList.add('ac');
  dtable.draw();
}

/* ── ce(): color + data-est + localStorage ──────────────────────
   FIX PRINCIPAL: función única con parámetro guardar=true.
   La segunda definición anterior (sin guardarCambio) causaba que
   los cambios del usuario no se persistieran en localStorage.
   ─────────────────────────────────────────────────────────────── */
function ce(sel, guardar = true) {
  const v = sel.value;
  const o = sel.dataset.original;
  sel.className = 'se';

  if      (v.includes('COMPLE'))                        sel.classList.add('complemento');
  else if (v.includes('PEND'))                          sel.classList.add('pendiente');
  else if (v.includes('NO DED') || v.includes('ERROR')) sel.classList.add('no-ded');
  else if (v.includes('EGRESO'))                        sel.classList.add('egreso');
  else if (v.includes('16 Y 0'))                        sel.classList.add('mix');
  else if (v.includes('0%') && !v.includes('16'))       sel.classList.add('ded0');
  else if (v.includes('16%'))                           sel.classList.add('ded16');
  else                                                  sel.classList.add('no-ded');

  if (v !== o) sel.classList.add('changed');
  else         sel.classList.remove('changed');

  const tr = sel.closest('tr');
  if (tr) {
    let est = 'no-ded';
    if      (v.includes('COMPLE'))                        est = 'complemento';
    else if (v.includes('PEND'))                          est = 'pendiente';
    else if (v.includes('NO DED') || v.includes('ERROR')) est = 'no-ded';
    else if (v.includes('EGRESO'))                        est = 'egreso';
    else if (v.includes('EFE'))                           est = 'efe';
    else if (v.includes('DED'))                           est = 'ded';
    tr.setAttribute('data-est', est);

    if (guardar) {
      const uuid = (tr.cells[0] ? tr.cells[0].getAttribute('title') || '' : '').trim().toUpperCase();
      if (uuid) {
        if (v !== o) guardarCambio(uuid, v);
        else         borrarCambio(uuid);
      }
    }
  }
  actualizarContadores();
}

/* ── limpiarCambios(): resetea al estado original ─────────────── */
function limpiarCambios() {
  if (!confirm('¿Borrar todos los cambios manuales y volver al estado original?')) return;
  try { localStorage.removeItem(STORE_KEY); } catch(e) {}
  location.reload();
}

/* ── beforeunload: guardar todas las filas al cerrar/recargar ─── */
window.addEventListener('beforeunload', function() {
  try {
    $(dtable.rows().nodes()).each(function() {
      guardarFila(this);
    });
  } catch(e) {}
});

/* ── recalcularTotal(): SUB2 + IVA16 + SUB0 = TOTAL ─────────────── */
function recalcularTotal(tr, s2, i16, s0) {
  const total  = Math.round((s2 + i16 + s0) * 100) / 100;
  const totVal = tr.querySelector('.tot-val');
  const totFrm = tr.querySelector('.tot-formula');
  if (totVal) {
    totVal.textContent = '$' + total.toLocaleString('es-MX', {
      minimumFractionDigits: 2, maximumFractionDigits: 2
    });
  }
  if (totFrm) {
    const fmt = v => v % 1 === 0 ? v.toFixed(0) : v.toFixed(2);
    totFrm.textContent = (s2||i16||s0) ? fmt(s2)+' + '+fmt(i16)+' + '+fmt(s0) : '';
  }
  return total;
}

/* ── rc(): recalcular estatus al editar montos ────────────────── */
function rc(inp) {
  const tr = inp.closest('tr');
  if (!tr) return;
  const ins = tr.querySelectorAll('.ip');
  if (ins.length < 3) return;
  const s2  = parseFloat(ins[0].value) || 0;
  const i16 = parseFloat(ins[1].value) || 0;
  const s0  = parseFloat(ins[2].value) || 0;
  const tot = recalcularTotal(tr, s2, i16, s0);
  guardarFila(tr);
  const sel = tr.querySelector('.se');
  if (!sel) return;
  const forma = (tr.cells[9]?.innerText  || '').trim();
  const uso   = (tr.cells[10]?.innerText || '').trim();
  const met   = (tr.cells[8]?.innerText  || '').trim();
  const fp = forma.substring(0,2);
  const uc = uso.substring(0,3).toUpperCase();
  const mc = met.substring(0,3).toUpperCase();
  if (uc === 'S01') { sel.value = 'NO DEDUCIBLE'; ce(sel); return; }
  if (!['G01','G02','G03'].includes(uc) ||
      !['PUE','PPD'].includes(mc)       ||
      !['01','02','03','04','28'].includes(fp)) return;
  if (fp==='01' && tot>=2000) { sel.value='NO DEDUCIBLE: Efectivo >= $2,000'; ce(sel); return; }
  if (uc==='G02') { sel.value='EGRESO'; ce(sel); return; }
  let suf = 'NO DEDUCIBLE';
  if      (s2>0 && i16>0 && s0===0) suf='16%';
  else if (s2>0 && i16>0 && s0>0)  suf='16 Y 0%';
  else if (s2===0 && i16===0 && s0>0) suf='0%';
  sel.value = (fp==='01' ? 'EFE ' : 'DED ') + suf;
  ce(sel);
}
"""

    # ── Ensamblado final del HTML ───────────────────────────────────
    # El <head> con CSS no usa variables Python → concatenación directa
    _head = (
        '<!DOCTYPE html>\n'
        '<html lang="es">\n'
        '<head>\n'
        '<meta charset="UTF-8">'
        '<meta name="viewport" content="width=device-width,initial-scale=1.0">\n'
        '<title>Reporte Fiscal ' + mes + ' - ReaDesF1.9</title>\n'
        '<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue'
        '&family=DM+Sans:wght@300;400;500;700'
        '&family=JetBrains+Mono:wght@400;700&display=swap" rel="stylesheet">\n'
        '<script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>\n'
        '<script src="https://cdn.datatables.net/1.13.7/js/jquery.dataTables.min.js"></script>\n'
        '<style>\n' + _CSS + '\n</style>\n'
        '</head><body>\n'
    )

    # El <body> con variables Python → format normal (sin CSS/JS)
    _body = (
        '<div class="hd"><div class="ht">\n'
        '  <div class="lg"><span class="r">Rea</span><span class="d">Des</span>'
        '<span class="f">F</span><span class="v">1.9</span></div>\n'
        '  <div class="hm"><strong>REPORTE FISCAL - REGIMEN {reg_cod}</strong>'
        '<br>{reg_nombre} . {mes}<br>Generado: {now}</div>\n'
        '</div></div>\n'
        '<div class="ps"><span class="pl">PUNTOS DEDUCIBLES:</span>'
        '<span class="pv">PUE</span>'
        '<span class="pv">01, 02, 03, 28</span>'
        '<span class="pv">G01 Y G03</span></div>\n'
        '<div class="cs">\n'
        '  <div class="cd cd1 cd-click" onclick="toggleDiot(this)" title="Ver DIOT">'
        '<div class="cl">DEDUCIBLES <span class="diot-ico">— DIOT</span></div>'
        '<div class="cn" id="cnt-ded">{nd}</div><div class="cm" id="mnt-ded">${md}</div></div>\n'
        '  <div class="cd cd5"><div class="cl">EFECTIVO</div>'
        '<div class="cn" id="cnt-efe">{ne}</div></div>\n'
        '  <div class="cd cd2"><div class="cl">NO DEDUCIBLES</div>'
        '<div class="cn" id="cnt-nod">{nn}</div><div class="cm" id="mnt-nod">${mn}</div></div>\n'
        '  <div class="cd cd3"><div class="cl">PENDIENTES</div>'
        '<div class="cn" id="cnt-pen">{np}</div><div class="cm" id="mnt-pen">${mp}</div></div>\n'
        '  <div class="cd cd4"><div class="cl">TOTAL FACTURAS</div>'
        '<div class="cn" id="cnt-tot">{tf}</div></div>\n'
        '</div>\n'
        '<div class="leyenda"><span class="ley-tit">COLORES:</span>\n'
        '  <span class="ley-item"><span class="ley-dot" style="background:#C6EFCE"></span>DED/EFE 16%</span>\n'
        '  <span class="ley-item"><span class="ley-dot" style="background:#BDD7EE"></span>16 Y 0%</span>\n'
        '  <span class="ley-item"><span class="ley-dot" style="background:#FFF3CD"></span>0%</span>\n'
        '  <span class="ley-item"><span class="ley-dot" style="background:#E2CFED"></span>EGRESO</span>\n'
        '  <span class="ley-item"><span class="ley-dot" style="background:#D5C6E0"></span>COMPLEMENTO</span>\n'
        '  <span class="ley-item"><span class="ley-dot" style="background:#E0E0E0"></span>PENDIENTE</span>\n'
        '  <span class="ley-item"><span class="ley-dot" style="background:#FFC7CE"></span>NO DEDUCIBLE</span>\n'
        '</div>\n'
        '<div class="fls"><span class="fll">FILTRAR:</span>\n'
        '  <button class="bf ac"  onclick="ft(\'todos\',this)">TODOS ({tf})</button>\n'
        '  <button class="bf bdd" onclick="ft(\'ded\',this)">DED ({ndd})</button>\n'
        '  <button class="bf bef" onclick="ft(\'efe\',this)">EFE ({ne})</button>\n'
        '  <button class="bf bno" onclick="ft(\'no-ded\',this)">NO DED ({nn})</button>\n'
        '  <button class="bf bpp" onclick="ft(\'pendiente\',this)">PEND ({np})</button>\n'
        '  <button class="bf beg" onclick="ft(\'egreso\',this)">EGRESO ({neg})</button>\n'
        '  <button class="bf bcp" onclick="ft(\'complemento\',this)">CP01 ({nc})</button>\n'
        '  <button class="bf breset" onclick="limpiarCambios()" title="Borrar cambios manuales y volver al original">↺ LIMPIAR</button>\n'
        '  <button class="bf breset" onclick="exportarCambiosCSV()" title="Descargar cambios manuales como CSV de respaldo" style="border-color:#1976D2;color:#1976D2">⬇ EXPORTAR CAMBIOS</button>\n'
        '</div>\n'
        '<div class="bw">'
        '<input id="bq-custom" class="bq" type="search" '
        'placeholder="Buscar por UUID, razon social, monto...">'
        '</div>\n'
        '<div id="diot-section" style="display:none">{diot_html}</div>\n'
        '<div class="tw"><table id="tbl" class="display" style="width:100%">\n'
        '  <thead><tr>'
        '<th>FOLIO FISCAL</th><th>UUID REL.</th><th>FECHA</th><th>RAZON SOCIAL</th>\n'
        '    <th>SUBTOTAL 16%</th><th>IVA 16%</th><th>SUB 0%</th><th>TOTAL</th>\n'
        '    <th>METODO</th><th>FORMA PAGO</th><th>USO CFDI</th>'
        '<th>ESTATUS</th><th>COMPLEMENTOS / OBS</th>\n'
        '  </tr></thead>\n'
        '  <tbody>{fs}</tbody>\n'
        '</table></div>\n'
        '<footer>\n'
        '  <div class="fl2"><span class="r">Rea</span>Des<span class="f">F</span></div>\n'
        '  <div class="fi">Sinergia REA . ReaDesF1.9 . Mexico 2026<br>'
        'Regimen {reg_cod} . {reg_nombre}<br>'
        '{nd2} deducibles . {nn} no deducibles . {np} pendientes</div>\n'
        '</footer>\n'
    ).format(
        mes=mes, reg_cod=reg_cod, reg_nombre=reg_nombre, now=now_str,
        nd=n_ded+n_efe, md='{:,.2f}'.format(m_ded), ne=n_efe,
        nn=n_no_ded,    mn='{:,.2f}'.format(m_no_ded),
        np=n_pend,      mp='{:,.2f}'.format(m_pend),
        tf=total_f, ndd=n_ded, neg=n_egreso, nc=n_comp,
        nd2=n_ded+n_efe, fs=fs,
        diot_html=generar_diot_html(filas, ppd_pend, idx_razones, mes)
    )

    # El JS se concatena directo — sin .format() para no escapar sus llaves
    html = _head + _body + '<script>\n' + _JS + '\n</script>\n</body></html>'

    with open(out, 'w', encoding='utf-8') as fp:
        fp.write(html)
    print('  HTML: %s' % out)


def generar_diot_html(filas, ppd_pend, idx_razones, mes):
    """
    Genera la sección HTML de la DIOT (Informativa de Operaciones con Terceros).
    Agrupa por RFC emisor, solo facturas deducibles (DED + EFE).

    FIX v2.4: acumula sub2/iva16/sub0 sin round() intermedio para evitar
    deriva de centavos al sumar N facturas. Se redondea solo al mostrar.
    """
    # Acumular por RFC — sin redondeo intermedio (float64 nativo)
    proveedores = {}
    for f in filas:
        is_cp   = es_complemento(f)
        estatus = calc_estatus(f, is_cp, ppd_pend, idx_razones).upper()

        # Solo deducibles — excluir NO DED, PENDIENTE, COMPLEMENTO, EGRESO, ERROR
        if not any(k in estatus for k in ('DED', 'EFE')):
            continue
        if any(k in estatus for k in ('NO DED', 'NO DEDUCIBLE', 'ERROR')):
            continue

        rfc   = f.get('rfc_em', '').strip().upper() or 'SIN RFC'
        razon = f.get('razon_em', '').strip() or rfc

        if rfc not in proveedores:
            proveedores[rfc] = {
                'razon': razon,
                'rfc':   rfc,
                'sub2':  0.0,
                'iva16': 0.0,
                'sub0':  0.0,
                'total': 0.0,
            }
        p = proveedores[rfc]
        # Usar float() sin round() — acumular con máxima precisión
        p['sub2']  += float(f.get('sub2',  0.0) or 0.0)
        p['iva16'] += float(f.get('iva16', 0.0) or 0.0)
        p['sub0']  += float(f.get('sub0',  0.0) or 0.0)
        p['total'] += float(f.get('total', 0.0) or 0.0)

    if not proveedores:
        return ''

    # Redondear acumulados a 2 decimales (solo al final, no por fila)
    for p in proveedores.values():
        p['sub2']  = round(p['sub2'],  2)
        p['iva16'] = round(p['iva16'], 2)
        p['sub0']  = round(p['sub0'],  2)
        p['total'] = round(p['total'], 2)

    # Ordenar por razón social
    filas_diot = sorted(proveedores.values(), key=lambda x: x['razon'])

    # Totales
    tot_sub2  = sum(p['sub2']  for p in filas_diot)
    tot_iva16 = sum(p['iva16'] for p in filas_diot)
    tot_sub0  = sum(p['sub0']  for p in filas_diot)
    tot_total = sum(p['total'] for p in filas_diot)

    def m(v):
        return '${:,.2f}'.format(v) if v else '-'

    # Construir filas HTML
    rows = ''
    for i, p in enumerate(filas_diot):
        alt = ' diot-alt' if i % 2 == 0 else ''
        rows += (
            '<tr class="diot-row%s">'
            '<td class="diot-rs">%s</td>'
            '<td class="diot-rfc">%s</td>'
            '<td class="diot-num">%s</td>'
            '<td class="diot-num">%s</td>'
            '<td class="diot-num">%s</td>'
            '<td class="diot-tot">%s</td>'
            '</tr>\n' % (
                alt,
                p['razon'], p['rfc'],
                m(p['sub2']), m(p['iva16']), m(p['sub0']), m(p['total'])
            )
        )

    # Fila de totales
    rows += (
        '<tr class="diot-totrow">'
        '<td colspan="2" class="diot-tl">TOTALES</td>'
        '<td class="diot-tot">%s</td>'
        '<td class="diot-tot">%s</td>'
        '<td class="diot-tot">%s</td>'
        '<td class="diot-tot">%s</td>'
        '</tr>\n' % (m(tot_sub2), m(tot_iva16), m(tot_sub0), m(tot_total))
    )

    return (
        '<div class="diot-wrap">\n'
        '  <div class="diot-hdr">\n'
        '    <span class="diot-dash">—</span>\n'
        '    <span class="diot-title">Informativa de operaciones con terceros (DIOT)</span>\n'
        '    <span class="diot-mes">%s</span>\n'
        '  </div>\n'
        '  <div class="diot-tw">\n'
        '  <table class="diot-tbl">\n'
        '    <thead><tr>'
        '<th>RAZÓN SOCIAL</th>'
        '<th>RFC</th>'
        '<th>ACT. PAGADAS 16%%</th>'
        '<th>IVA PAGADO</th>'
        '<th>ACT. PAGADO TASA 0%%</th>'
        '<th>TOTAL</th>'
        '</tr></thead>\n'
        '    <tbody>%s</tbody>\n'
        '  </table>\n'
        '  </div>\n'
        '</div>\n'
    ) % (mes.upper(), rows)


def generar_reporte(validado_path, mes_reporte=''):
    t0 = time.time()
    print('\n' + '='*60)
    print('  GENERADOR DE REPORTE FISCAL - ReaDesF1.8 v2.2')
    print('='*60)

    if not os.path.exists(validado_path):
        print('  ERROR: No encontrado: %s' % validado_path)
        return

    base     = re.sub(r'_validado\.xlsx$', '', validado_path, flags=re.IGNORECASE)
    base     = re.sub(r'\.xlsx$',          '', base,           flags=re.IGNORECASE)
    xlsx_out = base + '_reporte.xlsx'
    html_out = base + '_reporte.html'

    filas               = leer_validado(validado_path)
    reg_cod, reg_nombre = detectar_regimen(filas)
    idx_razones         = construir_indice_razones(filas)
    ppd_pend, cp01_a_ppd = detectar_ppd(filas)

    print('  Regimen: %s - %s' % (reg_cod, reg_nombre))
    print('  PPD pendientes (sin CP01): %d' % len(ppd_pend))
    print('  CP01 cruzados por RFC+total: %d' % len(cp01_a_ppd))

    stats = generar_excel(filas, ppd_pend, cp01_a_ppd, idx_razones,
                          xlsx_out, mes_reporte, reg_cod, reg_nombre)
    generar_html(filas, ppd_pend, cp01_a_ppd, idx_razones,
                 html_out, mes_reporte, stats, reg_cod, reg_nombre)

    t = time.time() - t0
    print('\n' + '='*60)
    print('  Listo en %.2fs' % t)
    print('  %s' % xlsx_out)
    print('  %s' % html_out)
    print('  DED:%d  NO:%d  PEND:%d  CP01:%d' % (
        stats['ded']+stats['efe'], stats['no_ded'], stats['pend'], stats['comp']))
    print('='*60 + '\n')


if __name__ == '__main__':
    import sys
    if len(sys.argv) >= 2:
        path = sys.argv[1]
        mes  = sys.argv[2] if len(sys.argv) >= 3 else ''
    else:
        carpeta = input('\n  Carpeta (Enter = Escritorio/GASTOS RESICO/2026/FEBRERO26): ').strip()
        if not carpeta:
            carpeta = os.path.join(os.path.expanduser('~'),'Desktop','GASTOS RESICO','2026','FEBRERO26')
        nombre = input('  Nombre del archivo _validado (sin .xlsx): ').strip()
        nombre = re.sub(r'\.xlsx$','',nombre,flags=re.IGNORECASE)
        if not nombre.endswith('_validado'):
            nombre += '_validado'
        mes  = input('  Mes del reporte (ej: FEBRERO 2026): ').strip()
        path = os.path.join(carpeta, nombre+'.xlsx')
    if not os.path.exists(path):
        print('\n  ERROR: No encontrado: %s' % path)
        carpeta_check = os.path.dirname(path)
        if os.path.exists(carpeta_check):
            excels = [f for f in os.listdir(carpeta_check) if f.endswith('.xlsx')]
            if excels:
                print('\n  Archivos disponibles:')
                for f in excels[:10]:
                    print('    . %s' % f)
    else:
        generar_reporte(path, mes_reporte=mes)
    input('\n>>> ENTER para salir...')
