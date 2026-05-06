"""
analizador_sistema.py — Adaptive Processing v1.8
ReaDesF1.8

NUEVO EN v1.8:
  ✓ Tercer motor: pandas_chunks para archivos >30,000 filas
  ✓ Tabla de decisión actualizada con 4 modos

MOTORES:
  🚀 TURBO        — pandas completo       (RAM ≥ 4GB, 5k-30k filas)
  📦 CHUNKS       — pandas por bloques    (RAM ≥ 4GB, >30k filas)
  🔧 SEGURO       — openpyxl optimizado   (RAM 2-4GB)
  🐢 MÍNIMO       — openpyxl conservador  (RAM < 2GB)

TABLA DE DECISIÓN:
  Filas      RAM        Motor
  ─────────  ─────────  ──────────────────
  < 5,000    cualquiera openpyxl  SEGURO
  5k-30k     < 4 GB     openpyxl  SEGURO
  5k-30k     ≥ 4 GB     pandas    TURBO
  > 30k      < 4 GB     openpyxl  MÍNIMO
  > 30k      ≥ 4 GB     chunks    CHUNKS
  cualquiera < 2 GB     openpyxl  MÍNIMO
"""

import os
import multiprocessing
import platform
from dataclasses import dataclass

try:
    import psutil
    PSUTIL_OK = True
except ImportError:
    PSUTIL_OK = False

try:
    import openpyxl
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


@dataclass
class ResultadoAnalisis:
    ram_total_gb:      float
    ram_disponible_gb: float
    cpu_cores:         int
    sistema_operativo: str
    archivo_mb:        float
    filas_reales:      int
    columnas:          int
    motor:             str    # 'pandas' | 'openpyxl' | 'pandas_chunks'
    modo:              str    # 'TURBO' | 'CHUNKS' | 'SEGURO' | 'MÍNIMO'
    razon:             str
    pandas_disponible: bool
    chunk_size:        int    # tamaño de bloque para motor chunks


def analizar_sistema() -> tuple:
    so = platform.system()
    if PSUTIL_OK:
        mem = psutil.virtual_memory()
        ram_total = mem.total     / (1024 ** 3)
        ram_disp  = mem.available / (1024 ** 3)
    else:
        ram_total = 2.0
        ram_disp  = 1.0
        print("  ⚠️  psutil no instalado — pip install psutil")
    return ram_total, ram_disp, multiprocessing.cpu_count(), so


def analizar_archivo(file_path: str) -> tuple:
    archivo_mb = os.path.getsize(file_path) / (1024 * 1024)
    filas      = max(100, int(archivo_mb * 1000))
    columnas   = 20

    if OPENPYXL_OK:
        try:
            wb    = openpyxl.load_workbook(file_path, read_only=True)
            sh    = wb.active
            filas = max(0, (sh.max_row or 1) - 1)
            columnas = sh.max_column or 20
            wb.close()
        except Exception as e:
            print(f"  ⚠️  No se pudieron contar filas: {e}")
    return archivo_mb, filas, columnas


def verificar_pandas() -> bool:
    try:
        import pandas  # noqa
        return True
    except ImportError:
        return False


def calcular_chunk_size(ram_gb: float, columnas: int) -> int:
    """
    Calcula el tamaño óptimo de bloque según RAM disponible.
    Más RAM → bloques más grandes → menos iteraciones.
    """
    if ram_gb >= 8:
        return 10000
    elif ram_gb >= 4:
        return 5000
    elif ram_gb >= 2:
        return 2000
    else:
        return 1000


def elegir_motor(ram_gb: float, archivo_mb: float,
                 filas: int, pandas_ok: bool) -> tuple:
    """
    Retorna: (motor, modo, razon, chunk_size)
    """
    chunk_size = calcular_chunk_size(ram_gb, 20)

    if not pandas_ok:
        return ('openpyxl', 'SEGURO', chunk_size,
                'pandas no instalado — pip install pandas')

    if ram_gb < 2.0:
        return ('openpyxl', 'MÍNIMO', chunk_size,
                f'RAM crítica ({ram_gb:.1f} GB) — modo mínimo para estabilidad')

    if filas < 5000:
        return ('openpyxl', 'SEGURO', chunk_size,
                f'Archivo pequeño ({filas:,} filas) — openpyxl es suficiente')

    if filas <= 30000 and ram_gb < 4.0:
        return ('openpyxl', 'SEGURO', chunk_size,
                f'RAM moderada ({ram_gb:.1f} GB) — openpyxl optimizado')

    if filas <= 30000 and ram_gb >= 4.0:
        return ('pandas', 'TURBO', chunk_size,
                f'RAM suficiente ({ram_gb:.1f} GB) + {filas:,} filas — pandas TURBO')

    if filas > 30000 and ram_gb < 4.0:
        return ('openpyxl', 'MÍNIMO', chunk_size,
                f'Archivo grande ({filas:,} filas) con RAM limitada — openpyxl conservador')

    if filas > 30000 and ram_gb >= 4.0:
        return ('pandas_chunks', 'CHUNKS', chunk_size,
                f'Archivo muy grande ({filas:,} filas) + RAM amplia ({ram_gb:.1f} GB) — '
                f'pandas chunks de {chunk_size:,} filas')

    return ('openpyxl', 'SEGURO', chunk_size, 'Configuración no clasificada — modo seguro')


def analizar_y_decidir(file_path: str) -> ResultadoAnalisis:

    print("\n╔" + "═" * 68 + "╗")
    print("║" + "  🖥️  ANÁLISIS DE RECURSOS — ReaDesF1.8".center(68) + "║")
    print("╠" + "═" * 68 + "╣")

    print("║  📊 HARDWARE:".ljust(69) + "║")
    ram_total, ram_disp, cpu, so = analizar_sistema()
    print(f"║    RAM total      : {ram_total:.2f} GB".ljust(69) + "║")
    print(f"║    RAM disponible : {ram_disp:.2f} GB".ljust(69) + "║")
    print(f"║    CPU núcleos    : {cpu}".ljust(69) + "║")
    print(f"║    Sistema        : {so}".ljust(69) + "║")

    estado_ram = ("🟢 Excelente" if ram_disp >= 4
                  else "🟡 Moderada" if ram_disp >= 2
                  else "🔴 Baja — modo conservador")
    print(f"║    Estado RAM     : {estado_ram}".ljust(69) + "║")

    print("╠" + "═" * 68 + "╣")
    print("║  📁 ARCHIVO:".ljust(69) + "║")
    arch_mb, filas, cols = analizar_archivo(file_path)
    print(f"║    Tamaño         : {arch_mb:.2f} MB".ljust(69) + "║")
    print(f"║    Filas          : {filas:,}".ljust(69) + "║")
    print(f"║    Columnas       : {cols}".ljust(69) + "║")

    estado_arch = ("🟢 Pequeño  (<5k filas)"   if filas < 5000
                   else "🟡 Mediano  (5k-30k)"  if filas <= 30000
                   else "🔴 Grande   (>30k — chunks recomendado)")
    print(f"║    Clasificación  : {estado_arch}".ljust(69) + "║")

    pandas_ok = verificar_pandas()
    print("╠" + "═" * 68 + "╣")
    print("║  📦 LIBRERÍAS:".ljust(69) + "║")
    print(f"║    pandas  : {'✅' if pandas_ok else '❌ pip install pandas'}".ljust(69) + "║")
    print(f"║    psutil  : {'✅' if PSUTIL_OK  else '⚠️  pip install psutil'}".ljust(69) + "║")
    print(f"║    openpyxl: {'✅' if OPENPYXL_OK else '❌ pip install openpyxl'}".ljust(69) + "║")

    motor, modo, chunk_size, razon = elegir_motor(ram_disp, arch_mb, filas, pandas_ok)

    iconos = {'TURBO':'🚀','CHUNKS':'📦','SEGURO':'🔧','MÍNIMO':'🐢'}
    vel    = {'TURBO':'~20,000 fact/min','CHUNKS':'~15,000 fact/min',
              'SEGURO':'~5,000 fact/min','MÍNIMO':'~2,000 fact/min'}

    print("╠" + "═" * 68 + "╣")
    print(f"║  {iconos.get(modo,'⚙️')} MOTOR: {motor.upper()} — MODO {modo}".ljust(69) + "║")

    # Mostrar razón en líneas cortas
    palabras, linea = razon.split(), ""
    for p in palabras:
        if len(linea) + len(p) + 1 > 64:
            print(f"║    {linea:<64}║")
            linea = p
        else:
            linea = (linea + " " + p).strip()
    if linea:
        print(f"║    {linea:<64}║")

    if motor == 'pandas_chunks':
        print(f"║    Tamaño de bloque: {chunk_size:,} filas".ljust(69) + "║")
        bloques = -(-filas // chunk_size)
        print(f"║    Bloques totales : ~{bloques}".ljust(69) + "║")

    print(f"║    Velocidad est. : {vel.get(modo,'N/D')}".ljust(69) + "║")
    print("╚" + "═" * 68 + "╝\n")

    return ResultadoAnalisis(
        ram_total_gb=ram_total, ram_disponible_gb=ram_disp,
        cpu_cores=cpu, sistema_operativo=so,
        archivo_mb=arch_mb, filas_reales=filas, columnas=cols,
        motor=motor, modo=modo, razon=razon,
        pandas_disponible=pandas_ok, chunk_size=chunk_size,
    )
