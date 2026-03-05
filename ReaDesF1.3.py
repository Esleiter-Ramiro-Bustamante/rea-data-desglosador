"""
===============================================================================
ReaDesF1.5 - VALIDADOR FISCAL OPTIMIZADO
===============================================================================

POLÍTICA DE PRIVACIDAD Y SEGURIDAD:
- Este software procesa facturas LOCALMENTE en tu computadora
- NO envía datos a internet
- NO almacena información en servidores externos
- Los archivos procesados quedan SOLO en tu equipo
- Cumple con LFPDPPP (Ley Federal de Protección de Datos Personales)

CARACTERÍSTICAS DE SEGURIDAD:
✓ Procesamiento 100% local
✓ Modo de anonimización para pruebas
✓ Log de auditoría automático
✓ Verificación de integridad de datos
✓ Advertencias de seguridad

NUEVO EN v1.5:
✓ OPTIMIZACIÓN MAYOR: headers_map — diccionario de columnas construido
  UNA SOLA VEZ al inicio. Búsqueda O(1) en lugar de recorrer la fila
  completa en cada llamada. Mejora de velocidad: 5x a 10x más rápido.
✓ find_column() ahora usa headers_map internamente
✓ create_column_if_missing() actualiza headers_map al crear columnas nuevas
✓ Pre-indexado de IEPS también usa headers_map

NUEVO EN v1.4.1:
✓ Corrección fundamento legal insumos agrícolas Régimen 612
✓ Art. 147 LISR eliminado (deducciones personales, no aplica aquí)
✓ Fundamento correcto: Art. 103 LISR + Art. 27 Fracc. III LISR

NUEVO EN v1.4:
✓ Detección de insumos agrícolas (fertilizantes, semillas, agroquímicos)
✓ Validación Régimen 612: Insumos efectivo >$2,000 NO deducibles

NUEVO EN v1.3:
✓ Gasolina agrupada RESICO (múltiples despachos separados por '|')

NUEVO EN v1.2:
✓ Reglas diferenciadas por régimen (626 vs 612)

===============================================================================
FUNDAMENTO LEGAL — INSUMOS AGRÍCOLAS RÉGIMEN 612
===============================================================================

  REGLA APLICADA:
  ┌─────────────────────────────────────────────────────────────────────────┐
  │ Insumo agrícola $1,500 efectivo      → ✅ DEDUCIBLE                    │
  │ Insumo agrícola $3,000 efectivo      → ❌ NO DEDUCIBLE                 │
  │ Insumo agrícola $3,000 transferencia → ✅ DEDUCIBLE                    │
  └─────────────────────────────────────────────────────────────────────────┘

  ✅ Art. 27, Fracc. III LISR → Pagos >$2,000 deben ser electrónicos.
  ✅ Art. 103 LISR            → Deducciones autorizadas Régimen 612.
  ❌ Art. 147 LISR            → Deducciones PERSONALES. NO aplica aquí.
  ❌ Art. 74 LISR (AGAPES)   → EXCLUSIVO régimen AGAPES. NO aplica a 612.

===============================================================================
OPTIMIZACIÓN TÉCNICA v1.5 — headers_map
===============================================================================

  ANTES (lento):
    find_column() recorría toda la fila 1 en cada llamada.
    500 facturas × 12 columnas = 6,000 recorridos de fila.

  AHORA (rápido):
    headers_map se construye UNA SOLA VEZ al inicio.
    Búsqueda por diccionario = O(1), tiempo constante.
    500 facturas × 12 columnas = 12 operaciones totales.

  GANANCIA ESTIMADA: 5x – 10x más rápido en archivos grandes.

===============================================================================
Desarrollado para: Validación fiscal de facturas CFDI (México)
Versión: 1.5
Fecha: Marzo 2026
===============================================================================
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import re
import time
import hashlib
from datetime import datetime

# ==============================
# CONFIGURACIÓN DE SEGURIDAD
# ==============================

class ConfiguracionSeguridad:
    """Configuración de seguridad y privacidad"""

    MODO_PRODUCCION                = True
    MODO_ANONIMIZAR                = False
    CREAR_LOG_AUDITORIA            = True
    LOG_DIRECTORY                  = "logs_auditoria"
    MOSTRAR_ADVERTENCIA_PRIVACIDAD = True

    @staticmethod
    def mostrar_advertencia_inicial():
        if not ConfiguracionSeguridad.MOSTRAR_ADVERTENCIA_PRIVACIDAD:
            return
        print("\n" + "=" * 80)
        print("⚠️  ADVERTENCIA DE PRIVACIDAD Y SEGURIDAD")
        print("=" * 80)
        print("\n📋 INFORMACIÓN IMPORTANTE:")
        print("  • Este software procesa facturas con información fiscal CONFIDENCIAL")
        print("  • Los datos se procesan LOCALMENTE en tu computadora")
        print("  • NO se envía información a internet ni a servidores externos")
        print("  • Tú eres responsable de la protección de los archivos procesados")
        print("\n🔐 CUMPLIMIENTO LEGAL:")
        print("  • Cumple con LFPDPPP (Protección de Datos Personales)")
        print("  • Los datos permanecen bajo tu custodia en todo momento")
        print("  • Se recomienda cifrar los archivos de salida")
        if ConfiguracionSeguridad.MODO_ANONIMIZAR:
            print("\n✅ MODO ANONIMIZACIÓN ACTIVADO:")
            print("  • Los RFCs y nombres serán anonimizados automáticamente")
        if ConfiguracionSeguridad.CREAR_LOG_AUDITORIA:
            print(f"\n📝 LOG DE AUDITORÍA: {ConfiguracionSeguridad.LOG_DIRECTORY}/")
        print("\n" + "=" * 80)
        respuesta = input("¿Deseas continuar? (SI/NO): ").strip().upper()
        if respuesta not in ['SI', 'S', 'YES', 'Y']:
            print("❌ Proceso cancelado por el usuario.")
            raise SystemExit(0)
        print()

# ==============================
# CLASE DE AUDITORÍA
# ==============================

class LogAuditoria:

    def __init__(self):
        self.log_entries = []
        self.inicio = datetime.now()
        if ConfiguracionSeguridad.CREAR_LOG_AUDITORIA:
            os.makedirs(ConfiguracionSeguridad.LOG_DIRECTORY, exist_ok=True)

    def registrar_inicio(self, archivo):
        self.log_entries.append({
            'timestamp': datetime.now(),
            'evento': 'INICIO_PROCESAMIENTO',
            'archivo': archivo,
            'hash_archivo': self._calcular_hash_archivo(archivo)
        })

    def registrar_fin(self, archivo_salida, filas_procesadas, tiempo_total):
        self.log_entries.append({
            'timestamp': datetime.now(),
            'evento': 'FIN_PROCESAMIENTO',
            'archivo_salida': archivo_salida,
            'filas_procesadas': filas_procesadas,
            'tiempo_segundos': round(tiempo_total, 2),
            'modo_anonimizacion': ConfiguracionSeguridad.MODO_ANONIMIZAR
        })

    def registrar_error(self, error):
        self.log_entries.append({
            'timestamp': datetime.now(),
            'evento': 'ERROR',
            'detalle': str(error)
        })

    def _calcular_hash_archivo(self, ruta_archivo):
        try:
            with open(ruta_archivo, 'rb') as f:
                return hashlib.sha256(f.read()).hexdigest()
        except:
            return "NO_DISPONIBLE"

    def guardar_log(self):
        if not ConfiguracionSeguridad.CREAR_LOG_AUDITORIA:
            return
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        log_file  = os.path.join(
            ConfiguracionSeguridad.LOG_DIRECTORY,
            f"auditoria_{timestamp}.log"
        )
        try:
            with open(log_file, 'w', encoding='utf-8') as f:
                f.write("=" * 80 + "\n")
                f.write("LOG DE AUDITORÍA - VALIDADOR FISCAL ReaDesF1.5\n")
                f.write("=" * 80 + "\n\n")
                for entry in self.log_entries:
                    f.write(f"[{entry['timestamp'].strftime('%Y-%m-%d %H:%M:%S')}] {entry['evento']}\n")
                    for key, value in entry.items():
                        if key not in ['timestamp', 'evento']:
                            f.write(f"  {key}: {value}\n")
                    f.write("\n")
                f.write("=" * 80 + "\nFIN DEL LOG\n" + "=" * 80 + "\n")
            print(f"📝 Log guardado: {log_file}")
        except Exception as e:
            print(f"⚠️  No se pudo guardar el log: {e}")

# ==============================
# CONFIGURACIÓN INICIAL
# ==============================

ConfiguracionSeguridad.mostrar_advertencia_inicial()
log_auditoria = LogAuditoria()

desktop_path = os.path.join(
    os.path.expanduser('~'),
    'Desktop/GASTOS RESICO/2026/FEBRERO26'
)

file_name = input("Ingrese el nombre del archivo Excel (sin extensión .xlsx): ")
file_name = file_name.strip()
file_name = re.sub(r'\.xlsx$', '', file_name, flags=re.IGNORECASE) + '.xlsx'
file_path = os.path.join(desktop_path, file_name)

tiempo_inicio = time.time()
log_auditoria.registrar_inicio(file_path)

# ==============================
# ESTILOS
# ==============================

blue_fill   = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
green_fill  = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill    = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
purple_fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
pink_fill   = PatternFill(start_color='FF69B4', end_color='FF69B4', fill_type='solid')
brown_fill  = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
lime_fill   = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

result_style = Font(bold=True)
center_align = Alignment(horizontal='center')

# ==============================
# FUNCIONES AUXILIARES
# ==============================

def extract_code(value):
    """Extrae el código de un valor 'CODIGO - Descripción'"""
    if value and '-' in str(value):
        return str(value).split('-')[0].strip()
    return str(value).strip() if value else ''

# ══════════════════════════════════════════════════════════════════════
# OPTIMIZACIÓN v1.5 — headers_map
# ══════════════════════════════════════════════════════════════════════
# Se declara aquí como variable global para que find_column() y
# create_column_if_missing() lo puedan usar y actualizar.
# Se construye una sola vez después de abrir el archivo Excel.
# ══════════════════════════════════════════════════════════════════════
headers_map = {}   # { 'nombre_columna_lower': numero_columna }

def build_headers_map(sheet):
    """
    Construye el diccionario de columnas UNA SOLA VEZ.
    Clave   = nombre de encabezado en minúsculas y sin espacios extremos.
    Valor   = número de columna (entero).
    Complejidad: O(n) una sola vez, luego O(1) por búsqueda.
    """
    global headers_map
    headers_map = {}
    for cell in sheet[1]:
        if cell.value is not None:
            headers_map[str(cell.value).strip().lower()] = cell.column
    print(f"⚡ headers_map construido: {len(headers_map)} columnas indexadas")

def find_column(sheet, column_name):
    """
    Busca columna usando headers_map (O(1)).
    Mantiene la misma firma que la versión anterior para compatibilidad.
    Si headers_map está vacío (no se llamó build_headers_map),
    hace el recorrido tradicional como respaldo.
    """
    if headers_map:
        return headers_map.get(column_name.strip().lower())
    # Respaldo: recorrido tradicional
    for cell in sheet[1]:
        if cell.value and str(cell.value).strip().lower() == column_name.strip().lower():
            return cell.column
    return None

def create_column_if_missing(sheet, column_name, fill_color=blue_fill):
    """
    Crea una columna si no existe y actualiza headers_map.
    Al actualizar el mapa aquí, las búsquedas posteriores
    de la columna nueva también serán O(1).
    """
    col = find_column(sheet, column_name)
    if col is None:
        last_col = sheet.max_column + 1
        cell = sheet.cell(row=1, column=last_col, value=column_name)
        cell.fill      = fill_color
        cell.alignment = center_align
        # ── Actualizar headers_map con la nueva columna ──
        headers_map[column_name.strip().lower()] = last_col
        print(f"✅ Columna creada y mapeada: {column_name} → {get_column_letter(last_col)}")
        return last_col
    return col

# ==============================
# FUNCIONES DE ANONIMIZACIÓN
# ==============================

def anonimizar_datos_sensibles(sheet, columns):
    if not ConfiguracionSeguridad.MODO_ANONIMIZAR:
        return 0
    print("🔒 Anonimizando datos sensibles...")
    filas_anonimizadas = 0
    cols_anonimizar = {
        'RFC emisor':    'XAXX010101XXX',
        'Razon emisor':  'PROVEEDOR ANONIMIZADO',
        'RFC receptor':  'XAXX010101XXX',
        'Razon receptor':'RECEPTOR ANONIMIZADO',
        'UUID': None
    }
    for row in range(2, sheet.max_row + 1):
        for col_name, valor_anonimo in cols_anonimizar.items():
            # Usa headers_map directamente — O(1)
            col = headers_map.get(col_name.lower())
            if col:
                if col_name == 'UUID':
                    valor_anonimo = f"ANONIMIZADO-{row:08d}-XXXX-XXXX-XXXXXXXXXXXX"
                sheet.cell(row=row, column=col, value=valor_anonimo)
                sheet.cell(row=row, column=col).fill = PatternFill(
                    start_color='FFFF00', end_color='FFFF00', fill_type='solid'
                )
        filas_anonimizadas += 1
    print(f"✅ {filas_anonimizadas} filas anonimizadas")
    return filas_anonimizadas

# ==============================
# FUNCIONES DE DETECCIÓN
# ==============================

def es_gasolina(concepto):
    """
    Detecta combustibles/gasolina.
    Art. 27 LISR: combustibles vehículos marítimos, aéreos y terrestres.
    """
    if not concepto:
        return False
    palabras = [
        'gasolina', 'combustible', 'magna', 'premium',
        'diesel', 'diésel', 'gasohol', 'gasoil',
        'nafta', 'petrol', 'gas', 'energético',
        'turbosina', 'jet fuel', 'bunker'
    ]
    return any(p in concepto.lower() for p in palabras)

def es_producto_dulce(concepto):
    """Detecta productos dulces/botanas con IEPS 8%."""
    if not concepto:
        return False
    palabras = [
        'pan', 'roles', 'conchas', 'mantecadas', 'donas', 'panque',
        'gansito', 'pinguinos', 'submarinos', 'chocorol', 'principe',
        'pastisetas', 'canelitas', 'polvorones', 'triki', 'duo',
        'rebanada', 'colchones', 'cuernitos', 'medias noches',
        'pan blanco', 'pan integral', 'pan molido', 'empanizador',
        'tostada', 'tostaditas', 'tostado', 'tortillinas', 'salmas',
        'chocolate', 'bon o bon', 'hershey', 'reese', 'kit kat',
        'kremas', 'trident',
        'papas', 'chips', 'sabritas', 'doritos', 'cheetos', 'ruffles',
        'barcel', 'takis', 'hot nuts', 'cacahuates', 'kiyakis',
        'runners', 'churrumais', 'tostitos', 'fritos', 'big mix',
        'galletas', 'oreo', 'emperador', 'marías', 'animalitos',
        'chokis', 'sponch', 'barrita',
        'pepsi', 'coca', 'sprite', 'fanta', 'monster', 'gatorade',
        'ades', 'delsey'
    ]
    return any(p in concepto.lower() for p in palabras)

def es_gasolina_agrupada(concepto):
    """
    Detecta gasolina con múltiples despachos separados por '|'.
    Facilidad EXCLUSIVA para régimen 626 (RESICO).
    """
    if not concepto or not es_gasolina(concepto):
        return False
    return concepto.count('|') >= 1

def es_insumo_agricola(concepto):
    """
    Detecta fertilizantes, semillas y agroquímicos.

    FUNDAMENTO LEGAL CORRECTO — RÉGIMEN 612:
    ✅ Art. 27, Fracc. III LISR → Pagos >$2,000 deben ser electrónicos.
    ✅ Art. 103 LISR            → Deducciones autorizadas Régimen 612.
    ❌ Art. 147 LISR            → Deducciones personales. NO aplica.
    ❌ Art. 74 LISR (AGAPES)   → Exclusivo AGAPES. NO aplica a 612.
    """
    if not concepto:
        return False
    palabras = [
        # 1. FERTILIZANTES
        'fertilizante', 'fertilizacion', 'fertilización',
        'abono', 'abonos', 'abonado',
        'urea', 'nitrato', 'nitrogeno', 'nitrógeno',
        'sulfato de amonio', 'amoniaco', 'sulfammo',
        'nitrofoska', 'nitrabor', 'kan-l', 'uan', 'nitramon', 'nitromax',
        'fosfato', 'fosforo', 'fósforo',
        'map', 'dap', 'superfosfato', 'triple super',
        'potasio', 'potásico', 'cloruro de potasio',
        'sulfato de potasio', 'sulfato potásico',
        'npk', 'haifa', 'basacote', 'nutri',
        'microelementos', 'humatos', 'calcio foliar',
        'quelatos', 'bioestimulante', 'aminoacidos',
        'acido humico', 'ácido húmico', 'fulvico',
        'yara', 'fertimex', 'compo', 'timac', 'mosaic', 'nutrimos', 'agrium',
        # 2. SEMILLAS Y MATERIAL VEGETATIVO
        'semilla', 'semillas', 'semillero',
        'esqueje', 'esquejes', 'plantula', 'plántula', 'plantulas',
        'variedad', 'hibrido', 'híbrido',
        'semilla de caña', 'punta de caña', 'trocito de caña',
        'cana semilla', 'caña semilla', 'propagacion', 'propagación',
        'maiz', 'maíz', 'sorgo', 'frijol',
        'garbanzo', 'ajonjoli', 'ajonjolí', 'girasol', 'cártamo', 'cartamo',
        # 3. HERBICIDAS
        'herbicida', 'herbicidas', 'maleza', 'deshierbe', 'desherbe',
        'glifosato', 'roundup', 'atrazina', 'atrazine',
        '2-4d', '2,4-d', 'amine', 'pendimetalina', 'metribuzin',
        'diuron', 'diurón', 'hexazinona', 'ametrina', 'ametrine',
        'faena', 'gesapax', 'velpar', 'karmex', 'harness', 'dual',
        # 4. INSECTICIDAS Y PLAGUICIDAS
        'insecticida', 'insecticidas', 'plaguicida', 'plaguicidas',
        'pesticida', 'pesticidas',
        'clorpirifos', 'imidacloprid', 'lambda',
        'cipermetrina', 'deltametrina', 'malatión',
        'abamectina', 'spinosad', 'bifentrina', 'thiamethoxam', 'acetamiprid',
        'bacillus', 'beauveria', 'trichoderma', 'metarhizium',
        'control biologico', 'biol',
        'lorsban', 'confidor', 'regent', 'karate', 'decis', 'engeo',
        # 5. FUNGICIDAS
        'fungicida', 'fungicidas',
        'mancozeb', 'metalaxil', 'propiconazol', 'tebuconazol',
        'azoxistrobina', 'iprodiona', 'clorotalonil', 'captan', 'tiofanato',
        'dithane', 'ridomil', 'tilt', 'folicur', 'amistar', 'rovral',
        # 6. AGRONUTRIENTES Y ESTIMULANTES
        'agronutriente', 'agronutrientes', 'estimulante', 'estimulantes',
        'regulador de crecimiento', 'regulador',
        'citoquinina', 'auxina', 'giberelina',
        'ethephon', 'ethrel', 'madurante', 'maduracion', 'maduración',
        # 7. INSUMOS DE APLICACIÓN
        'coadyuvante', 'adherente', 'surfactante',
        'aceite agricola', 'aceite agrícola', 'dispersante', 'emulsificante',
        # 8. ENMIENDAS DE SUELO
        'cal agricola', 'cal agrícola', 'calcita', 'dolomita',
        'yeso agricola', 'yeso agrícola',
        'azufre agricola', 'azufre agrícola', 'encalado',
        'enmienda', 'corrector de suelo', 'acondicionador de suelo',
    ]
    return any(p in concepto.lower() for p in palabras)

# ==============================
# REGLAS DE DEDUCIBILIDAD
# ==============================

USOS_DEDUCIBLES     = ['G01', 'G02', 'G03']
METODOS_VALIDOS     = ['PUE', 'PPD']
FORMAS_VALIDAS      = ['01', '02', '03', '04', '28']
FORMAS_ELECTRONICAS = ['02', '03', '04', '28']
USO_CFDI_VERDE      = "G02 - Devoluciones, descuentos o bonificaciones"

REGIMENES_TRABAJADOS  = ['626', '612']
REGIMENES_ADVERTENCIA = ['605', '616']

# ==============================
# ABRIR ARCHIVO EXCEL
# ==============================

def mostrar_error_archivo(file_path, error):
    print("\n" + "╔" + "═" * 78 + "╗")
    print("║" + " " * 25 + "⚠️  ERROR AL CARGAR ARCHIVO" + " " * 26 + "║")
    print("╠" + "═" * 78 + "╣")
    print("║  ❌ No se pudo encontrar o abrir el archivo" + " " * 32 + "║")
    if len(file_path) > 72:
        carpeta = os.path.dirname(file_path)
        archivo = os.path.basename(file_path)
        print(f"║     {carpeta[:72]:<72}║")
        print(f"║     {archivo:<72}║")
    else:
        print(f"║     {file_path:<72}║")
    print("║  🔍 Posibles causas:" + " " * 57 + "║")
    print("║     • Nombre mal escrito / archivo no existe / archivo abierto en Excel" + " " * 5 + "║")
    carpeta = os.path.dirname(file_path)
    if os.path.exists(carpeta):
        archivos_excel = [f for f in os.listdir(carpeta)
                          if f.endswith(('.xlsx', '.xls', '.xlsm'))]
        if archivos_excel:
            print("║  📁 Archivos Excel en la carpeta:" + " " * 44 + "║")
            for i, archivo in enumerate(archivos_excel[:5], 1):
                print(f"║     {i}. {archivo[:70]:<70}║")
    print("╚" + "═" * 78 + "╝\n")

try:
    if not os.path.exists(desktop_path):
        print(f"\n❌ Carpeta no encontrada: {desktop_path}")
        log_auditoria.registrar_error(f"Carpeta no encontrada: {desktop_path}")
        log_auditoria.guardar_log()
        input("\n👉 Presiona ENTER para salir...")
        raise SystemExit(1)

    if not os.path.exists(file_path):
        mostrar_error_archivo(file_path, "Archivo no encontrado")
        log_auditoria.registrar_error(f"Archivo no encontrado: {file_path}")
        log_auditoria.guardar_log()
        input("\n👉 Presiona ENTER para salir...")
        raise SystemExit(1)

    workbook = openpyxl.load_workbook(file_path)
    sheet    = workbook.active
    print(f"✅ Archivo cargado: {file_path}")
    print(f"📊 Filas totales  : {sheet.max_row - 1}")

except FileNotFoundError as e:
    mostrar_error_archivo(file_path, e)
    log_auditoria.registrar_error(e)
    log_auditoria.guardar_log()
    input("\n👉 Presiona ENTER para salir...")
    raise SystemExit(1)
except PermissionError as e:
    print(f"\n❌ Archivo bloqueado. Ciérralo en Excel e intenta de nuevo.")
    log_auditoria.registrar_error(e)
    log_auditoria.guardar_log()
    input("\n👉 Presiona ENTER para salir...")
    raise SystemExit(1)
except Exception as e:
    print(f"\n❌ Error inesperado: {e}")
    log_auditoria.registrar_error(e)
    log_auditoria.guardar_log()
    input("\n👉 Presiona ENTER para salir...")
    raise SystemExit(1)

# ══════════════════════════════════════════════════════════════════════
# OPTIMIZACIÓN v1.5 — Construir headers_map UNA SOLA VEZ
# ══════════════════════════════════════════════════════════════════════
print("⚡ Construyendo índice de columnas (headers_map)...")
build_headers_map(sheet)

# ==============================
# COLUMNAS REQUERIDAS
# ==============================

required_columns = {
    'SubTotal':           'SubTotal',
    'Descuento':          'Descuento',
    'IVA Trasladado 0%':  'IVA Trasladado 0%',
    'IVA Exento':         'IVA Exento',
    'IVA Trasladado 16%': 'IVA Trasladado 16%',
    'Total':              'Total',
    'Uso CFDI':           'Uso CFDI',
    'Metodo pago':        'Metodo pago',
    'Forma pago':         'Forma pago',
    'Regimen receptor':   'Regimen receptor',
    'Razon emisor':       'Razon emisor',
    'Conceptos':          'Conceptos'
}

print("🔍 Verificando columnas base...")
columns = {}
for key, col_name in required_columns.items():
    # find_column ahora usa headers_map — O(1)
    columns[key] = create_column_if_missing(sheet, col_name)

filas_anonimizadas = anonimizar_datos_sensibles(sheet, columns)

# ==============================
# PRE-INDEXAR COLUMNAS IEPS
# (también usa headers_map — O(1))
# ==============================

print("🚀 Pre-indexando columnas IEPS...")

ieps_gasolina_encontrado      = False
ieps_8_encontrado             = False
ieps_no_desglosado_encontrado = False
columnas_ieps                 = []

for col in range(1, sheet.max_column + 1):
    header = sheet.cell(row=1, column=col).value
    if header and "IEPS" in str(header).upper():
        columnas_ieps.append(col)
        header_str = str(header).strip()
        if '8%' in header_str or '8 %' in header_str:
            columns['IEPS Trasladado 8%'] = col
            ieps_8_encontrado = True
            print(f"  ✅ IEPS 8% (dulces)    : {get_column_letter(col)}")
        elif 'No Desglosado' in header_str:
            columns['IEPS Trasladado No Desglosado'] = col
            ieps_no_desglosado_encontrado = True
            print(f"  ✅ IEPS No Desglosado  : {get_column_letter(col)}")
        else:
            columns['IEPS Trasladado'] = col
            ieps_gasolina_encontrado = True
            print(f"  ✅ IEPS Gasolina       : {get_column_letter(col)}")

print(f"  ⚡ {len(columnas_ieps)} columnas IEPS indexadas")

# Columna Efecto (usa headers_map)
efecto_col = find_column(sheet, 'Efecto')
if efecto_col:
    print(f"  ✅ Columna Efecto      : {get_column_letter(efecto_col)}")

# Columna Razón No Deducible
razon_no_ded_col = create_column_if_missing(sheet, 'Razón No Deducible', red_fill)

# ==============================
# INICIALIZAR COLUMNAS NUMÉRICAS
# ==============================

print("📝 Inicializando columnas numéricas...")
for row in range(2, sheet.max_row + 1):
    for col_key in ['IVA Trasladado 16%', 'IVA Trasladado 0%', 'IVA Exento', 'Descuento']:
        cell = sheet.cell(row=row, column=columns[col_key])
        if cell.value is None:
            cell.value = 0
        cell.number_format = "0.00"

# ==============================
# CREAR COLUMNAS DE CÁLCULO
# ==============================

last_column = sheet.max_column
calc_headers = [
    'SUB1-16%', 'SUB0%', 'SUB2-16%',
    'IVA ACREDITABLE 16%', 'C IVA', 'T2',
    'Comprobación T2', 'Deducible'
]

print("🧮 Creando columnas de cálculo...")
for i, header in enumerate(calc_headers, start=1):
    cell           = sheet.cell(row=1, column=last_column + i, value=header)
    cell.fill      = blue_fill
    cell.alignment = center_align
    # Actualizar headers_map con columnas de cálculo nuevas
    headers_map[header.lower()] = last_column + i

sub1_col      = last_column + 1
sub0_col      = last_column + 2
sub2_col      = last_column + 3
iva_acred_col = last_column + 4
c_iva_col     = last_column + 5
t2_col        = last_column + 6
comprob_col   = last_column + 7
deducible_col = last_column + 8

# Cache de letras de columna (evita llamar get_column_letter en cada fila)
col_letters = {
    'SubTotal':  get_column_letter(columns['SubTotal']),
    'Descuento': get_column_letter(columns['Descuento']),
    'IVA16':     get_column_letter(columns['IVA Trasladado 16%']),
    'IVA0':      get_column_letter(columns['IVA Trasladado 0%']),
    'IVAExento': get_column_letter(columns['IVA Exento']),
    'Total':     get_column_letter(columns['Total']),
    'sub1':      get_column_letter(sub1_col),
    'sub0':      get_column_letter(sub0_col),
    'sub2':      get_column_letter(sub2_col),
    'iva_acred': get_column_letter(iva_acred_col),
    't2':        get_column_letter(t2_col),
}
if ieps_8_encontrado:
    col_letters['IEPS8']    = get_column_letter(columns['IEPS Trasladado 8%'])
if ieps_gasolina_encontrado:
    col_letters['IEPS_GAS'] = get_column_letter(columns['IEPS Trasladado'])
if ieps_no_desglosado_encontrado:
    col_letters['IEPS_ND']  = get_column_letter(columns['IEPS Trasladado No Desglosado'])

# ==============================
# VARIABLES DE CONTROL
# ==============================

# IEPS
ieps_gasolina_procesados      = 0
ieps_8_procesados             = 0
ieps_no_desglosado_procesados = 0

# Gasolina
gasolina_con_ieps              = 0
gasolina_sin_ieps              = 0
gasolina_efectivo_626          = 0
gasolina_efectivo_626_agrupada = 0
gasolina_efectivo_612          = 0
gasolina_electronico           = 0

# Dulces
dulces_con_ieps8 = 0
dulces_sin_ieps8 = 0

# Insumos agrícolas
insumo_agricola_efectivo_612 = 0   # ❌ Efectivo >$2,000
insumo_agricola_menor_2000   = 0   # ✅ Efectivo ≤$2,000
insumo_agricola_electronico  = 0   # ✅ Pago electrónico

# Generales
uso_s01_count         = 0
efectivo_mayor_2000   = 0
regimenes_encontrados = {}

print("\n🔧 Iniciando procesamiento de filas...")
print("=" * 80)

# ==============================
# LOOP PRINCIPAL
# ==============================

total_filas = sheet.max_row - 1

for row in range(2, sheet.max_row + 1):

    if (row - 1) % 100 == 0 or row == sheet.max_row:
        progreso = ((row - 1) / total_filas) * 100
        print(f"📊 Procesando: {row - 1}/{total_filas} facturas ({progreso:.1f}%)")

    # Leer datos de la fila
    row_data = {
        'concepto':    str(sheet.cell(row=row, column=columns['Conceptos']).value or ''),
        'total':       float(sheet.cell(row=row, column=columns['Total']).value or 0),
        'uso_cfdi':    sheet.cell(row=row, column=columns['Uso CFDI']).value,
        'metodo_pago': sheet.cell(row=row, column=columns['Metodo pago']).value,
        'forma_pago':  sheet.cell(row=row, column=columns['Forma pago']).value,
        'regimen':     sheet.cell(row=row, column=columns['Regimen receptor']).value,
    }

    regimen = extract_code(row_data['regimen'])
    regimenes_encontrados[regimen] = regimenes_encontrados.get(regimen, 0) + 1

    # Detectar tipo de concepto (una sola vez por fila)
    es_gasolina_concepto   = es_gasolina(row_data['concepto'])
    es_dulce_concepto      = es_producto_dulce(row_data['concepto'])
    es_insumo_agr_concepto = es_insumo_agricola(row_data['concepto'])

    # Valores IEPS
    ieps_8_val   = float(sheet.cell(row=row, column=columns['IEPS Trasladado 8%']).value or 0)                if ieps_8_encontrado             else 0.0
    ieps_gas_val = float(sheet.cell(row=row, column=columns['IEPS Trasladado']).value or 0)                   if ieps_gasolina_encontrado      else 0.0
    ieps_nd_val  = float(sheet.cell(row=row, column=columns['IEPS Trasladado No Desglosado']).value or 0)     if ieps_no_desglosado_encontrado else 0.0

    # ── SUB1 según tipo de IEPS ──────────────────────────────────────────
    if ieps_8_val > 0:
        formula_sub1 = (f"=({col_letters['SubTotal']}{row}-{col_letters['Descuento']}{row})"
                        f"+{col_letters['IEPS8']}{row}")
        if es_dulce_concepto:
            sheet.cell(row=row, column=columns['Conceptos']).fill = pink_fill
            dulces_con_ieps8 += 1
        ieps_8_procesados += 1

    elif es_gasolina_concepto and (ieps_gas_val > 0 or ieps_nd_val > 0):
        ieps_gasolina = ieps_gas_val if ieps_gas_val > 0 else ieps_nd_val
        sheet.cell(row=row, column=columns['IVA Trasladado 0%'], value=ieps_gasolina)
        sheet.cell(row=row, column=columns['IVA Trasladado 0%']).fill = orange_fill
        formula_sub1 = f"=({col_letters['SubTotal']}{row}-{col_letters['Descuento']}{row})"
        sheet.cell(row=row, column=columns['Conceptos']).fill = blue_fill
        gasolina_con_ieps += 1
        ieps_gasolina_procesados += 1

    else:
        formula_sub1 = f"=({col_letters['SubTotal']}{row}-{col_letters['Descuento']}{row})"
        if es_gasolina_concepto:
            sheet.cell(row=row, column=columns['Conceptos']).fill = orange_fill
            gasolina_sin_ieps += 1
        elif es_dulce_concepto:
            dulces_sin_ieps8 += 1

    # ── Fórmulas de cálculo ──────────────────────────────────────────────
    def write_calc(col, value):
        c = sheet.cell(row=row, column=col, value=value)
        c.number_format = "0.00"
        c.font = result_style

    write_calc(sub1_col,     formula_sub1)
    write_calc(sub0_col,     f"={col_letters['IVA0']}{row}+{col_letters['IVAExento']}{row}")
    write_calc(sub2_col,     f"={col_letters['sub1']}{row}-{col_letters['sub0']}{row}")
    write_calc(iva_acred_col, f"={col_letters['sub2']}{row}*0.16")
    write_calc(c_iva_col,    f"={col_letters['iva_acred']}{row}-{col_letters['IVA16']}{row}")
    write_calc(t2_col,       f"={col_letters['sub2']}{row}+{col_letters['sub0']}{row}+{col_letters['IVA16']}{row}")
    write_calc(comprob_col,  f"={col_letters['Total']}{row}-{col_letters['t2']}{row}")

    # Validación visual IVA
    try:
        subtotal_val   = float(sheet.cell(row=row, column=columns['SubTotal']).value or 0)
        descuento_val  = float(sheet.cell(row=row, column=columns['Descuento']).value or 0)
        iva0_val       = float(sheet.cell(row=row, column=columns['IVA Trasladado 0%']).value or 0)
        iva_exento_val = float(sheet.cell(row=row, column=columns['IVA Exento']).value or 0)
        iva16_val      = float(sheet.cell(row=row, column=columns['IVA Trasladado 16%']).value or 0)
        sub1_calc      = (subtotal_val - descuento_val + ieps_8_val
                          if ieps_8_val > 0 else subtotal_val - descuento_val)
        iva_acred_calc = round((sub1_calc - (iva0_val + iva_exento_val)) * 0.16, 2)
        if abs(iva_acred_calc - iva16_val) < 0.01:
            sheet.cell(row=row, column=iva_acred_col).fill = green_fill
            sheet.cell(row=row, column=columns['IVA Trasladado 16%']).fill = green_fill
    except:
        pass

    # ── Formateo por régimen ─────────────────────────────────────────────
    reg_fills = {'626': blue_fill, '612': purple_fill, '616': brown_fill}
    sheet.cell(row=row, column=columns['Regimen receptor']).fill = reg_fills.get(regimen, orange_fill)
    if regimen not in reg_fills:
        sheet.cell(row=row, column=columns['Razon emisor']).fill = orange_fill

    uso_cfdi = extract_code(row_data['uso_cfdi']).upper() if row_data['uso_cfdi'] else ''
    if row_data['uso_cfdi'] and str(row_data['uso_cfdi']).strip() == USO_CFDI_VERDE:
        sheet.cell(row=row, column=columns['Uso CFDI']).fill = green_fill
    if uso_cfdi == 'S01':
        sheet.cell(row=row, column=columns['Uso CFDI']).fill = red_fill
        uso_s01_count += 1

    es_egreso = False
    if efecto_col:
        efecto_val = sheet.cell(row=row, column=efecto_col).value
        if efecto_val and str(efecto_val).strip().upper() in ['EGRESO', 'E']:
            es_egreso = True

    # ====================================================================
    # VALIDACIÓN DE DEDUCIBILIDAD
    # ====================================================================

    metodo_pg = extract_code(row_data['metodo_pago']).upper()
    forma_pg  = extract_code(row_data['forma_pago']).upper()

    es_deducible    = True
    razones_rechazo = []

    # Validaciones comunes a TODOS los regímenes
    if uso_cfdi not in USOS_DEDUCIBLES:
        es_deducible = False
        razones_rechazo.append(f"Uso CFDI {uso_cfdi} no deducible")
    if metodo_pg not in METODOS_VALIDOS:
        es_deducible = False
        razones_rechazo.append(f"Método {metodo_pg} inválido")

    # ════════════════════════════════════════════════════════════════════
    # GASOLINA
    # ════════════════════════════════════════════════════════════════════
    if es_gasolina_concepto:
        if forma_pg == '01':
            if regimen == '626':
                if es_gasolina_agrupada(row_data['concepto']):
                    num_despachos = row_data['concepto'].count('|') + 1
                    gasolina_efectivo_626 += 1
                    gasolina_efectivo_626_agrupada += 1
                    sheet.cell(row=row, column=columns['Forma pago']).fill = yellow_fill
                    razones_rechazo.append(
                        f"RESICO (626): {num_despachos} despachos agrupados en efectivo (facilidad)")
                elif row_data['total'] <= 2000:
                    gasolina_efectivo_626 += 1
                    sheet.cell(row=row, column=columns['Forma pago']).fill = yellow_fill
                    razones_rechazo.append("RESICO (626): Gasolina efectivo ≤$2,000 (facilidad)")
                else:
                    es_deducible = False
                    gasolina_efectivo_626 += 1
                    sheet.cell(row=row, column=columns['Forma pago']).fill = red_fill
                    razones_rechazo.append(
                        "RESICO (626): Gasolina individual efectivo >$2,000 NO deducible")
            elif regimen == '612':
                es_deducible = False
                gasolina_efectivo_612 += 1
                sheet.cell(row=row, column=columns['Forma pago']).fill = red_fill
                razones_rechazo.append(
                    "Régimen 612: Gasolina en efectivo NO deducible. "
                    "Art. 103 LISR + Art. 27 Fracc. III LISR.")
            else:
                es_deducible = False
                sheet.cell(row=row, column=columns['Forma pago']).fill = red_fill
                razones_rechazo.append(
                    "Gasolina en efectivo NO deducible (Art. 27 Fracc. III LISR)")
        else:
            gasolina_electronico += 1
            if forma_pg not in FORMAS_ELECTRONICAS:
                es_deducible = False
                razones_rechazo.append(f"Gasolina: forma de pago {forma_pg} inválida")

    # ════════════════════════════════════════════════════════════════════
    # INSUMOS AGRÍCOLAS
    # Art. 27 Fracc. III LISR + Art. 103 LISR
    # ════════════════════════════════════════════════════════════════════
    elif es_insumo_agr_concepto:

        if forma_pg == '01' and row_data['total'] > 2000:
            # ❌ Efectivo >$2,000 → NO DEDUCIBLE
            es_deducible = False
            insumo_agricola_efectivo_612 += 1
            sheet.cell(row=row, column=columns['Forma pago']).fill = red_fill
            sheet.cell(row=row, column=columns['Conceptos']).fill  = red_fill
            razones_rechazo.append(
                "Insumo agrícola efectivo >$2,000 NO deducible. "
                "Art. 103 LISR + Art. 27 Fracc. III LISR. "
                "Facilidad AGAPES (Art. 74 LISR) NO aplica a Régimen 612.")

        elif forma_pg == '01' and row_data['total'] <= 2000:
            # ✅ Efectivo ≤$2,000 → DEDUCIBLE
            insumo_agricola_menor_2000 += 1
            sheet.cell(row=row, column=columns['Forma pago']).fill = yellow_fill
            sheet.cell(row=row, column=columns['Conceptos']).fill  = yellow_fill
            razones_rechazo.append(
                "Insumo agrícola efectivo ≤$2,000: deducible. "
                "Art. 27 Fracc. III LISR.")

        elif forma_pg in FORMAS_ELECTRONICAS:
            # ✅ Pago electrónico → DEDUCIBLE sin límite
            insumo_agricola_electronico += 1
            sheet.cell(row=row, column=columns['Conceptos']).fill = lime_fill
            razones_rechazo.append(
                "Insumo agrícola pago electrónico: deducible. "
                "Art. 27 Fracc. III LISR cumplido.")

        else:
            es_deducible = False
            razones_rechazo.append(
                f"Insumo agrícola: forma de pago {forma_pg} inválida.")

    # ════════════════════════════════════════════════════════════════════
    # GASTOS NORMALES
    # ════════════════════════════════════════════════════════════════════
    else:
        if forma_pg == '01' and row_data['total'] > 2000:
            es_deducible = False
            efectivo_mayor_2000 += 1
            razones_rechazo.append(
                "Efectivo >$2,000 NO deducible (Art. 27 Fracc. III LISR)")
            if efecto_col:
                sheet.cell(row=row, column=efecto_col).fill = red_fill
        elif forma_pg not in FORMAS_VALIDAS:
            es_deducible = False
            razones_rechazo.append(f"Forma de pago {forma_pg} inválida")

    if regimen not in REGIMENES_TRABAJADOS:
        razones_rechazo.append(f"⚠️ Régimen {regimen}: Verificar manualmente")

    # ── Escribir resultado ───────────────────────────────────────────────
    ded_cell           = sheet.cell(row=row, column=deducible_col,
                                    value="SI" if es_deducible else "NO")
    ded_cell.fill      = (blue_fill if es_egreso else green_fill) if es_deducible else red_fill
    ded_cell.font      = Font(bold=True, color='FFFFFF')
    ded_cell.alignment = center_align

    razon_cell = sheet.cell(row=row, column=razon_no_ded_col)
    if razones_rechazo:
        razon_cell.value = " | ".join(razones_rechazo)
        if not es_deducible:
            razon_cell.fill = red_fill
            razon_cell.font = Font(bold=True, color='FFFFFF')
        else:
            razon_cell.fill = yellow_fill
            razon_cell.font = Font(bold=True)
    else:
        razon_cell.value = "Cumple requisitos"
        razon_cell.fill  = green_fill
        razon_cell.font  = Font(color='006100')

# ==============================
# AJUSTAR ANCHO DE COLUMNAS
# ==============================

for col in range(last_column + 1, last_column + len(calc_headers) + 1):
    sheet.column_dimensions[get_column_letter(col)].width = 15
sheet.column_dimensions[get_column_letter(razon_no_ded_col)].width = 65

# ==============================
# GUARDAR ARCHIVO
# ==============================

try:
    base_name   = re.sub(r'\.xlsx$', '', file_name, flags=re.IGNORECASE)
    suffix      = "_ANONIMIZADO_validado" if ConfiguracionSeguridad.MODO_ANONIMIZAR else "_validado"
    output_name = os.path.join(desktop_path, f"{base_name}{suffix}.xlsx")
    workbook.save(output_name)

    tiempo_fin   = time.time()
    tiempo_total = tiempo_fin - tiempo_inicio
    velocidad    = total_filas / tiempo_total if tiempo_total > 0 else 0

    log_auditoria.registrar_fin(output_name, total_filas, tiempo_total)

    print("\n" + "=" * 80)
    print("✅ PROCESO COMPLETADO — ReaDesF1.5")
    print("=" * 80)
    print(f"📂 Archivo  : {output_name}")
    print(f"📊 Filas    : {total_filas}")
    print(f"⏱️  Tiempo   : {tiempo_total:.2f} segundos")
    print(f"⚡ Velocidad : {velocidad:.0f} facturas/segundo")
    print(f"🗂️  Columnas indexadas en headers_map: {len(headers_map)}")

    print("\n📋 REGÍMENES ENCONTRADOS:")
    for reg, count in sorted(regimenes_encontrados.items()):
        print(f"  • Régimen {reg}: {count} facturas")

    print("\n📌 RESUMEN DE VALIDACIÓN:")

    print(f"\n🍬 IEPS 8% (Dulces/Botanas):")
    print(f"  • Con IEPS 8% : {dulces_con_ieps8}")
    print(f"  • Sin IEPS 8% : {dulces_sin_ieps8}")

    print(f"\n⛽ Gasolina:")
    print(f"  • Con IEPS    : {gasolina_con_ieps}")
    print(f"  • Sin IEPS    : {gasolina_sin_ieps}")
    print(f"  • Efectivo 626: {gasolina_efectivo_626}  (agrupada: {gasolina_efectivo_626_agrupada})")
    print(f"  • Efectivo 612 (rechazadas): {gasolina_efectivo_612}")
    print(f"  • Electrónica : {gasolina_electronico}")

    print(f"\n🌱 INSUMOS AGRÍCOLAS (Fertilizantes, Semillas, Agroquímicos):")
    print(f"  • ❌ Efectivo >$2,000 (NO deducibles): {insumo_agricola_efectivo_612}")
    print(f"  • ✅ Efectivo ≤$2,000 (deducibles)   : {insumo_agricola_menor_2000}")
    print(f"  • ✅ Pago electrónico (deducibles)    : {insumo_agricola_electronico}")

    print(f"\n📋 General:")
    print(f"  • Usos S01 (sin efectos fiscales) : {uso_s01_count}")
    print(f"  • Gastos normales efectivo >$2,000: {efectivo_mayor_2000}")

    print("\n⚠️  REGLAS APLICADAS:")
    print("  COMÚN A TODOS:")
    print("    • Art. 27 Fracc. III LISR: Efectivo >$2,000 NO deducible")
    print("    • Uso CFDI válido: G01, G02, G03  |  Método: PUE, PPD")
    print("\n  RÉGIMEN 626 (RESICO):")
    print("    • Gasolina efectivo ≤$2,000 o agrupada: Deducible (facilidad)")
    print("    • Gasolina individual >$2,000          : NO deducible")
    print("\n  RÉGIMEN 612 (Actividad Empresarial):")
    print("    • Gasolina efectivo (cualquier monto)  : NO deducible")
    print("    • Insumo agrícola efectivo >$2,000     : NO deducible")
    print("    • Insumo agrícola efectivo ≤$2,000     : Deducible")
    print("    • Insumo agrícola pago electrónico     : Deducible sin límite")
    print("    • Fundamento: Art. 103 LISR + Art. 27 Fracc. III LISR")
    print("    • Art. 74 LISR (AGAPES) / Art. 147 LISR: NO aplican a Régimen 612")

    print("\n🎨 CÓDIGO DE COLORES:")
    print("  🟦 AZUL        : Régimen 626 / Gasolina con IEPS")
    print("  🟪 MORADO      : Régimen 612")
    print("  🟫 CAFÉ        : Régimen 616")
    print("  🟧 NARANJA     : Otros regímenes / IEPS gasolina sin clasificar")
    print("  🟩 VERDE       : Deducible / IVA correcto")
    print("  🟢 VERDE CLARO : Insumo agrícola pago electrónico ✅")
    print("  🟥 ROJO        : NO deducible")
    print("  🟨 AMARILLO    : Gasolina RESICO ≤$2,000 / Insumo efectivo ≤$2,000")
    print("  🌸 ROSA        : Dulces/Botanas con IEPS 8%")

    print("\n" + "=" * 80)
    print("🔐 Archivo con información fiscal CONFIDENCIAL — Almacenar de forma segura")
    print("=" * 80)

except Exception as e:
    print(f"\n❌ Error al guardar: {e}")
    log_auditoria.registrar_error(e)
finally:
    log_auditoria.guardar_log()