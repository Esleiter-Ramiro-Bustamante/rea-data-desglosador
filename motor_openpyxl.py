"""
motor_openpyxl.py — Motor para computadoras básicas
ReaDesF1.8  (sin cambios funcionales respecto a v1.7 — fórmulas auditables ya correctas)
"""

import zipfile
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from validaciones_fiscales import (
    detectar_tipo, es_gasolina_agrupada, extraer_codigo,
    evaluar_deducibilidad, formulas_auditables,
    USOS_DEDUCIBLES, USO_CFDI_VERDE, REGIMENES_TRABAJADOS,
    LIMITE_EFECTIVO, FORMAS_ELECTRONICAS
)

blue_fill   = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
green_fill  = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
red_fill    = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
purple_fill = PatternFill(start_color='800080', end_color='800080', fill_type='solid')
orange_fill = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
pink_fill   = PatternFill(start_color='FF69B4', end_color='FF69B4', fill_type='solid')
brown_fill  = PatternFill(start_color='8B4513', end_color='8B4513', fill_type='solid')
lime_fill   = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')

bold_white   = Font(bold=True, color='FFFFFF')
bold_black   = Font(bold=True)
bold_result  = Font(bold=True)
center_align = Alignment(horizontal='center')
REG_FILLS    = {'626': blue_fill, '612': purple_fill, '616': brown_fill}


def _parchear_cache_formulas(xlsx_path: str, cache: dict) -> None:
    """
    Inserta valores cacheados en celdas con fórmulas del xlsx vía XML.
    cache = { 'AG5': 1234.56, 'AH5': 0.0, ... }

    Antes: <c r="AG5"><f>E5-F5</f><v></v></c>   ← openpyxl data_only lee None
    Después:<c r="AG5"><f>E5-F5</f><v>1234.56</v></c> ← lee 1234.56 sin Excel

    Esta es la única forma correcta de tener fórmula viva + valor cacheado
    en openpyxl. c._value = num sobreescribe la fórmula con el número.
    """
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        contents = {n: zin.read(n) for n in zin.namelist()}

    sheet_key = 'xl/worksheets/sheet1.xml'
    xml = contents[sheet_key].decode('utf-8')

    def replacer(m):
        ref = m.group(1)
        if ref not in cache:
            return m.group(0)
        val = cache[ref]
        if isinstance(val, float) and val == int(val):
            sv = str(int(val))
        else:
            sv = f'{val:.10f}'.rstrip('0').rstrip('.')
        return m.group(0).replace('<v></v>', f'<v>{sv}</v>')

    xml_patched = re.sub(
        r'<c r="([A-Z]+\d+)"[^>]*><f>[^<]*</f><v></v></c>',
        replacer, xml
    )
    contents[sheet_key] = xml_patched.encode('utf-8')

    with zipfile.ZipFile(xlsx_path, 'w', zipfile.ZIP_DEFLATED) as zout:
        for name, data in contents.items():
            zout.writestr(name, data)


def procesar_con_openpyxl(file_path: str, output_path: str,
                           modo: str = 'SEGURO') -> dict:

    print(f"  🔧 Motor: openpyxl ({modo})")
    wb    = openpyxl.load_workbook(file_path)
    sheet = wb.active
    wb.calculation.fullCalcOnLoad = False
    total_filas = sheet.max_row - 1
    print(f"  📊 Filas: {total_filas:,}")

    # headers_map
    headers_map = {
        str(c.value).strip().lower(): c.column
        for c in sheet[1] if c.value
    }

    def fc(n): return headers_map.get(n.strip().lower())
    def ec(n, fill=None):
        col = fc(n)
        if col is None:
            col = sheet.max_column + 1
            c   = sheet.cell(1, col, value=n)
            if fill: c.fill = fill
            c.alignment = center_align
            headers_map[n.strip().lower()] = col
        return col

    COL = {k: ec(v) for k, v in {
        'concepto':'Conceptos','total':'Total','uso':'Uso CFDI',
        'metodo':'Metodo pago','forma':'Forma pago',
        'regimen':'Regimen receptor','subtotal':'SubTotal',
        'descuento':'Descuento','iva0':'IVA Trasladado 0%',
        'iva_ex':'IVA Exento','iva16':'IVA Trasladado 16%',
        'razon_em':'Razon emisor',
    }.items()}

    ieps_8_col  = fc('IEPS Trasladado 8%')
    ieps_g_col  = fc('IEPS Trasladado')
    ieps_nd_col = fc('IEPS Trasladado No Desglosado')
    efecto_col  = fc('Efecto')
    razon_col   = ec('Razón No Deducible', red_fill)

    # Inicializar numéricos
    for row in range(2, sheet.max_row + 1):
        for ck in ['iva16','iva0','iva_ex','descuento']:
            c = sheet.cell(row, COL[ck])
            if c.value is None: c.value = 0
            c.number_format = "0.00"

    # Columnas de cálculo
    lc = sheet.max_column
    for i, h in enumerate(['SUB1-16%','SUB0%','SUB2-16%',
                            'IVA ACREDITABLE 16%','C IVA','T2',
                            'Comprobación T2','Deducible'], 1):
        c = sheet.cell(1, lc+i, value=h)
        c.fill = blue_fill
        c.alignment = center_align
        headers_map[h.lower()] = lc+i

    sub1_c,sub0_c,sub2_c = lc+1,lc+2,lc+3
    iva_ac,c_iva,t2_c    = lc+4,lc+5,lc+6
    comp_c,ded_c         = lc+7,lc+8

    # Cache letras de columna
    CL = {k: get_column_letter(v) for k,v in {
        'ST':COL['subtotal'],'DC':COL['descuento'],
        'I16':COL['iva16'],'I0':COL['iva0'],'IE':COL['iva_ex'],
        'TOT':COL['total'],'S1':sub1_c,'S0':sub0_c,
        'S2':sub2_c,'IA':iva_ac,'T2':t2_c,
    }.items()}
    if ieps_8_col:  CL['I8']  = get_column_letter(ieps_8_col)
    if ieps_g_col:  CL['IG']  = get_column_letter(ieps_g_col)
    if ieps_nd_col: CL['IND'] = get_column_letter(ieps_nd_col)

    stats = {k:0 for k in ['dulces_ieps8','dulces_sin','gas_ieps','gas_sin',
             'gas_626','gas_626_agrup','gas_612','gas_elec',
             'ins_nd','ins_menor','ins_elec','s01','ef_mayor']}
    stats['regimenes'] = {}

    # Cache fórmulas: { 'AG5': 1234.56 } — se parcha en XML al final
    _cache_formulas = {}

    # wc() definida UNA VEZ fuera del loop
    def wc(col, formula, num_val):
        c = sheet.cell(rn, col, formula)
        c.number_format = "0.00"
        c.font = bold_result
        # Acumular en cache para parche XML posterior
        _cache_formulas[f'{get_column_letter(col)}{rn}'] = num_val

    print("  🔄 Procesando...")

    for idx, row_cells in enumerate(
            sheet.iter_rows(min_row=2, max_row=sheet.max_row), 1):

        if idx % 1000 == 0 or idx == total_filas:
            print(f"    📊 {idx:,}/{total_filas:,} ({idx/total_filas*100:.0f}%)")

        rn = row_cells[0].row

        # Variables locales UNA VEZ por fila
        concepto_lower = str(row_cells[COL['concepto']-1].value or '').lower()
        total    = float(row_cells[COL['total']    -1].value or 0)
        st_v     = float(row_cells[COL['subtotal'] -1].value or 0)
        dc_v     = float(row_cells[COL['descuento']-1].value or 0)
        iva0_v   = float(row_cells[COL['iva0']     -1].value or 0)
        iva_ex_v = float(row_cells[COL['iva_ex']   -1].value or 0)
        iva16_v  = float(row_cells[COL['iva16']    -1].value or 0)
        regimen  = extraer_codigo(row_cells[COL['regimen']-1].value)
        uso_cfdi = extraer_codigo(row_cells[COL['uso']    -1].value)
        metodo   = extraer_codigo(row_cells[COL['metodo'] -1].value)
        forma    = extraer_codigo(row_cells[COL['forma']  -1].value)

        stats['regimenes'][regimen] = stats['regimenes'].get(regimen,0) + 1

        es_gas, es_dulce, es_insumo = detectar_tipo(concepto_lower)

        ieps_8_v  = float(row_cells[ieps_8_col -1].value or 0) if ieps_8_col  else 0.0
        ieps_g_v  = float(row_cells[ieps_g_col -1].value or 0) if ieps_g_col  else 0.0
        ieps_nd_v = float(row_cells[ieps_nd_col-1].value or 0) if ieps_nd_col else 0.0

        # SUB1 según IEPS
        if ieps_8_v > 0:
            f_sub1 = formulas_auditables(rn, CL)['sub1_ieps8']
            if es_dulce:
                sheet.cell(rn, COL['concepto']).fill = pink_fill
                stats['dulces_ieps8'] += 1
        elif es_gas and (ieps_g_v > 0 or ieps_nd_v > 0):
            ig = ieps_g_v if ieps_g_v > 0 else ieps_nd_v
            c  = sheet.cell(rn, COL['iva0'])
            c.value = ig
            c.fill  = orange_fill
            f_sub1  = formulas_auditables(rn, CL)['sub1']
            sheet.cell(rn, COL['concepto']).fill = blue_fill
            stats['gas_ieps'] += 1
        else:
            f_sub1 = formulas_auditables(rn, CL)['sub1']
            if es_gas:
                sheet.cell(rn, COL['concepto']).fill = orange_fill
                stats['gas_sin'] += 1
            elif es_dulce:
                stats['dulces_sin'] += 1

        # ── FÓRMULAS AUDITABLES — siempre escritas como fórmulas Excel
        # Al pararse en la celda se ve la operación completa.
        # sub1, sub0, sub2, iva_acred son la segunda validación contable.
        fa = formulas_auditables(rn, CL)

        # sub0: gasolina con IEPS → solo iva0 (el exento ya está incluido, no duplicar)
        f_sub0 = fa['sub0_gas'] if (es_gas and (ieps_g_v > 0 or ieps_nd_v > 0)) else fa['sub0']

        # ── Valores numéricos cacheados ────────────────────────────────
        # Se escriben junto con la fórmula para que openpyxl pueda
        # leerlos sin depender de que Excel evalúe las fórmulas.
        ieps_activo = ieps_nd_v if ieps_nd_v > 0 else ieps_g_v
        v_sub1 = round(st_v - dc_v + (ieps_8_v if ieps_8_v > 0 else 0), 2)
        if es_gas and (ieps_g_v > 0 or ieps_nd_v > 0):
            v_sub1 = round(st_v - dc_v, 2)
        v_sub0 = round(iva0_v, 2) if (ieps_activo > 0 and abs(iva0_v - iva_ex_v) < 0.01) \
                 else round(iva0_v + iva_ex_v, 2)
        v_sub2      = round(max(v_sub1 - v_sub0, 0), 2)
        v_iva_acred = round(v_sub2 * 0.16, 2)
        v_c_iva     = round(v_iva_acred - iva16_v, 2)
        v_t2        = round(v_sub2 + v_sub0 + iva16_v, 2)
        v_comprob   = round(total - v_t2, 2)

        wc(sub1_c, f_sub1,      v_sub1)
        wc(sub0_c, f_sub0,      v_sub0)
        wc(sub2_c, fa['sub2'],  v_sub2)
        wc(iva_ac,  fa['iva_acred'], v_iva_acred)
        wc(c_iva,   fa['c_iva'],     v_c_iva)
        wc(t2_c,    fa['t2'],        v_t2)
        wc(comp_c,  fa['comprob'],   v_comprob)

        # Validación visual IVA
        sub1_calc = st_v - dc_v + (ieps_8_v if ieps_8_v > 0 else 0)
        iva_calc  = round((sub1_calc - (iva0_v + iva_ex_v)) * 0.16, 2)
        if abs(iva_calc - iva16_v) < 0.01:
            sheet.cell(rn, iva_ac).fill        = green_fill
            sheet.cell(rn, COL['iva16']).fill  = green_fill

        # Color régimen
        sheet.cell(rn, COL['regimen']).fill = REG_FILLS.get(regimen, orange_fill)
        if regimen not in REG_FILLS:
            sheet.cell(rn, COL['razon_em']).fill = orange_fill

        uso_raw = row_cells[COL['uso']-1].value
        if uso_raw and str(uso_raw).strip() == USO_CFDI_VERDE:
            sheet.cell(rn, COL['uso']).fill = green_fill
        if uso_cfdi == 'S01':
            sheet.cell(rn, COL['uso']).fill = red_fill
            stats['s01'] += 1

        es_egreso = False
        if efecto_col:
            ev = row_cells[efecto_col-1].value
            if ev and str(ev).strip().upper() in {'EGRESO','E'}:
                es_egreso = True

        # Deducibilidad
        es_ded, razones = evaluar_deducibilidad(
            uso_cfdi, metodo, forma, regimen, total,
            es_gas, es_insumo, concepto_lower)

        # Contadores específicos
        if es_gas and forma == '01':
            if regimen == '626':
                stats['gas_626'] += 1
                if es_gasolina_agrupada(concepto_lower): stats['gas_626_agrup'] += 1
                fill = yellow_fill if (total <= LIMITE_EFECTIVO or es_gasolina_agrupada(concepto_lower)) else red_fill
                sheet.cell(rn, COL['forma']).fill = fill
            else:
                stats['gas_612'] += 1
                sheet.cell(rn, COL['forma']).fill = red_fill
        elif es_gas:
            stats['gas_elec'] += 1

        if es_insumo:
            if forma == '01' and total > LIMITE_EFECTIVO:
                stats['ins_nd'] += 1
                sheet.cell(rn, COL['forma']).fill    = red_fill
                sheet.cell(rn, COL['concepto']).fill = red_fill
            elif forma == '01':
                stats['ins_menor'] += 1
                sheet.cell(rn, COL['forma']).fill    = yellow_fill
                sheet.cell(rn, COL['concepto']).fill = yellow_fill
            else:
                stats['ins_elec'] += 1
                sheet.cell(rn, COL['concepto']).fill = lime_fill

        if not es_ded and forma == '01' and total > LIMITE_EFECTIVO and not es_gas and not es_insumo:
            stats['ef_mayor'] += 1
            if efecto_col: sheet.cell(rn, efecto_col).fill = red_fill

        dc = sheet.cell(rn, ded_c, value="SI" if es_ded else "NO")
        dc.fill      = (blue_fill if es_egreso else green_fill) if es_ded else red_fill
        dc.font      = bold_white
        dc.alignment = center_align

        rc = sheet.cell(rn, razon_col)
        if razones:
            rc.value = " | ".join(razones)
            rc.fill  = red_fill   if not es_ded else yellow_fill
            rc.font  = bold_white if not es_ded else bold_black
        else:
            rc.value = "Cumple requisitos"
            rc.fill  = green_fill
            rc.font  = Font(color='006100')

    # Anchos fuera del loop
    for col in range(lc+1, lc+9):
        sheet.column_dimensions[get_column_letter(col)].width = 15
    sheet.column_dimensions[get_column_letter(razon_col)].width = 65

    wb.save(output_path)
    print(f"  🔧 Insertando valores cacheados en fórmulas ({len(_cache_formulas):,} celdas)...")
    _parchear_cache_formulas(output_path, _cache_formulas)
    print(f"  ✅ Guardado: {output_path}")
    return stats
